from flask import Flask, request, jsonify, session, send_from_directory
from flask_cors import CORS
import os
import random
import string
import time
import subprocess
import json
import sys
from datetime import datetime, timedelta
from dotenv import load_dotenv
from flask_mail import Mail, Message
from models import db, User, Property, PropertyAmenity, PropertyImage, AgentProfile, Region, AgentRegion, PropertyView, BusinessInquiry, PricePrediction, FAQEntry, FAQSection, ContentSection, Bookmark, TeamSection, TeamMember, LegalContent, SubscriptionPlan, SubscriptionPlanFeature, ImportantFeature, FeaturesSection, FeaturesStep, UserProfile
from sqlalchemy import text
import jwt
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps

# Import ML review filter
try:
    from ml_review_filter import get_ml_filter
    ML_FILTER_AVAILABLE = True
except ImportError as e:
    print(f"Warning: ML review filter not available: {e}")
    ML_FILTER_AVAILABLE = False

# Load environment variables
load_dotenv()

app = Flask(__name__)
# CORS configuration to allow frontend origins and Authorization header
CORS(
    app,
    resources={r"/api/*": {"origins": "*"}},
    supports_credentials=True,
    allow_headers=["Content-Type", "Authorization"],
    methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
)

# JWT Configuration
JWT_SECRET_KEY = os.getenv('JWT_SECRET_KEY', 'your-super-secret-jwt-key-change-in-production')
JWT_ACCESS_TOKEN_EXPIRES = timedelta(hours=24)  # 24 hours

# Database Helper Functions
def validate_subscription_status(status):
    """Validate subscription status against database constraints"""
    valid_statuses = ['active', 'inactive', 'cancelled']
    return status.lower() in valid_statuses

def validate_user_type(user_type):
    """Validate user type against database constraints"""
    valid_types = ['free', 'premium', 'agent', 'admin']
    return user_type.lower() in valid_types
#test
# JWT Helper Functions
def create_access_token(user_id, email, user_type):
    """Create JWT access token"""
    payload = {
        'user_id': user_id,
        'email': email,
        'user_type': user_type,
        'exp': datetime.utcnow() + JWT_ACCESS_TOKEN_EXPIRES
    }
    return jwt.encode(payload, JWT_SECRET_KEY, algorithm='HS256')

def verify_token(token):
    """Verify JWT token and return user data"""
    try:
        payload = jwt.decode(token, JWT_SECRET_KEY, algorithms=['HS256'])
        return payload
    except jwt.ExpiredSignatureError:
        return None
    except jwt.InvalidTokenError:
        return None

# Authentication Middleware
def require_auth(f):
    """Decorator to require authentication"""
    def decorated_function(*args, **kwargs):
        print(f"🔍 Auth check for {f.__name__}")
        token = request.headers.get('Authorization')
        print(f"🔍 Token received: {token[:20] + '...' if token else 'None'}")
        
        if not token:
            print("❌ No token provided")
            return jsonify({'error': 'No token provided'}), 401
        
        # Remove 'Bearer ' prefix if present
        if token.startswith('Bearer '):
            token = token[7:]
        
        payload = verify_token(token)
        if not payload:
            return jsonify({'error': 'Invalid or expired token'}), 401
        
        # Add user info to request
        request.user = payload
        return f(*args, **kwargs)
    
    decorated_function.__name__ = f.__name__
    return decorated_function

# Database configuration
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL', 'postgresql://postgres:password@localhost:5432/fyp_app')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'your-secret-key-here')

# Initialize database with app
db.init_app(app)

# Gmail SMTP Configuration
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = os.getenv('MAIL_USERNAME', 'bearmeh00@gmail.com')
app.config['MAIL_PASSWORD'] = os.getenv('MAIL_PASSWORD', 'tuei rmyg tfqq szai')

# Initialize Flask-Mail
mail = Mail(app)

# Basic route to test if server is running
@app.route('/')
def home():
    return jsonify({'message': 'FYP App Backend is running!'})

# Test endpoint to check properties
@app.route('/api/test/properties')
def test_properties():
    """Test endpoint to check properties in database"""
    try:
        properties = Property.query.limit(10).all()
        properties_data = []
        for prop in properties:
            properties_data.append({
                'id': prop.id,
                'title': prop.title,
                'agent_id': prop.agent_id,
                'address': prop.address,
                'property_type': prop.property_type,
                'status': prop.status
            })
        return jsonify({'properties': properties_data, 'count': len(properties_data)})
    except Exception as e:
        print(f"Error in test endpoint: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# Test endpoint to check specific property
@app.route('/api/test/property/<int:property_id>')
def test_property(property_id):
    """Test endpoint to check a specific property"""
    try:
        property = Property.query.get(property_id)
        if not property:
            return jsonify({'error': f'Property {property_id} not found'}), 404
        
        return jsonify({
            'id': property.id,
            'title': property.title,
            'agent_id': property.agent_id,
            'address': property.address,
            'property_type': property.property_type,
            'status': property.status
        })
    except Exception as e:
        print(f"Error in test property endpoint: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# Health check route
@app.route('/health')
def health():
    return jsonify({'status': 'healthy', 'message': 'Server is running'})

# Database check route
@app.route('/api/db/check')
def check_database():
    """Check what tables exist in the database"""
    try:
        # Get list of tables
        query = text("""
            SELECT table_name 
            FROM information_schema.tables 
            WHERE table_schema = 'public'
            ORDER BY table_name
        """)
        result = db.session.execute(query)
        tables = [row[0] for row in result]
        
        return jsonify({
            'status': 'success',
            'tables': tables,
            'message': f'Found {len(tables)} tables in database'
        }), 200
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'error': str(e),
            'message': 'Failed to check database'
        }), 500

# Create missing tables route
@app.route('/api/db/setup', methods=['POST'])
def setup_database():
    """Create missing tables if they don't exist"""
    try:
        # Create user_reviews table if it doesn't exist
        create_reviews_table = text("""
            CREATE TABLE IF NOT EXISTS user_reviews (
                id SERIAL PRIMARY KEY,
                user_id INTEGER NOT NULL,
                review_type VARCHAR(20) CHECK (review_type IN ('platform', 'property', 'agent')) NOT NULL,
                rating INTEGER CHECK (rating >= 1 AND rating <= 5),
                review_text TEXT,
                review_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                is_verified BOOLEAN DEFAULT FALSE,
                FOREIGN KEY (user_id) REFERENCES users(id)
            )
        """)
        
        # Create review_interactions table if it doesn't exist
        create_interactions_table = text("""
            CREATE TABLE IF NOT EXISTS review_interactions (
                id SERIAL PRIMARY KEY,
                review_id INTEGER NOT NULL,
                user_id INTEGER NOT NULL,
                interaction_type VARCHAR(20) CHECK (interaction_type IN ('like', 'dislike')) NOT NULL,
                interaction_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE (review_id, user_id),
                FOREIGN KEY (review_id) REFERENCES user_reviews(id),
                FOREIGN KEY (user_id) REFERENCES users(id)
            )
        """)
        
        db.session.execute(create_reviews_table)
        db.session.execute(create_interactions_table)
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'message': 'Database tables created successfully'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'status': 'error',
            'error': str(e),
            'message': 'Failed to create database tables'
        }), 500

# Note: Sample data routes removed - system now uses real database data

# Test feedback submission without authentication (for debugging)
@app.route('/api/feedback/test-submit', methods=['POST'])
def test_submit_feedback():
    """Test feedback submission without authentication"""
    try:
        data = request.get_json()
        
        if not data or not data.get('rating') or not data.get('review_text'):
            return jsonify({'error': 'Rating and review text are required'}), 400
        
        # Get the first user from the database for testing
        user = User.query.first()
        if not user:
            return jsonify({'error': 'No users found in database'}), 400
        
        # Create new review
        new_review = {
            'user_id': user.id,
            'review_type': data.get('review_type', 'platform'),
            'rating': data['rating'],
            'review_text': data['review_text'],
            'review_date': datetime.utcnow(),
            'is_verified': False
        }
        
        # Insert into database
        query = text("""
            INSERT INTO user_reviews (user_id, review_type, rating, review_text, review_date, is_verified)
            VALUES (:user_id, :review_type, :rating, :review_text, :review_date, :is_verified)
            RETURNING id
        """)
        
        result = db.session.execute(query, new_review)
        review_id = result.fetchone()[0]
        db.session.commit()
        
        return jsonify({
            'message': 'Test feedback submitted successfully!',
            'review_id': review_id
        }), 201
        
    except Exception as e:
        db.session.rollback()
        print(f"Error submitting test feedback: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to submit test feedback: {str(e)}'}), 500

# Example API route
@app.route('/api/test')
def test_api():
    return jsonify({'message': 'API is working!'})

# User registration
@app.route('/api/auth/register', methods=['POST'])
def register():
    data = request.get_json()
    
    if not data or not data.get('email') or not data.get('password') or not data.get('full_name'):
        return jsonify({'error': 'Missing required fields'}), 400
    
    # Validate password length
    password = data.get('password', '')
    if len(password) < 8:
        return jsonify({'error': 'Password must be at least 8 characters long'}), 400
    
    # Check if user already exists
    if User.query.filter_by(email=data['email']).first():
        return jsonify({'error': 'User already exists'}), 409
    
    # Create new user
    user = User(
        email=data['email'],
        full_name=data['full_name'],
        phone_number=data.get('phone_number'),
        user_type=data.get('user_type', 'free')
    )
    user.set_password(password)
    
    try:
        db.session.add(user)
        db.session.commit()
        
        # Create agent profile if user is an agent
        if user.user_type == 'agent':
            agent_profile = AgentProfile(
                user_id=user.id,
                first_time_agent=True  # Default to True for new agents
            )
            db.session.add(agent_profile)
            db.session.commit()
        
        return jsonify({
            'message': 'User created successfully',
            'user_id': user.id,
            'email': user.email,
            'full_name': user.full_name
        }), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'Failed to create user'}), 500

# User login
@app.route('/api/auth/login', methods=['POST'])
def login():
    data = request.get_json()
    
    if not data or not data.get('email') or not data.get('password'):
        return jsonify({'error': 'Missing email or password'}), 400
    
    user = User.query.filter_by(email=data['email']).first()
    
    if user and user.check_password(data['password']):
        # Update last login
        user.last_login = datetime.utcnow()
        db.session.commit()
        
        # Create JWT token
        access_token = create_access_token(user.id, user.email, user.user_type)
        
        return jsonify({
            'message': 'Login successful',
            'access_token': access_token,
            'user': {
                'user_id': user.id,
                'email': user.email,
                'full_name': user.full_name,
                'user_type': user.user_type
            }
        }), 200
    else:
        return jsonify({'error': 'Invalid credentials'}), 401

# Email verification storage (in production, use Redis or database)
email_verification_codes = {}

# Password reset storage (in production, use Redis or database)
password_reset_codes = {}

# Generate OTP code
def generate_otp():
    return ''.join(random.choices(string.digits, k=6))

# Send OTP to email using Gmail SMTP
def send_otp_email(email, otp):
    try:
        msg = Message(
            'Your Valuez Verification Code',
            sender='bearmeh00@gmail.com',  # ← Now matches your MAIL_USERNAME
            recipients=[email]
        )
        msg.body = f'Your verification code is: {otp}'
        mail.send(msg)
        print(f"📧 Email sent successfully to {email}")
        return True
    except Exception as e:
        print(f"❌ Failed to send email: {e}")
        return False

# Request email verification
@app.route('/api/auth/send-otp', methods=['POST'])
def send_otp():
    data = request.get_json()
    
    if not data or not data.get('email'):
        return jsonify({'error': 'Email is required'}), 400
    
    email = data['email']
    
    # Check if user already exists
    if User.query.filter_by(email=email).first():
        return jsonify({'error': 'User already exists with this email'}), 409
    
    # Generate OTP
    otp = generate_otp()
    
    # Store OTP with expiration (5 minutes)
    email_verification_codes[email] = {
        'otp': otp,
        'expires_at': datetime.utcnow() + timedelta(minutes=5),
        'attempts': 0
    }
    
    # Send OTP (mock)
    if send_otp_email(email, otp):
        return jsonify({
            'message': 'OTP sent successfully',
            'email': email
        }), 200
    else:
        return jsonify({'error': 'Failed to send OTP'}), 500

# Verify OTP
@app.route('/api/auth/verify-otp', methods=['POST'])
def verify_otp():
    data = request.get_json()
    
    if not data or not data.get('email') or not data.get('otp'):
        return jsonify({'error': 'Email and OTP are required'}), 400
    
    email = data['email']
    otp = data['otp']
    
    # Check if OTP exists and is valid
    if email not in email_verification_codes:
        return jsonify({'error': 'No OTP requested for this email'}), 400
    
    verification_data = email_verification_codes[email]
    
    # Check if OTP is expired
    if datetime.utcnow() > verification_data['expires_at']:
        del email_verification_codes[email]
        return jsonify({'error': 'OTP has expired'}), 400
    
    # Check if too many attempts
    if verification_data['attempts'] >= 3:
        del email_verification_codes[email]
        return jsonify({'error': 'Too many failed attempts. Please request a new OTP'}), 400
    
    # Check if OTP matches
    if verification_data['otp'] != otp:
        verification_data['attempts'] += 1
        return jsonify({'error': 'Invalid OTP'}), 400
    
    # OTP is valid - mark email as verified
    verification_data['verified'] = True
    verification_data['verified_at'] = datetime.utcnow()
    
    return jsonify({
        'message': 'Email verified successfully',
        'email': email
    }), 200

# Get all properties
@app.route('/api/properties', methods=['GET'])
def get_properties():
    try:
        # Get query parameters for pagination
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 8, type=int)
        
        # Query active properties with pagination
        query = Property.query.filter_by(status='active')
        
        # Pagination
        pagination = query.order_by(Property.created_at.desc()).paginate(
            page=page, per_page=per_page, error_out=False
        )
        
        properties = pagination.items
        property_list = []
        
        for prop in properties:
            # Get basic agent information with profile
            agent = User.query.get(prop.agent_id)
            agent_name = agent.full_name if agent else 'Property Agent'
            
            # Get agent profile for additional info
            agent_profile = None
            if agent:
                agent_profile = AgentProfile.query.filter_by(user_id=agent.id).first()
            
            property_data = {
                'id': prop.id,
                'title': prop.title,
                'description': prop.description,
                'property_type': prop.property_type,
                'address': prop.address,
                'city': prop.city,
                'state': prop.state,
                'zip_code': prop.zip_code,
                'size_sqft': float(prop.size_sqft),
                'asking_price': float(prop.asking_price),
                'price_type': prop.price_type,
                'status': prop.status,
                'latitude': float(prop.latitude) if prop.latitude else None,
                'longitude': float(prop.longitude) if prop.longitude else None,
                'created_at': prop.created_at.isoformat() if prop.created_at else None,
                'agent_name': agent_name,
                'agent_license': agent_profile.license_number if agent_profile else None,
                'agent_company': agent_profile.company_name if agent_profile else 'Valuez Real Estate'
            }
            property_list.append(property_data)
        
        return jsonify({
            'properties': property_list,
            'total': pagination.total,
            'page': page,
            'per_page': per_page,
            'pages': pagination.pages
        }), 200
    except Exception as e:
        print(f"Error fetching properties: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Failed to fetch properties'}), 500

# Filter properties by nearby amenities
@app.route('/api/properties/filter-by-amenities', methods=['POST'])
def filter_properties_by_amenities():
    """
    Filter properties based on proximity to selected amenities.
    Uses Google Maps Places API to find nearby amenities and returns property IDs.
    """
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Request body is required'}), 400
        
        amenity_types = data.get('amenity_types', [])
        radius = data.get('radius', 1000)  # Default 1km in meters
        
        if not amenity_types or len(amenity_types) == 0:
            # If no amenities selected, return all active property IDs
            all_properties = Property.query.filter_by(status='active').all()
            return jsonify({
                'property_ids': [prop.id for prop in all_properties],
                'count': len(all_properties)
            }), 200
        
        # Map amenity types to Google Places API types
        places_type_mapping = {
            'school': 'school',
            'hospital': 'hospital',
            'supermarket': 'supermarket',
            'restaurant': 'restaurant',
            'cafe': 'cafe',
            'bank': 'bank',
            'atm': 'atm',
            'shopping_mall': 'shopping_mall',
            'bus_station': 'bus_station',
            'subway_station': 'subway_station',
            'gym': 'gym',
            'pharmacy': 'pharmacy',
            'park': 'park'
        }
        
        # Get all active properties with coordinates
        properties = Property.query.filter_by(status='active').filter(
            Property.latitude.isnot(None),
            Property.longitude.isnot(None)
        ).all()
        
        if not properties:
            return jsonify({
                'property_ids': [],
                'count': 0,
                'message': 'No properties with coordinates found'
            }), 200
        
        # For each amenity type, find properties that have nearby amenities
        # Since we don't have Google Maps API in backend, we'll use a simpler approach:
        # For now, return properties that have latitude/longitude (real implementation would use Places API)
        matching_property_ids = set()
        
        try:
            import requests
            GOOGLE_MAPS_API_KEY = os.getenv('GOOGLE_MAPS_API_KEY')
            api_key_configured = bool(GOOGLE_MAPS_API_KEY)  # Set based on whether key exists
            
            if GOOGLE_MAPS_API_KEY:
                # Use Google Places API to find nearby amenities
                for prop in properties:
                    lat = float(prop.latitude)
                    lng = float(prop.longitude)
                    
                    # Check each selected amenity type
                    has_nearby_amenity = False
                    for amenity_type in amenity_types:
                        places_type = places_type_mapping.get(amenity_type, amenity_type)
                        
                        # Query Google Places API for nearby places
                        places_url = 'https://maps.googleapis.com/maps/api/place/nearbysearch/json'
                        params = {
                            'location': f'{lat},{lng}',
                            'radius': radius,
                            'type': places_type,
                            'key': GOOGLE_MAPS_API_KEY
                        }
                        
                        try:
                            response = requests.get(places_url, params=params, timeout=5)
                            if response.status_code == 200:
                                places_data = response.json()
                                if places_data.get('status') == 'OK' and len(places_data.get('results', [])) > 0:
                                    has_nearby_amenity = True
                                    break  # Found at least one nearby amenity of the selected types
                        except Exception as e:
                            print(f"Error checking nearby places for property {prop.id}: {e}")
                            continue
                    
                    if has_nearby_amenity:
                        matching_property_ids.add(prop.id)
            else:
                # Fallback: If no API key, return properties with coordinates (for testing)
                # In production, this should return empty list
                # For now, return all properties with coordinates to allow testing
                matching_property_ids = {prop.id for prop in properties}
                print("Warning: GOOGLE_MAPS_API_KEY not found. Returning all properties with coordinates (for testing). Please set GOOGLE_MAPS_API_KEY for accurate filtering.")
        
        except ImportError:
            # If requests library is not available, return all properties with coordinates (for testing)
            matching_property_ids = {prop.id for prop in properties}
            api_key_configured = False
            print("Warning: requests library not available. Returning all properties with coordinates (for testing). Please install requests: pip install requests")
        
        return jsonify({
            'property_ids': list(matching_property_ids),
            'count': len(matching_property_ids),
            'api_key_configured': api_key_configured if 'api_key_configured' in locals() else bool(os.getenv('GOOGLE_MAPS_API_KEY'))
        }), 200
        
    except Exception as e:
        print(f"Error filtering properties by amenities: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to filter properties: {str(e)}'}), 500

# Get property by ID
@app.route('/api/properties/<int:property_id>', methods=['GET'])
def get_property(property_id):
    try:
        property = Property.query.get_or_404(property_id)
        
        # Track the property view (anonymous tracking)
        try:
            new_view = PropertyView(
                property_id=property_id,
                user_id=None,  # Anonymous view
                ip_address=request.remote_addr,
                user_agent=request.headers.get('User-Agent')
            )
            db.session.add(new_view)
            db.session.commit()
        except Exception as e:
            # Don't fail the main request if view tracking fails
            print(f"Error tracking property view (table might not exist yet): {e}")
            db.session.rollback()
        
        # Get agent information with profile
        agent = User.query.get(property.agent_id)
        agent_data = None
        if agent:
            # Get agent profile information
            agent_profile = AgentProfile.query.filter_by(user_id=agent.id).first()
            
            agent_data = {
                'id': agent.id,
                'full_name': agent.full_name,
                'email': agent.email,
                'phone_number': agent.phone_number,
                'profile_image_url': agent.profile_image_url,
                'license_number': agent_profile.license_number if agent_profile else None,
                'company_name': agent_profile.company_name if agent_profile else 'Valuez Real Estate',
                'years_experience': agent_profile.years_experience if agent_profile else None,
                'specializations': agent_profile.specializations if agent_profile else None,
                'bio': agent_profile.bio if agent_profile else None,
                'contact_preferences': agent_profile.contact_preferences if agent_profile else 'phone,email,whatsapp'
            }
        
        # Get property amenities
        amenities = []
        property_amenities = PropertyAmenity.query.filter_by(property_id=property_id).all()
        for amenity in property_amenities:
            if amenity.has_amenity:
                amenities.append(amenity.amenity_name)
        
        # Get property images
        images = []
        property_images = PropertyImage.query.filter_by(property_id=property_id).all()
        for image in property_images:
            images.append({
                'id': image.id,
                'url': image.image_url,
                'name': image.image_name,
                'is_primary': image.is_primary
            })
        
        property_data = {
            'id': property.id,
            'title': property.title,
            'description': property.description,
            'property_type': property.property_type,
            'address': property.address,
            'street_address': property.street_address,
            'city': property.city,
            'state': property.state,
            'zip_code': property.zip_code,
            'size_sqft': float(property.size_sqft),
            'floors': property.floors,
            'year_built': property.year_built,
            'zoning': property.zoning,
            'parking_spaces': property.parking_spaces,
            'asking_price': float(property.asking_price),
            'price_type': property.price_type,
            'status': property.status,
            'latitude': float(property.latitude) if property.latitude else None,
            'longitude': float(property.longitude) if property.longitude else None,
            'created_at': property.created_at.isoformat() if property.created_at else None,
            'updated_at': property.updated_at.isoformat() if property.updated_at else None,
            'agent': agent_data,
            'amenities': amenities,
            'images': images
        }
        
        print(f"Returning property data for ID {property_id}:")
        print(f"Coordinates: lat={property_data['latitude']}, lng={property_data['longitude']}")
        print(f"Address: {property_data['address']}")
        
        return jsonify(property_data), 200
    except Exception as e:
        print(f"Error fetching property {property_id}: {e}")
        return jsonify({'error': 'Property not found'}), 404

# Get nearby properties for comparison/prediction
@app.route('/api/properties/nearby', methods=['POST'])
def get_nearby_properties():
    try:
        data = request.get_json()
        if not data or not data.get('address'):
            return jsonify({'error': 'Address is required'}), 400
        
        target_address = data['address']
        limit = data.get('limit', 10)  # Default to 10 properties
        
        # Get properties with the same property type first, then fallback to others
        target_property_type = data.get('propertyType')
        
        if target_property_type:
            # Only get properties with exact property type match
            properties = Property.query.filter_by(
                status='active', 
                property_type=target_property_type
            ).limit(limit).all()
        else:
            # If no property type specified, get all active properties
            properties = Property.query.filter_by(status='active').limit(limit).all()
        
        # Enhanced proximity-based filtering using address components
        nearby_properties = []
        
        # Extract key location components from target address
        target_parts = target_address.lower().split()
        target_city = None
        target_district = None
        
        # Map address keywords to exact region_value from FirstTimerAgent regionsData
        singapore_districts = {
            'raffles': 'Raffles Place, Cecil, Marina, People\'s Park',
            'cecil': 'Raffles Place, Cecil, Marina, People\'s Park',
            'marina': 'Raffles Place, Cecil, Marina, People\'s Park',
            'people\'s park': 'Raffles Place, Cecil, Marina, People\'s Park',
            'anson': 'Anson, Tanjong Pagar',
            'tanjong pagar': 'Anson, Tanjong Pagar',
            'queenstown': 'Queenstown, Tiong Bahru',
            'tiong bahru': 'Queenstown, Tiong Bahru',
            'telok blangah': 'Telok Blangah, Harbourfront',
            'harbourfront': 'Telok Blangah, Harbourfront',
            'pasir panjang': 'Pasir Panjang, Hong Leong Garden, Clementi New Town',
            'clementi': 'Pasir Panjang, Hong Leong Garden, Clementi New Town',
            'high street': 'High Street, Beach Road (part)',
            'beach road': 'High Street, Beach Road (part)',
            'middle road': 'Middle Road, Golden Mile',
            'golden mile': 'Middle Road, Golden Mile',
            'little india': 'Little India',
            'orchard': 'Orchard, Cairnhill, River Valley',
            'cairnhill': 'Orchard, Cairnhill, River Valley',
            'river valley': 'Orchard, Cairnhill, River Valley',
            'ardmore': 'Ardmore, Bukit Timah, Holland Road, Tanglin',
            'bukit timah': 'Ardmore, Bukit Timah, Holland Road, Tanglin',
            'holland road': 'Ardmore, Bukit Timah, Holland Road, Tanglin',
            'tanglin': 'Ardmore, Bukit Timah, Holland Road, Tanglin',
            'watten estate': 'Watten Estate, Novena, Thomson',
            'novena': 'Watten Estate, Novena, Thomson',
            'thomson': 'Watten Estate, Novena, Thomson',
            'balestier': 'Balestier, Toa Payoh, Serangoon',
            'toa payoh': 'Balestier, Toa Payoh, Serangoon',
            'serangoon': 'Balestier, Toa Payoh, Serangoon',
            'macpherson': 'Macpherson, Braddell',
            'braddell': 'Macpherson, Braddell',
            'geylang': 'Geylang, Eunos',
            'eunos': 'Geylang, Eunos',
            'katong': 'Katong, Joo Chiat, Amber Road',
            'joo chiat': 'Katong, Joo Chiat, Amber Road',
            'bedok': 'Bedok, Upper East Coast, Eastwood, Kew Drive',
            'loyang': 'Loyang, Changi',
            'changi': 'Loyang, Changi',
            'tampines': 'Tampines, Pasir Ris',
            'pasir ris': 'Tampines, Pasir Ris',
            'hougang': 'Serangoon Garden, Hougang, Punggol',
            'punggol': 'Serangoon Garden, Hougang, Punggol',
            'bishan': 'Bishan, Ang Mo Kio',
            'ang mo kio': 'Bishan, Ang Mo Kio',
            'jurong': 'Jurong',
            'hillview': 'Hillview, Dairy Farm, Bukit Panjang, Choa Chu Kang',
            'choa chu kang': 'Hillview, Dairy Farm, Bukit Panjang, Choa Chu Kang',
            'lim chu kang': 'Lim Chu Kang, Tengah',
            'kranji': 'Kranji, Woodgrove',
            'woodgrove': 'Kranji, Woodgrove',
            'upper thomson': 'Upper Thomson, Springleaf',
            'springleaf': 'Upper Thomson, Springleaf',
            'yishun': 'Yishun, Sembawang',
            'sembawang': 'Yishun, Sembawang',
            'seletar': 'Seletar'
        }
        
        # Find the region based on address keywords
        target_region = None
        for keyword, region in singapore_districts.items():
            if keyword in target_address:
                target_region = region
                break
        
        if not target_region:
            target_region = 'General Singapore Areas'  # Default fallback
        
        for prop in properties:
            # Skip if it's the same address
            if prop.address.lower() == target_address.lower():
                continue
                
            # Get agent information
            agent = User.query.get(prop.agent_id)
            agent_name = agent.full_name if agent else 'Property Agent'
            
            # Get agent profile
            agent_profile = None
            if agent:
                agent_profile = AgentProfile.query.filter_by(user_id=agent.id).first()
            
            # Calculate enhanced similarity score
            similarity_score = 0
            
            # City match (highest weight)
            if prop.city and target_address.lower().find(prop.city.lower()) != -1:
                similarity_score += 5
            
            # District/area match (high weight)
            if target_district and prop.address.lower().find(target_district) != -1:
                similarity_score += 4
            
            # State match (medium weight)
            if prop.state and target_address.lower().find(prop.state.lower()) != -1:
                similarity_score += 3
            
            # Postal code proximity (if available)
            if prop.zip_code and target_address.lower().find(prop.zip_code) != -1:
                similarity_score += 2
            
            # Property type similarity (highest priority)
            if prop.property_type and data.get('propertyType'):
                if prop.property_type.lower() == data['propertyType'].lower():
                    similarity_score += 10  # Much higher weight for property type match
            
            # Include properties with meaningful similarity or as fallback
            if similarity_score >= 2 or (len(nearby_properties) < limit and similarity_score >= 0):
                property_data = {
                    'id': prop.id,
                    'address': prop.address,
                    'city': prop.city,
                    'state': prop.state,
                    'property_type': prop.property_type,
                    'size_sqft': float(prop.size_sqft) if prop.size_sqft else None,
                    'asking_price': float(prop.asking_price) if prop.asking_price else None,
                    'price_type': prop.price_type,
                    'status': prop.status,
                    'agent_name': agent_name,
                    'agent_company': agent_profile.company_name if agent_profile else 'Valuez Real Estate',
                    'similarity_score': similarity_score
                }
                nearby_properties.append(property_data)
        
        # Sort by similarity score (highest first) and limit results
        nearby_properties.sort(key=lambda x: x['similarity_score'], reverse=True)
        nearby_properties = nearby_properties[:limit]
        
        return jsonify({
            'nearby_properties': nearby_properties,
            'total_found': len(nearby_properties),
            'target_address': target_address
        }), 200
        
    except Exception as e:
        print(f"Error fetching nearby properties: {e}")
        return jsonify({'error': 'Failed to fetch nearby properties'}), 500

# Get agents assigned to a specific region based on address
@app.route('/api/agents/region', methods=['POST'])
def get_agents_by_region():
    try:
        data = request.get_json()
        if not data or not data.get('address'):
            return jsonify({'error': 'Address is required'}), 400
        
        target_address = data['address'].lower()
        property_type = data.get('property_type')  # Get property type from request
        
        # Determine region based on address (using the same logic as nearby properties)
        target_region = None
        
        # Map address keywords to exact region_value from FirstTimerAgent regionsData
        singapore_districts = {
            'raffles': 'Raffles Place, Cecil, Marina, People\'s Park',
            'cecil': 'Raffles Place, Cecil, Marina, People\'s Park',
            'marina': 'Raffles Place, Cecil, Marina, People\'s Park',
            'people\'s park': 'Raffles Place, Cecil, Marina, People\'s Park',
            'anson': 'Anson, Tanjong Pagar',
            'tanjong pagar': 'Anson, Tanjong Pagar',
            'queenstown': 'Queenstown, Tiong Bahru',
            'tiong bahru': 'Queenstown, Tiong Bahru',
            'telok blangah': 'Telok Blangah, Harbourfront',
            'harbourfront': 'Telok Blangah, Harbourfront',
            'pasir panjang': 'Pasir Panjang, Hong Leong Garden, Clementi New Town',
            'clementi': 'Pasir Panjang, Hong Leong Garden, Clementi New Town',
            'high street': 'High Street, Beach Road (part)',
            'beach road': 'High Street, Beach Road (part)',
            'middle road': 'Middle Road, Golden Mile',
            'golden mile': 'Middle Road, Golden Mile',
            'little india': 'Little India',
            'orchard': 'Orchard, Cairnhill, River Valley',
            'cairnhill': 'Orchard, Cairnhill, River Valley',
            'river valley': 'Orchard, Cairnhill, River Valley',
            'ardmore': 'Ardmore, Bukit Timah, Holland Road, Tanglin',
            'bukit timah': 'Ardmore, Bukit Timah, Holland Road, Tanglin',
            'holland road': 'Ardmore, Bukit Timah, Holland Road, Tanglin',
            'tanglin': 'Ardmore, Bukit Timah, Holland Road, Tanglin',
            'watten estate': 'Watten Estate, Novena, Thomson',
            'novena': 'Watten Estate, Novena, Thomson',
            'thomson': 'Watten Estate, Novena, Thomson',
            'balestier': 'Balestier, Toa Payoh, Serangoon',
            'toa payoh': 'Balestier, Toa Payoh, Serangoon',
            'serangoon': 'Balestier, Toa Payoh, Serangoon',
            'macpherson': 'Macpherson, Braddell',
            'braddell': 'Macpherson, Braddell',
            'geylang': 'Geylang, Eunos',
            'eunos': 'Geylang, Eunos',
            'katong': 'Katong, Joo Chiat, Amber Road',
            'joo chiat': 'Katong, Joo Chiat, Amber Road',
            'bedok': 'Bedok, Upper East Coast, Eastwood, Kew Drive',
            'loyang': 'Loyang, Changi',
            'changi': 'Loyang, Changi',
            'tampines': 'Tampines, Pasir Ris',
            'pasir ris': 'Tampines, Pasir Ris',
            'hougang': 'Serangoon Garden, Hougang, Punggol',
            'punggol': 'Serangoon Garden, Hougang, Punggol',
            'bishan': 'Bishan, Ang Mo Kio',
            'ang mo kio': 'Bishan, Ang Mo Kio',
            'jurong': 'Jurong',
            'hillview': 'Hillview, Dairy Farm, Bukit Panjang, Choa Chu Kang',
            'choa chu kang': 'Hillview, Dairy Farm, Bukit Panjang, Choa Chu Kang',
            'lim chu kang': 'Lim Chu Kang, Tengah',
            'kranji': 'Kranji, Woodgrove',
            'woodgrove': 'Kranji, Woodgrove',
            'upper thomson': 'Upper Thomson, Springleaf',
            'springleaf': 'Upper Thomson, Springleaf',
            'yishun': 'Yishun, Sembawang',
            'sembawang': 'Yishun, Sembawang',
            'seletar': 'Seletar'
        }
        
        # Find the region based on address keywords
        for keyword, region in singapore_districts.items():
            if keyword in target_address:
                target_region = region
                break
        
        if not target_region:
            target_region = 'General Singapore Areas'  # Default fallback
        
        # Get agents assigned to this region
        agents = []
        

        
        try:
            agent_regions = AgentRegion.query.filter_by(region_value=target_region).all()
            print(f"Found {len(agent_regions)} agent regions for region: {target_region}")
        except Exception as e:
            print(f"Error querying AgentRegion table: {e}")
            # Return empty agents list if table doesn't exist
            return jsonify({
                'agents': [],
                'region': target_region,
                'total_agents': 0,
                'note': 'Agent regions table not available'
            }), 200
        
        # Map property types to match specialization names
        # Property types from price prediction: Office, Retail, Shop House, Single-user Factory, Multiple-user Factory, Warehouse, Business Parks
        property_type_mapping = {
            'office': 'Office',
            'retail': 'Retail',
            'shop house': 'Shop House',
            'shophouse': 'Shop House',
            'single-user factory': 'Single-user Factory',
            'single-user-factory': 'Single-user Factory',
            'singleuserfactory': 'Single-user Factory',
            'multiple-user factory': 'Multiple-user Factory',
            'multiple-user-factory': 'Multiple-user Factory',
            'multipleuserfactory': 'Multiple-user Factory',
            'warehouse': 'Warehouse',
            'business parks': 'Business Parks',
            'businessparks': 'Business Parks',
            'business-park': 'Business Parks'
        }
        
        # Normalize property type for matching
        normalized_property_type = None
        if property_type:
            normalized_property_type = property_type_mapping.get(property_type.lower(), property_type)
        
        for agent_region in agent_regions:
            try:
                # Use agent_id instead of user_id based on your database structure
                user = User.query.get(agent_region.agent_id)
                if user and user.user_type == 'agent':
                    # Get agent profile
                    agent_profile = AgentProfile.query.filter_by(user_id=user.id).first()
                    
                    # Filter by specializations if property_type is provided
                    if normalized_property_type and agent_profile and agent_profile.specializations:
                        # Check if agent has this property type in their specializations
                        agent_specializations = agent_profile.specializations
                        # Handle both list and JSON string formats
                        if isinstance(agent_specializations, str):
                            try:
                                agent_specializations = json.loads(agent_specializations)
                            except:
                                agent_specializations = []
                        elif agent_specializations is None:
                            agent_specializations = []
                        
                        # Only include agent if they specialize in this property type
                        if normalized_property_type not in agent_specializations:
                            continue  # Skip this agent
                    
                    agent_data = {
                        'id': user.id,
                        'name': user.full_name,
                        'email': user.email,
                        'phone': user.phone_number,
                        'company': agent_profile.company_name if agent_profile else 'Valuez Real Estate',
                        'license': agent_profile.license_number if agent_profile else None,
                        'experience': agent_profile.years_experience if agent_profile else None,
                        'specializations': agent_profile.specializations if agent_profile else None,
                        'region': target_region
                    }
                    agents.append(agent_data)
            except Exception as e:
                print(f"Error processing agent region {agent_region.id}: {e}")
                continue
        
        return jsonify({
            'agents': agents,
            'region': target_region,
            'total_agents': len(agents)
        }), 200
        
    except Exception as e:
        print(f"Error fetching agents by region: {e}")
        return jsonify({'error': 'Failed to fetch agents by region'}), 500

# Get FAQ entries
@app.route('/api/faq', methods=['GET'])
def get_faq():
    try:
        faq_entries = FAQEntry.query.filter_by(is_active=True).order_by(FAQEntry.display_order).all()
        faq_list = []
        
        for entry in faq_entries:
            faq_data = {
                'id': entry.id,
                'question': entry.question,
                'answer': entry.answer,
                'category': entry.category,
                'display_order': entry.display_order
            }
            faq_list.append(faq_data)
        
        return jsonify({'faq_entries': faq_list}), 200
    except Exception as e:
        return jsonify({'error': 'Failed to fetch FAQ'}), 500

# Password recovery - Send OTP
@app.route('/api/auth/forgot-password', methods=['POST'])
def forgot_password():
    data = request.get_json()
    
    if not data or not data.get('email'):
        return jsonify({'error': 'Email is required'}), 400
    
    email = data['email']
    user = User.query.filter_by(email=email).first()
    
    if not user:
        # Don't reveal if user exists or not for security
        return jsonify({'message': 'If the email exists, a recovery code has been sent'}), 200
    
    # Generate OTP for password recovery
    otp = generate_otp()
    
    # Store OTP with expiration (5 minutes)
    password_reset_codes[email] = {
        'otp': otp,
        'expires_at': datetime.utcnow() + timedelta(minutes=5),
        'attempts': 0
    }
    
    try:
        # Send recovery email
        msg = Message(
            subject='Password Recovery - Valuez',
            sender='bearmeh00@gmail.com',
            recipients=[email]
        )
        msg.body = f'''Hello {user.full_name},

You have requested to reset your password for your Valuez account.

Your recovery code is: {otp}

This code will expire in 5 minutes.

If you did not request this password reset, please ignore this email.

Best regards,
Valuez Team'''
        
        mail.send(msg)
        
        return jsonify({'message': 'Recovery code sent successfully'}), 200
        
    except Exception as e:
        # Remove the stored OTP if email fails
        if email in password_reset_codes:
            del password_reset_codes[email]
        return jsonify({'error': 'Failed to send recovery email'}), 500

# Get user profile
@app.route('/api/auth/profile', methods=['GET'])
@require_auth
def get_profile():
    # Get user from authenticated token
    user_id = request.user['user_id']
    
    user = User.query.get(user_id)
    if not user:
        return jsonify({'error': 'User not found'}), 404
    
    return jsonify({
        'id': user.id,
        'email': user.email,
        'full_name': user.full_name,
        'phone_number': user.phone_number,
        'user_type': user.user_type,
        'subscription_status': user.subscription_status,
        'subscription_start_date': user.subscription_start_date.isoformat() if user.subscription_start_date else None,
        'account_created_date': user.account_created_date.isoformat() if user.account_created_date else None,
        'last_login': user.last_login.isoformat() if user.last_login else None,
        'is_active': user.is_active
    }), 200

# Get agent profile
@app.route('/api/auth/agent-profile', methods=['GET'])
@require_auth
def get_agent_profile():
    # Get user from authenticated token
    user_id = request.user['user_id']
    
    user = User.query.get(user_id)
    if not user:
        return jsonify({'error': 'User not found'}), 404
    
    # Check if user is an agent
    if user.user_type != 'agent':
        return jsonify({'error': 'Access denied. Agent profile only available for agent users.'}), 403
    
    # Get agent profile from agent_profiles table
    agent_profile = AgentProfile.query.filter_by(user_id=user_id).first()
    
    if not agent_profile:
        return jsonify({'error': 'Agent profile not found'}), 404
    
    return jsonify({
        'id': agent_profile.id,
        'user_id': agent_profile.user_id,
        'cea_number': agent_profile.license_number,
        'company_name': agent_profile.company_name,
        'company_phone': agent_profile.company_phone if hasattr(agent_profile, 'company_phone') else None,
        'company_email': agent_profile.company_email if hasattr(agent_profile, 'company_email') else None,
        'years_experience': agent_profile.years_experience,
        'specializations': agent_profile.specializations,
        'bio': agent_profile.bio,
        'contact_preferences': agent_profile.contact_preferences,
        'license_picture_url': agent_profile.license_picture_url if hasattr(agent_profile, 'license_picture_url') else None
    }), 200

# Upload agent license picture
@app.route('/api/auth/upload-license-picture', methods=['POST'])
@require_auth
def upload_license_picture():
    print("=== License Upload Debug ===")
    print(f"Request method: {request.method}")
    print(f"Request files: {list(request.files.keys())}")
    print(f"Request form: {list(request.form.keys())}")
    
    # Get user from authenticated token
    user_id = request.user['user_id']
    print(f"User ID: {user_id}")
    
    user = User.query.get(user_id)
    if not user:
        print("User not found")
        return jsonify({'error': 'User not found'}), 404
    
    print(f"User type: {user.user_type}")
    
    # Check if user is an agent
    if user.user_type != 'agent':
        print("User is not an agent")
        return jsonify({'error': 'Access denied. License upload only available for agent users.'}), 403
    
    # Check if file was uploaded
    if 'license_picture' not in request.files:
        print("No license_picture in request.files")
        return jsonify({'error': 'No license picture file provided'}), 400
    
    file = request.files['license_picture']
    print(f"File name: {file.filename}")
    print(f"File content type: {file.content_type}")
    
    # Check if file is empty
    if file.filename == '':
        print("File filename is empty")
        return jsonify({'error': 'No file selected'}), 400
    
    # Check file type
    allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'pdf'}
    if not file.filename.lower().endswith(tuple('.' + ext for ext in allowed_extensions)):
        print(f"Invalid file type: {file.filename}")
        return jsonify({'error': 'Invalid file type. Only PNG, JPG, JPEG, GIF, and PDF files are allowed.'}), 400
    
    try:
        # Get agent profile
        agent_profile = AgentProfile.query.filter_by(user_id=user_id).first()
        if not agent_profile:
            print("Agent profile not found")
            return jsonify({'error': 'Agent profile not found'}), 404
        
        print(f"Found agent profile ID: {agent_profile.id}")
        
        # Generate unique filename
        import os
        from werkzeug.utils import secure_filename
        filename = secure_filename(file.filename)
        timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
        unique_filename = f"license_{user_id}_{timestamp}_{filename}"
        print(f"Generated filename: {unique_filename}")
        
        # Create upload directory if it doesn't exist
        upload_dir = os.path.join(os.getcwd(), 'uploads', 'licenses')
        print(f"Upload directory: {upload_dir}")
        os.makedirs(upload_dir, exist_ok=True)
        print(f"Upload directory created/exists: {os.path.exists(upload_dir)}")
        
        # Save file
        file_path = os.path.join(upload_dir, unique_filename)
        print(f"Full file path: {file_path}")
        file.save(file_path)
        print(f"File saved successfully: {os.path.exists(file_path)}")
        
        # Update agent profile with license picture URL
        license_url = f"/uploads/licenses/{unique_filename}"
        print(f"License URL: {license_url}")
        agent_profile.license_picture_url = license_url
        
        print("About to commit to database...")
        db.session.commit()
        print("Database commit successful")
        
        return jsonify({
            'message': 'License picture uploaded successfully',
            'license_picture_url': license_url
        }), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"Error uploading license picture: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Failed to upload license picture'}), 500

# Serve uploaded license pictures
@app.route('/uploads/licenses/<filename>')
def serve_license_picture(filename):
    """Serve uploaded license pictures"""
    print(f"=== Serving License Picture ===")
    print(f"Requested filename: {filename}")
    
    upload_dir = os.path.join(os.getcwd(), 'uploads', 'licenses')
    print(f"Upload directory: {upload_dir}")
    
    file_path = os.path.join(upload_dir, filename)
    print(f"Full file path: {file_path}")
    print(f"File exists: {os.path.exists(file_path)}")
    
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        return jsonify({'error': 'File not found'}), 404
    
    try:
        response = send_from_directory(upload_dir, filename)
        # Add CORS headers
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Methods'] = 'GET'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        return response
    except Exception as e:
        print(f"Error serving file: {e}")
        return jsonify({'error': 'Error serving file'}), 500

# Get user properties and recommendations
@app.route('/api/user/properties', methods=['GET'])
@require_auth
def get_user_properties():
    """Get properties for user dashboard with recommendations"""
    try:
        # Get properties from database (you can add filtering logic here)
        properties = Property.query.limit(10).all()
        
        # Convert to JSON format
        properties_data = []
        for prop in properties:
            properties_data.append({
                'id': prop.id,
                'title': prop.title or f"{prop.property_type} Property",
                'address': prop.address,
                'price': f"${prop.asking_price:,}" if prop.asking_price else "Price on request",
                'propertyType': prop.property_type,
                'image': 'https://images.unsplash.com/photo-1560472354-b33ff0c44a43?w=400&h=300&fit=crop'
            })
        
        return jsonify({
            'properties': properties_data,
            'message': 'Properties retrieved successfully'
        }), 200
        
    except Exception as e:
        print(f"Error fetching properties: {e}")
        # Return empty properties if there's an error
        return jsonify({
            'properties': [],
            'message': 'No properties available at the moment'
        }), 200

# Get review statistics (overall rating, total reviews, star distribution)
@app.route('/api/reviews/statistics', methods=['GET'])
def get_review_statistics():
    """Get overall review statistics including average rating, total reviews, and star distribution"""
    try:
        # Get overall statistics
        stats_query = text("""
            SELECT 
                COUNT(*) as total_reviews,
                AVG(rating)::numeric(3,2) as average_rating,
                COUNT(CASE WHEN rating = 5 THEN 1 END) as five_stars,
                COUNT(CASE WHEN rating = 4 THEN 1 END) as four_stars,
                COUNT(CASE WHEN rating = 3 THEN 1 END) as three_stars,
                COUNT(CASE WHEN rating = 2 THEN 1 END) as two_stars,
                COUNT(CASE WHEN rating = 1 THEN 1 END) as one_star
            FROM user_reviews
            WHERE is_verified = TRUE
        """)
        
        result = db.session.execute(stats_query).fetchone()
        
        if not result or result.total_reviews == 0:
            return jsonify({
                'total_reviews': 0,
                'average_rating': 0,
                'star_distribution': {
                    '5': 0,
                    '4': 0,
                    '3': 0,
                    '2': 0,
                    '1': 0
                },
                'star_percentages': {
                    '5': 0,
                    '4': 0,
                    '3': 0,
                    '2': 0,
                    '1': 0
                }
            }), 200
        
        total_reviews = result.total_reviews
        average_rating = float(result.average_rating) if result.average_rating else 0
        five_stars = result.five_stars or 0
        four_stars = result.four_stars or 0
        three_stars = result.three_stars or 0
        two_stars = result.two_stars or 0
        one_star = result.one_star or 0
        
        return jsonify({
            'total_reviews': total_reviews,
            'average_rating': round(average_rating, 1),
            'star_distribution': {
                '5': five_stars,
                '4': four_stars,
                '3': three_stars,
                '2': two_stars,
                '1': one_star
            },
            'star_percentages': {
                '5': round((five_stars / total_reviews) * 100, 1),
                '4': round((four_stars / total_reviews) * 100, 1),
                '3': round((three_stars / total_reviews) * 100, 1),
                '2': round((two_stars / total_reviews) * 100, 1),
                '1': round((one_star / total_reviews) * 100, 1)
            }
        }), 200
        
    except Exception as e:
        print(f"Error fetching review statistics: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to retrieve review statistics: {str(e)}'}), 500

# Get published user reviews for feedback page
@app.route('/api/feedback/reviews', methods=['GET'])
def get_published_reviews():
    """Get published reviews from user_reviews table"""
    try:
        # User model is already imported at the top
        
        # Get verified/published reviews with user information and real like/dislike counts
        query = text("""
            SELECT 
                ur.id,
                ur.rating,
                ur.review_text,
                ur.review_date,
                ur.review_type,
                ur.admin_response,
                ur.admin_response_date,
                u.full_name,
                u.email,
                COALESCE(ri.likes_count, 0) as likes,
                COALESCE(ri.dislikes_count, 0) as dislikes
            FROM user_reviews ur
            JOIN users u ON ur.user_id = u.id
            LEFT JOIN (
                SELECT 
                    review_id,
                    COUNT(CASE WHEN interaction_type = 'like' THEN 1 END) as likes_count,
                    COUNT(CASE WHEN interaction_type = 'dislike' THEN 1 END) as dislikes_count
                FROM review_interactions
                GROUP BY review_id
            ) ri ON ur.id = ri.review_id
            WHERE ur.is_verified = TRUE
            ORDER BY ur.review_date DESC
            LIMIT 10
        """)
        
        result = db.session.execute(query)
        reviews = []
        
        for row in result:
            # Generate a default avatar based on user's name
            avatar_url = f"https://ui-avatars.com/api/?name={row.full_name}&background=random&size=40"
            
            reviews.append({
                'id': row.id,
                'name': row.full_name,
                'time': format_time_ago(row.review_date),
                'rating': row.rating,
                'text': row.review_text,
                'likes': row.likes,
                'dislikes': row.dislikes,
                'image': avatar_url,
                'review_type': row.review_type,
                'review_date': row.review_date.isoformat() if row.review_date else None,
                'admin_response': row.admin_response,
                'admin_response_date': row.admin_response_date.isoformat() if row.admin_response_date else None
            })
        
        return jsonify({
            'reviews': reviews,
            'message': 'Reviews retrieved successfully'
        }), 200
        
    except Exception as e:
        print(f"Error fetching reviews: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to retrieve reviews: {str(e)}'}), 500

# Submit new user review (rating + review text)
@app.route('/api/review/submit', methods=['POST'])
@require_auth
def submit_review():
    """Submit new user review with rating"""
    try:
        data = request.get_json()
        user_id = request.user['user_id']
        
        if not data or not data.get('rating') or not data.get('review_text'):
            return jsonify({'error': 'Rating and review text are required'}), 400
        
        # Create new review
        new_review = {
            'user_id': user_id,
            'review_type': data.get('review_type', 'platform'),
            'rating': data['rating'],
            'review_text': data['review_text'],
            'review_date': datetime.utcnow(),
            'is_verified': False  # Admin needs to verify/publish
        }
        
        # Insert into database
        query = text("""
            INSERT INTO user_reviews (user_id, review_type, rating, review_text, review_date, is_verified)
            VALUES (:user_id, :review_type, :rating, :review_text, :review_date, :is_verified)
            RETURNING id
        """)
        
        result = db.session.execute(query, new_review)
        review_id = result.fetchone()[0]
        db.session.commit()
        
        return jsonify({
            'message': 'Review submitted successfully!',
            'review_id': review_id
        }), 201
        
    except Exception as e:
        db.session.rollback()
        print(f"Error submitting review: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to submit review: {str(e)}'}), 500

# Get user's feedbacks
@app.route('/api/feedback/my-feedbacks', methods=['GET'])
@require_auth
def get_my_feedbacks():
    """Get current user's feedbacks/inquiries"""
    try:
        current_user = request.user
        user_id = current_user['user_id']
        
        # Use raw SQL to check if admin_response columns exist and fetch them
        from sqlalchemy import text
        
        try:
            # Check if admin_response columns exist
            check_query = text("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name = 'business_inquiries' AND column_name IN ('admin_response', 'admin_response_date')
            """)
            existing_columns = [row[0] for row in db.session.execute(check_query).fetchall()]
            has_admin_response_col = 'admin_response' in existing_columns
            has_admin_response_date_col = 'admin_response_date' in existing_columns
            
            if has_admin_response_col and has_admin_response_date_col:
                # Query with admin_response columns
                query = text("""
                    SELECT id, inquiry_type, message, status, created_at, admin_response, admin_response_date
                    FROM business_inquiries
                    WHERE user_id = :user_id
                    ORDER BY created_at DESC
                """)
                result = db.session.execute(query, {'user_id': user_id})
                
                feedbacks = []
                for row in result:
                    feedbacks.append({
                        'id': row.id,
                        'inquiry_type': row.inquiry_type,
                        'message': row.message,
                        'status': row.status,
                        'created_at': row.created_at.isoformat() if row.created_at else None,
                        'admin_response': row.admin_response if row.admin_response else None,
                        'admin_response_date': row.admin_response_date.isoformat() if row.admin_response_date else None
                    })
            else:
                # Fallback: use ORM without admin_response columns
                inquiries = BusinessInquiry.query.filter_by(user_id=user_id).order_by(BusinessInquiry.created_at.desc()).all()
                feedbacks = []
                for inquiry in inquiries:
                    feedbacks.append({
                        'id': inquiry.id,
                        'inquiry_type': inquiry.inquiry_type,
                        'message': inquiry.message,
                        'status': inquiry.status,
                        'created_at': inquiry.created_at.isoformat() if inquiry.created_at else None,
                        'admin_response': None,
                        'admin_response_date': None
                    })
        except Exception as e:
            print(f"Error checking/fetching admin_response columns: {e}")
            # Fallback: use ORM without admin_response columns
            inquiries = BusinessInquiry.query.filter_by(user_id=user_id).order_by(BusinessInquiry.created_at.desc()).all()
            feedbacks = []
            for inquiry in inquiries:
                feedbacks.append({
                    'id': inquiry.id,
                    'inquiry_type': inquiry.inquiry_type,
                    'message': inquiry.message,
                    'status': inquiry.status,
                    'created_at': inquiry.created_at.isoformat() if inquiry.created_at else None,
                    'admin_response': None,
                    'admin_response_date': None
                })
        
        return jsonify({
            'feedbacks': feedbacks
        }), 200
        
    except Exception as e:
        print(f"Error getting user feedbacks: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to get feedbacks: {str(e)}'}), 500

# Submit new user feedback (business inquiry - feature request, bug report, general feedback)
@app.route('/api/feedback/submit', methods=['POST'])
@require_auth
def submit_feedback():
    """Submit new user feedback/inquiry"""
    try:
        data = request.get_json()
        current_user = request.user
        user_id = current_user['user_id']
        user = User.query.get(user_id)
        
        if not data or not data.get('inquiry_type') or not data.get('message'):
            return jsonify({'error': 'Feedback type and message are required'}), 400
        
        # Validate that the inquiry_type exists and is active in feedback_form_types
        inquiry_type = data['inquiry_type']
        try:
            from sqlalchemy import text
            # First, try to remove the CHECK constraint if it exists (to allow dynamic types)
            try:
                alter_query = text("""
                    ALTER TABLE business_inquiries 
                    DROP CONSTRAINT IF EXISTS business_inquiries_inquiry_type_check
                """)
                db.session.execute(alter_query)
                db.session.commit()
                print("Removed restrictive CHECK constraint on inquiry_type to allow dynamic feedback types")
            except Exception as alter_error:
                # Constraint might not exist or have different name, try alternative
                try:
                    # PostgreSQL might name it differently
                    alter_query2 = text("""
                        DO $$ 
                        BEGIN
                            IF EXISTS (
                                SELECT 1 FROM pg_constraint 
                                WHERE conname LIKE '%inquiry_type%check%'
                                AND conrelid = 'business_inquiries'::regclass
                            ) THEN
                                EXECUTE 'ALTER TABLE business_inquiries DROP CONSTRAINT ' || 
                                    (SELECT conname FROM pg_constraint 
                                     WHERE conname LIKE '%inquiry_type%check%' 
                                     AND conrelid = 'business_inquiries'::regclass LIMIT 1);
                            END IF;
                        END $$;
                    """)
                    db.session.execute(alter_query2)
                    db.session.commit()
                except:
                    pass  # If constraint removal fails, continue anyway
            
            # Validate against feedback_form_types table
            try:
                type_check_query = text("""
                    SELECT id, name, value, status 
                    FROM feedback_form_types 
                    WHERE value = :value AND status = 'active'
                """)
                type_result = db.session.execute(type_check_query, {'value': inquiry_type}).fetchone()
                
                # If type not found in database, check if it's a legacy allowed type
                if not type_result:
                    # Check if feedback_form_types table has any entries (might be empty)
                    table_check = text("SELECT COUNT(*) FROM feedback_form_types")
                    type_count = db.session.execute(table_check).scalar()
                    
                    if type_count == 0:
                        # Table exists but empty, allow legacy types
                        legacy_types = ['property_viewing', 'price_quote', 'general', 'support']
                        if inquiry_type not in legacy_types:
                            return jsonify({
                                'error': f'Invalid feedback type: {inquiry_type}. Please select a valid type from the dropdown.'
                            }), 400
                    else:
                        # Table has entries but this type is not active
                        return jsonify({
                            'error': f'Invalid or inactive feedback type: {inquiry_type}. Please select an active type from the dropdown.'
                        }), 400
            except Exception as table_error:
                # If feedback_form_types table doesn't exist, allow legacy types
                print(f"Note: feedback_form_types table may not exist: {table_error}")
                legacy_types = ['property_viewing', 'price_quote', 'general', 'support']
                if inquiry_type not in legacy_types:
                    return jsonify({
                        'error': f'Invalid feedback type: {inquiry_type}. Please select a valid type from the dropdown.'
                    }), 400
        except Exception as type_check_error:
            # If validation fails completely, still try to submit (database will validate)
            print(f"Warning: Could not validate feedback type {inquiry_type}: {type_check_error}")
        
        # Get user information
        name = user.full_name if user else data.get('name', 'User')
        email = user.email if user else data.get('email', '')
        phone = user.phone_number if user else data.get('phone', '')
        
        # Create new business inquiry - set status to 'in_progress' for not yet responded feedback
        new_inquiry = BusinessInquiry(
            user_id=user_id,
            name=name,
            email=email,
            phone=phone,
            inquiry_type=inquiry_type,
            message=data['message'],
            status='in_progress'
        )
        
        db.session.add(new_inquiry)
        db.session.commit()
        
        return jsonify({
            'message': 'Feedback submitted successfully! Our team will review it and get back to you.',
            'inquiry_id': new_inquiry.id
        }), 201
        
    except Exception as e:
        db.session.rollback()
        print(f"Error submitting feedback: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to submit feedback: {str(e)}'}), 500

# Like a review
@app.route('/api/feedback/like', methods=['POST'])
@require_auth
def like_review():
    """Like or unlike a review"""
    try:
        data = request.get_json()
        user_id = request.user['user_id']
        review_id = data.get('review_id')
        
        if not review_id:
            return jsonify({'error': 'Review ID is required'}), 400
        
        # Check if user already liked this review
        check_query = text("""
            SELECT interaction_type FROM review_interactions 
            WHERE user_id = :user_id AND review_id = :review_id
        """)
        
        result = db.session.execute(check_query, {'user_id': user_id, 'review_id': review_id})
        existing_interaction = result.fetchone()
        
        if existing_interaction:
            if existing_interaction[0] == 'like':
                # User already liked, remove the like
                delete_query = text("""
                    DELETE FROM review_interactions 
                    WHERE user_id = :user_id AND review_id = :review_id
                """)
                db.session.execute(delete_query, {'user_id': user_id, 'review_id': review_id})
                action = 'removed'
            else:
                # User disliked, switch to like
                update_query = text("""
                    UPDATE review_interactions 
                    SET interaction_type = 'like', interaction_date = CURRENT_TIMESTAMP
                    WHERE user_id = :user_id AND review_id = :review_id
                """)
                db.session.execute(update_query, {'user_id': user_id, 'review_id': review_id})
                action = 'switched_to_like'
        else:
            # User hasn't interacted, add like
            insert_query = text("""
                INSERT INTO review_interactions (user_id, review_id, interaction_type)
                VALUES (:user_id, :review_id, 'like')
            """)
            db.session.execute(insert_query, {'user_id': user_id, 'review_id': review_id})
            action = 'added'
        
        db.session.commit()
        
        # Get updated counts
        count_query = text("""
            SELECT 
                COUNT(CASE WHEN interaction_type = 'like' THEN 1 END) as likes,
                COUNT(CASE WHEN interaction_type = 'dislike' THEN 1 END) as dislikes
            FROM review_interactions 
            WHERE review_id = :review_id
        """)
        
        count_result = db.session.execute(count_query, {'review_id': review_id})
        counts = count_result.fetchone()
        
        return jsonify({
            'message': f'Like {action} successfully',
            'action': action,
            'likes': counts[0],
            'dislikes': counts[1]
        }), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"Error handling like: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to handle like: {str(e)}'}), 500

# Dislike a review
@app.route('/api/feedback/dislike', methods=['POST'])
@require_auth
def dislike_review():
    """Dislike or undislike a review"""
    try:
        data = request.get_json()
        user_id = request.user['user_id']
        review_id = data.get('review_id')
        
        if not review_id:
            return jsonify({'error': 'Review ID is required'}), 400
        
        # Check if user already disliked this review
        check_query = text("""
            SELECT interaction_type FROM review_interactions 
            WHERE user_id = :user_id AND review_id = :review_id
        """)
        
        result = db.session.execute(check_query, {'user_id': user_id, 'review_id': review_id})
        existing_interaction = result.fetchone()
        
        if existing_interaction:
            if existing_interaction[0] == 'dislike':
                # User already disliked, remove the dislike
                delete_query = text("""
                    DELETE FROM review_interactions 
                    WHERE user_id = :user_id AND review_id = :review_id
                """)
                db.session.execute(delete_query, {'user_id': user_id, 'review_id': review_id})
                action = 'removed'
            else:
                # User liked, switch to dislike
                update_query = text("""
                    UPDATE review_interactions 
                    SET interaction_type = 'dislike', interaction_date = CURRENT_TIMESTAMP
                    WHERE user_id = :user_id AND review_id = :review_id
                """)
                db.session.execute(update_query, {'user_id': user_id, 'review_id': review_id})
                action = 'switched_to_dislike'
        else:
            # User hasn't interacted, add dislike
            insert_query = text("""
                INSERT INTO review_interactions (user_id, review_id, interaction_type)
                VALUES (:user_id, :review_id, 'dislike')
            """)
            db.session.execute(insert_query, {'user_id': user_id, 'review_id': review_id})
            action = 'added'
        
        db.session.commit()
        
        # Get updated counts
        count_query = text("""
            SELECT 
                COUNT(CASE WHEN interaction_type = 'like' THEN 1 END) as likes,
                COUNT(CASE WHEN interaction_type = 'dislike' THEN 1 END) as dislikes
            FROM review_interactions 
            WHERE review_id = :review_id
        """)
        
        count_result = db.session.execute(count_query, {'review_id': review_id})
        counts = count_result.fetchone()
        
        return jsonify({
            'message': f'Dislike {action} successfully',
            'action': action,
            'likes': counts[0],
            'dislikes': counts[1]
        }), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"Error handling dislike: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to handle dislike: {str(e)}'}), 500

# Get user's interaction state for reviews
@app.route('/api/feedback/interactions', methods=['GET'])
@require_auth
def get_user_interactions():
    """Get current user's interaction state for all reviews"""
    try:
        user_id = request.user['user_id']
        
        # Get all user interactions
        query = text("""
            SELECT review_id, interaction_type
            FROM review_interactions
            WHERE user_id = :user_id
        """)
        
        result = db.session.execute(query, {'user_id': user_id})
        interactions = {}
        
        for row in result:
            interactions[row.review_id] = row.interaction_type
        
        return jsonify({
            'interactions': interactions,
            'message': 'User interactions retrieved successfully'
        }), 200
        
    except Exception as e:
        print(f"Error fetching user interactions: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to retrieve user interactions: {str(e)}'}), 500

# Get current user's reviews
@app.route('/api/feedback/my-reviews', methods=['GET'])
@require_auth
def get_my_reviews():
    """Get reviews submitted by the current user"""
    try:
        user_id = request.user['user_id']
        print(f"Fetching reviews for user_id: {user_id}")
        
        # First, check if the user_reviews table exists
        try:
            # Simple query to check if table exists
            check_query = text("SELECT COUNT(*) FROM user_reviews LIMIT 1")
            db.session.execute(check_query)
            print("user_reviews table exists")
        except Exception as table_error:
            print(f"Table check error: {table_error}")
            return jsonify({
                'reviews': [],
                'message': 'No reviews table found. Please contact administrator.'
            }), 200
        
        # Get user reviews with real like/dislike counts
        query = text("""
            SELECT 
                ur.id,
                ur.rating,
                ur.review_text,
                ur.review_date,
                ur.review_type,
                ur.is_verified,
                ur.admin_response,
                ur.admin_response_date,
                COALESCE(ri.likes_count, 0) as likes,
                COALESCE(ri.dislikes_count, 0) as dislikes
            FROM user_reviews ur
            LEFT JOIN (
                SELECT 
                    review_id,
                    COUNT(CASE WHEN interaction_type = 'like' THEN 1 END) as likes_count,
                    COUNT(CASE WHEN interaction_type = 'dislike' THEN 1 END) as dislikes_count
                FROM review_interactions
                GROUP BY review_id
            ) ri ON ur.id = ri.review_id
            WHERE ur.user_id = :user_id
            ORDER BY ur.review_date DESC
        """)
        
        print(f"Executing query: {query}")
        result = db.session.execute(query, {'user_id': user_id})
        reviews = []
        
        for row in result:
            reviews.append({
                'id': row.id,
                'time': format_time_ago(row.review_date),
                'rating': row.rating,
                'text': row.review_text,
                'likes': row.likes,
                'dislikes': row.dislikes,
                'is_verified': row.is_verified,
                'review_type': row.review_type,
                'review_date': row.review_date.isoformat() if row.review_date else None,
                'admin_response': row.admin_response,
                'admin_response_date': row.admin_response_date.isoformat() if row.admin_response_date else None
            })
        
        print(f"Found {len(reviews)} reviews for user {user_id}")
        
        return jsonify({
            'reviews': reviews,
            'message': 'User reviews retrieved successfully'
        }), 200
        
    except Exception as e:
        print(f"Error fetching user reviews: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to retrieve user reviews: {str(e)}'}), 500

def format_time_ago(timestamp):
    """Format timestamp to relative time string"""
    if not timestamp:
        return 'Unknown time'
    
    now = datetime.utcnow()
    diff = now - timestamp
    
    if diff.days > 0:
        if diff.days == 1:
            return '1 day ago'
        elif diff.days < 7:
            return f'{diff.days} days ago'
        elif diff.days < 30:
            weeks = diff.days // 7
            return f'{weeks} week{"s" if weeks > 1 else ""} ago'
        else:
            months = diff.days // 30
            return f'{months} month{"s" if months > 1 else ""} ago'
    elif diff.seconds > 3600:
        hours = diff.seconds // 3600
        return f'{hours} hour{"s" if hours > 1 else ""} ago'
    elif diff.seconds > 60:
        minutes = diff.seconds // 60
        return f'{minutes} minute{"s" if minutes > 1 else ""} ago'
    else:
        return 'Just now'

# Update user profile
@app.route('/api/auth/profile', methods=['PUT'])
@require_auth
def update_profile():
    data = request.get_json()
    
    # Get user from authenticated token
    user_id = request.user['user_id']
    user = User.query.get(user_id)
    
    if not user:
        return jsonify({'error': 'User not found'}), 404
    
    try:
        # Update allowed fields
        if 'full_name' in data:
            user.full_name = data['full_name']
        if 'phone_number' in data:
            user.phone_number = data['phone_number']
        
        # Handle agent information updates
        agent_profile = None
        if user.user_type == 'agent' and 'agent_info' in data:
            agent_profile = AgentProfile.query.filter_by(user_id=user_id).first()
            if not agent_profile:
                return jsonify({'error': 'Agent profile not found'}), 404
            
            agent_info = data['agent_info']
            if 'cea_number' in agent_info:
                agent_profile.license_number = agent_info['cea_number']
            if 'company_name' in agent_info:
                agent_profile.company_name = agent_info['company_name']
            if 'company_phone' in agent_info:
                agent_profile.company_phone = agent_info['company_phone']
            if 'company_email' in agent_info:
                agent_profile.company_email = agent_info['company_email']
            if 'specializations' in agent_info:
                agent_profile.specializations = agent_info['specializations']
        
        db.session.commit()
        
        # Prepare response
        response_data = {
            'message': 'Profile updated successfully',
            'user': {
                'id': user.id,
                'email': user.email,
                'full_name': user.full_name,
                'phone_number': user.phone_number,
                'user_type': user.user_type
            }
        }
        
        # Include agent profile data if updated
        if agent_profile:
            response_data['agent_profile'] = {
                'cea_number': agent_profile.license_number,
                'company_name': agent_profile.company_name,
                'company_phone': agent_profile.company_phone if hasattr(agent_profile, 'company_phone') else None,
                'company_email': agent_profile.company_email if hasattr(agent_profile, 'company_email') else None,
                'specializations': agent_profile.specializations if agent_profile.specializations else [],
                'license_picture_url': agent_profile.license_picture_url if hasattr(agent_profile, 'license_picture_url') else None
            }
        
        return jsonify(response_data), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"Error updating profile: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Failed to update profile'}), 500

# Change password
@app.route('/api/auth/change-password', methods=['POST'])
@require_auth
def change_password():
    data = request.get_json()
    
    if not data or not data.get('current_password') or not data.get('new_password'):
        return jsonify({'error': 'Current password and new password are required'}), 400
    
    # Get user from authenticated token
    user_id = request.user['user_id']
    user = User.query.get(user_id)
    
    if not user:
        return jsonify({'error': 'User not found'}), 404
    
    # Verify current password
    if not user.check_password(data['current_password']):
        return jsonify({'error': 'Current password is incorrect'}), 400
    
    # Validate new password length
    new_password = data.get('new_password', '')
    if len(new_password) < 8:
        return jsonify({'error': 'New password must be at least 8 characters long'}), 400
    
    try:
        # Set new password
        user.set_password(new_password)
        db.session.commit()
        
        return jsonify({'message': 'Password changed successfully'}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'Failed to change password'}), 500

# Upgrade to Premium
@app.route('/api/auth/upgrade-to-premium', methods=['POST'])
@require_auth
def upgrade_to_premium():
    # Get user from authenticated token
    user_id = request.user['user_id']
    user = User.query.get(user_id)
    
    if not user:
        return jsonify({'error': 'User not found'}), 404
    
    if user.user_type == 'premium':
        return jsonify({'error': 'User is already premium'}), 400
    
    try:
        # Update user to premium
        user.user_type = 'premium'
        user.subscription_status = 'active'
        user.subscription_start_date = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify({
            'message': 'Successfully upgraded to Premium',
            'user': {
                'id': user.id,
                'email': user.email,
                'user_type': user.user_type,
                'subscription_status': user.subscription_status,
                'subscription_start_date': user.subscription_start_date.isoformat()
            }
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'Failed to upgrade to premium'}), 500

# Downgrade to Free
@app.route('/api/auth/downgrade-to-free', methods=['POST'])
@require_auth
def downgrade_to_free():
    # Get user from authenticated token
    user_id = request.user['user_id']
    user = User.query.get(user_id)
    
    if not user:
        return jsonify({'error': 'User not found'}), 404
    
    if user.user_type == 'free':
        return jsonify({'error': 'User is already on free plan'}), 400
    
    try:
        # Update user to free
        user.user_type = 'free'
        user.subscription_status = 'inactive'
        user.subscription_start_date = None
        
        db.session.commit()
        
        return jsonify({
            'message': 'Successfully downgraded to Free plan',
            'user': {
                'id': user.id,
                'email': user.email,
                'user_type': user.user_type,
                'subscription_status': user.subscription_status
            }
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'Failed to downgrade to free'}), 500

# Delete account
@app.route('/api/auth/deactivate', methods=['POST'])
def delete_account():
    data = request.get_json()
    
    if not data or not data.get('email'):
        return jsonify({'error': 'Email is required'}), 400
    
    email = data['email']
    user = User.query.filter_by(email=email).first()
    
    if not user:
        return jsonify({'error': 'User not found'}), 404
    
    try:
        # Completely delete user account from database
        db.session.delete(user)
        db.session.commit()
        
        return jsonify({'message': 'Account deleted successfully'}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'Failed to delete account'}), 500

# ==================== SUBSCRIPTION PLANS API ====================

# Get all subscription plans
@app.route('/api/subscription-plans', methods=['GET'])
def get_subscription_plans():
    try:
        plans = SubscriptionPlan.query.filter_by(is_active=True).order_by(SubscriptionPlan.display_order).all()
        
        plans_data = []
        for plan in plans:
            plan_data = {
                'id': plan.id,
                'plan_name': plan.plan_name,
                'plan_type': plan.plan_type,
                'monthly_price': float(plan.monthly_price),
                'yearly_price': float(plan.yearly_price),
                'description': plan.description,
                'is_active': plan.is_active,
                'is_popular': plan.is_popular,
                'display_order': plan.display_order,
                'features': []
            }
            
            # Get features for this plan
            features = SubscriptionPlanFeature.query.filter_by(plan_id=plan.id).order_by(SubscriptionPlanFeature.display_order).all()
            for feature in features:
                plan_data['features'].append({
                    'id': feature.id,
                    'feature_name': feature.feature_name,
                    'feature_description': feature.feature_description,
                    'is_included': feature.is_included,
                    'display_order': feature.display_order
                })
            
            plans_data.append(plan_data)
        
        return jsonify({
            'success': True,
            'plans': plans_data
        }), 200
        
    except Exception as e:
        return jsonify({'error': 'Failed to fetch subscription plans'}), 500

# Get single subscription plan
@app.route('/api/subscription-plans/<int:plan_id>', methods=['GET'])
def get_subscription_plan(plan_id):
    try:
        plan = SubscriptionPlan.query.get(plan_id)
        
        if not plan:
            return jsonify({'error': 'Plan not found'}), 404
        
        plan_data = {
            'id': plan.id,
            'plan_name': plan.plan_name,
            'plan_type': plan.plan_type,
            'monthly_price': float(plan.monthly_price),
            'yearly_price': float(plan.yearly_price),
            'description': plan.description,
            'is_active': plan.is_active,
            'is_popular': plan.is_popular,
            'display_order': plan.display_order,
            'features': []
        }
        
        # Get features for this plan
        features = SubscriptionPlanFeature.query.filter_by(plan_id=plan.id).order_by(SubscriptionPlanFeature.display_order).all()
        for feature in features:
            plan_data['features'].append({
                'id': feature.id,
                'feature_name': feature.feature_name,
                'feature_description': feature.feature_description,
                'is_included': feature.is_included,
                'display_order': feature.display_order
            })
        
        return jsonify({
            'success': True,
            'plan': plan_data
        }), 200
        
    except Exception as e:
        return jsonify({'error': 'Failed to fetch subscription plan'}), 500

# Create subscription plan (Admin only)
@app.route('/api/admin/subscription-plans', methods=['POST'])
@require_auth
def create_subscription_plan():
    try:
        # Check if user is admin
        user_id = request.user['user_id']
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        data = request.get_json()
        
        if not data or not data.get('plan_name') or not data.get('plan_type'):
            return jsonify({'error': 'Plan name and type are required'}), 400
        
        # Create new plan
        plan = SubscriptionPlan(
            plan_name=data['plan_name'],
            plan_type=data['plan_type'],
            monthly_price=data.get('monthly_price', 0.00),
            yearly_price=data.get('yearly_price', 0.00),
            description=data.get('description', ''),
            display_order=data.get('display_order', 0)
        )
        
        db.session.add(plan)
        db.session.flush()  # Get the plan ID
        
        # Add features if provided
        if data.get('features'):
            for feature_data in data['features']:
                feature = SubscriptionPlanFeature(
                    plan_id=plan.id,
                    feature_name=feature_data['feature_name'],
                    feature_description=feature_data.get('feature_description', ''),
                    is_included=feature_data.get('is_included', True),
                    display_order=feature_data.get('display_order', 0)
                )
                db.session.add(feature)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Subscription plan created successfully',
            'plan_id': plan.id
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'Failed to create subscription plan'}), 500

# Update subscription plan (Admin only)
@app.route('/api/admin/subscription-plans/<int:plan_id>', methods=['PUT'])
@require_auth
def update_subscription_plan(plan_id):
    try:
        print(f"🔍 Update subscription plan called for ID: {plan_id}")
        print(f"🔍 Request data: {request.get_json()}")
        
        # Check if user is admin
        user_id = request.user['user_id']
        print(f"🔍 User ID: {user_id}")
        user = User.query.get(user_id)
        print(f"🔍 User found: {user}, User type: {user.user_type if user else 'None'}")
        
        if not user or user.user_type != 'admin':
            print("❌ Admin access denied")
            return jsonify({'error': 'Admin access required'}), 403
        
        plan = SubscriptionPlan.query.get(plan_id)
        if not plan:
            return jsonify({'error': 'Plan not found'}), 404
        
        data = request.get_json()
        
        # Update plan details
        if 'plan_name' in data:
            plan.plan_name = data['plan_name']
        if 'plan_type' in data:
            plan.plan_type = data['plan_type']
        if 'monthly_price' in data:
            plan.monthly_price = data['monthly_price']
        if 'yearly_price' in data:
            plan.yearly_price = data['yearly_price']
        if 'description' in data:
            plan.description = data['description']
        if 'display_order' in data:
            plan.display_order = data['display_order']
        if 'is_active' in data:
            plan.is_active = data['is_active']
        if 'is_popular' in data:
            plan.is_popular = data['is_popular']
        
        # Update features if provided
        if 'features' in data:
            # Delete existing features
            SubscriptionPlanFeature.query.filter_by(plan_id=plan_id).delete()
            
            # Add new features
            for feature_data in data['features']:
                feature = SubscriptionPlanFeature(
                    plan_id=plan_id,
                    feature_name=feature_data['feature_name'],
                    feature_description=feature_data.get('feature_description', ''),
                    is_included=feature_data.get('is_included', True),
                    display_order=feature_data.get('display_order', 0)
                )
                db.session.add(feature)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Subscription plan updated successfully'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'Failed to update subscription plan'}), 500

# Delete subscription plan (Admin only)
@app.route('/api/admin/subscription-plans/<int:plan_id>', methods=['DELETE'])
@require_auth
def delete_subscription_plan(plan_id):
    try:
        # Check if user is admin
        user_id = request.user['user_id']
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        plan = SubscriptionPlan.query.get(plan_id)
        if not plan:
            return jsonify({'error': 'Plan not found'}), 404
        
        # Delete plan (features will be deleted automatically due to cascade)
        db.session.delete(plan)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Subscription plan deleted successfully'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'Failed to delete subscription plan'}), 500

# ==================== IMPORTANT FEATURES API ====================

# Get all important features
@app.route('/api/important-features', methods=['GET'])
def get_important_features():
    try:
        features = ImportantFeature.query.filter_by(is_active=True).order_by(ImportantFeature.display_order).all()
        
        features_data = []
        for feature in features:
            features_data.append({
                'id': feature.id,
                'feature_name': feature.feature_name,
                'feature_description': feature.feature_description,
                'display_order': feature.display_order
            })
        
        return jsonify({
            'success': True,
            'features': features_data
        }), 200
        
    except Exception as e:
        return jsonify({'error': 'Failed to fetch important features'}), 500

# Create important feature (Admin only)
@app.route('/api/admin/important-features', methods=['POST'])
@require_auth
def create_important_feature():
    try:
        # Check if user is admin
        user_id = request.user['user_id']
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        data = request.get_json()
        print(f"🔍 Received data for creating important feature: {data}")
        
        if not data or not data.get('feature_name'):
            print("❌ Missing feature_name in request data")
            return jsonify({'error': 'Feature name is required'}), 400
        
        feature = ImportantFeature(
            feature_name=data['feature_name'],
            feature_description=data.get('feature_description', ''),
            display_order=data.get('display_order', 0)
        )
        
        db.session.add(feature)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Important feature created successfully',
            'feature_id': feature.id
        }), 201
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error creating important feature: {str(e)}")
        print(f"❌ Error type: {type(e).__name__}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to create important feature: {str(e)}'}), 500

# Update important feature (Admin only)
@app.route('/api/admin/important-features/<int:feature_id>', methods=['PUT'])
@require_auth
def update_important_feature(feature_id):
    try:
        # Check if user is admin
        user_id = request.user['user_id']
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        feature = ImportantFeature.query.get(feature_id)
        if not feature:
            return jsonify({'error': 'Feature not found'}), 404
        
        data = request.get_json()
        
        if 'feature_name' in data:
            feature.feature_name = data['feature_name']
        if 'feature_description' in data:
            feature.feature_description = data['feature_description']
        if 'display_order' in data:
            feature.display_order = data['display_order']
        if 'is_active' in data:
            feature.is_active = data['is_active']
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Important feature updated successfully'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error updating important feature: {str(e)}")
        print(f"❌ Error type: {type(e).__name__}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to update important feature: {str(e)}'}), 500

# Delete important feature (Admin only)
@app.route('/api/admin/important-features/<int:feature_id>', methods=['DELETE'])
@require_auth
def delete_important_feature(feature_id):
    try:
        # Check if user is admin
        user_id = request.user['user_id']
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        feature = ImportantFeature.query.get(feature_id)
        if not feature:
            return jsonify({'error': 'Feature not found'}), 404
        
        db.session.delete(feature)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Important feature deleted successfully'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error deleting important feature: {str(e)}")
        print(f"❌ Error type: {type(e).__name__}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to delete important feature: {str(e)}'}), 500

# Password recovery - Verify OTP and reset password
@app.route('/api/auth/reset-password', methods=['POST'])
def reset_password():
    data = request.get_json()
    
    if not data or not data.get('email') or not data.get('otp') or not data.get('new_password'):
        return jsonify({'error': 'Missing required fields'}), 400
    
    email = data['email']
    otp = data['otp']
    new_password = data['new_password']
    
    # Check if OTP exists and is valid
    if email not in password_reset_codes:
        return jsonify({'error': 'Invalid or expired recovery code'}), 400
    
    reset_data = password_reset_codes[email]
    
    # Check if OTP has expired
    if datetime.utcnow() > reset_data['expires_at']:
        del password_reset_codes[email]
        return jsonify({'error': 'Recovery code has expired'}), 400
    
    # Check if OTP matches
    if reset_data['otp'] != otp:
        reset_data['attempts'] += 1
        
        # Remove OTP if too many attempts
        if reset_data['attempts'] >= 3:
            del password_reset_codes[email]
            return jsonify({'error': 'Too many failed attempts. Please request a new recovery code'}), 400
        
        return jsonify({'error': 'Invalid recovery code'}), 400
    
    # Validate new password length
    if len(new_password) < 8:
        return jsonify({'error': 'New password must be at least 8 characters long'}), 400
    
    # OTP is valid, reset password
    user = User.query.filter_by(email=email).first()
    if not user:
        del password_reset_codes[email]
        return jsonify({'error': 'User not found'}), 404
    
    try:
        user.set_password(new_password)
        db.session.commit()
        
        # Remove the used OTP
        del password_reset_codes[email]
        
        return jsonify({'message': 'Password reset successfully'}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'Failed to reset password'}), 500

# Admin Dashboard Endpoints

# Get admin dashboard metrics
@app.route('/api/admin/metrics', methods=['GET'])
@require_auth
def get_admin_metrics():
    try:
        print("Admin metrics endpoint called")
        
        # Get current user to verify admin status
        current_user = request.user
        print(f"Current user: {current_user}")
        
        if current_user['user_type'] != 'admin':
            print(f"User type {current_user['user_type']} is not admin")
            return jsonify({'error': 'Admin access required'}), 403
        
        print("User is admin, proceeding with metrics calculation")
        
        # Count total users
        try:
            total_users = User.query.count()
            print(f"Total users: {total_users}")
        except Exception as e:
            print(f"Error counting total users: {e}")
            total_users = 0
        
        # Count active subscriptions (premium and agent users)
        try:
            active_subscriptions = User.query.filter(
                User.subscription_status == 'active',
                User.user_type.in_(['premium', 'agent'])
            ).count()
            print(f"Active subscriptions: {active_subscriptions}")
        except Exception as e:
            print(f"Error counting active subscriptions: {e}")
            active_subscriptions = 0
        
        # Count total feedback/reviews (handle case where table might not exist)
        try:
            total_feedback = db.session.execute(text('SELECT COUNT(*) FROM business_inquiries')).scalar()
            print(f"Total feedback: {total_feedback}")
        except Exception as e:
            print(f"Error counting feedback: {e}")
            total_feedback = 0
        
        # Calculate revenue (mock calculation for now)
        try:
            premium_users = User.query.filter_by(user_type='premium', subscription_status='active').count()
            agent_users = User.query.filter_by(user_type='agent', subscription_status='active').count()
            print(f"Premium users: {premium_users}, Agent users: {agent_users}")
            
            # Mock pricing: Premium $29/month, Agent $99/month
            monthly_revenue = (premium_users * 29) + (agent_users * 99)
            print(f"Monthly revenue: {monthly_revenue}")
        except Exception as e:
            print(f"Error calculating revenue: {e}")
            monthly_revenue = 0
        
        metrics = {
            'total_users': total_users,
            'active_subscriptions': active_subscriptions,
            'total_feedback': total_feedback,
            'monthly_revenue': monthly_revenue
        }
        
        print(f"Final metrics: {metrics}")
        return jsonify(metrics), 200
        
    except Exception as e:
        print(f"Error getting admin metrics: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Failed to get admin metrics'}), 500

# Get all user accounts for admin
@app.route('/api/admin/users', methods=['GET'])
@require_auth
def get_all_users():
    try:
        # Get current user to verify admin status
        current_user = request.user
        if current_user['user_type'] != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        # Get all users with pagination
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        
        users = User.query.paginate(
            page=page, 
            per_page=per_page, 
            error_out=False
        )
        
        user_list = []
        for user in users.items:
            user_list.append({
                'id': user.id,
                'full_name': user.full_name,
                'email': user.email,
                'user_type': user.user_type,
                'subscription_status': user.subscription_status,
                'account_created_date': user.account_created_date.isoformat() if user.account_created_date else None,
                'last_login': user.last_login.isoformat() if user.last_login else None,
                'is_active': user.is_active,
                'referral_code': user.referral_code
            })
        
        return jsonify({
            'users': user_list,
            'total': users.total,
            'pages': users.pages,
            'current_page': users.page
        }), 200
        
    except Exception as e:
        print(f"Error getting users: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Failed to get users'}), 500

# Get user details by ID for admin
@app.route('/api/admin/users/<int:user_id>', methods=['GET'])
@require_auth
def get_user_by_id(user_id):
    try:
        # Get current user to verify admin status
        current_user = request.user
        if current_user['user_type'] != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        # Get user by ID
        user = User.query.get(user_id)
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        # Get user's recent activities (reviews, interactions, etc.)
        activities = []
        
        # Get user's reviews
        reviews_query = text("""
            SELECT review_date, 'Review Submitted' as type, review_text as details
            FROM user_reviews 
            WHERE user_id = :user_id
            ORDER BY review_date DESC
            LIMIT 10
        """)
        reviews_result = db.session.execute(reviews_query, {'user_id': user_id})
        for row in reviews_result:
            activities.append({
                'date': row.review_date.isoformat() if row.review_date else None,
                'type': row.type,
                'details': row.details[:100] + '...' if len(row.details) > 100 else row.details
            })
        
        # Get user's review interactions (likes/dislikes)
        interactions_query = text("""
            SELECT interaction_date, 
                   CASE WHEN interaction_type = 'like' THEN 'Liked Review' ELSE 'Disliked Review' END as type,
                   'Interacted with review' as details
            FROM review_interactions 
            WHERE user_id = :user_id
            ORDER BY interaction_date DESC
            LIMIT 5
        """)
        interactions_result = db.session.execute(interactions_query, {'user_id': user_id})
        for row in interactions_result:
            activities.append({
                'date': row.interaction_date.isoformat() if row.interaction_date else None,
                'type': row.type,
                'details': row.details
            })
        
        # Sort activities by date (most recent first)
        activities.sort(key=lambda x: x['date'] or '', reverse=True)
        
        # Limit to 15 most recent activities
        activities = activities[:15]
        
        user_data = {
            'id': user.id,
            'full_name': user.full_name,
            'email': user.email,
            'phone_number': user.phone_number,
            'user_type': user.user_type,
            'subscription_status': user.subscription_status,
            'account_created_date': user.account_created_date.isoformat() if user.account_created_date else None,
            'last_login': user.last_login.isoformat() if user.last_login else None,
            'is_active': user.is_active,
            'referral_code': user.referral_code,
            'activities': activities
        }
        
        return jsonify(user_data), 200
        
    except Exception as e:
        print(f"Error getting user details: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Failed to get user details'}), 500

# Deactivate user account
@app.route('/api/admin/users/<int:user_id>/deactivate', methods=['POST'])
@require_auth
def deactivate_user(user_id):
    try:
        # Get current user to verify admin status
        current_user = request.user
        if current_user['user_type'] != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        # Get user by ID
        user = User.query.get(user_id)
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        # Check if trying to deactivate self
        if user.id == current_user['user_id']:
            return jsonify({'error': 'Cannot deactivate your own account'}), 400
        
        # Deactivate user
        user.is_active = False
        user.subscription_status = 'inactive'  # Fixed: lowercase to match database constraint
        db.session.commit()
        
        return jsonify({
            'message': 'User deactivated successfully',
            'user_id': user_id,
            'status': 'deactivated'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"Error deactivating user: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Failed to deactivate user'}), 500

# User deactivates their own account
@app.route('/api/user/deactivate', methods=['POST'])
@require_auth
def deactivate_own_account():
    try:
        # Get current user
        current_user = request.user
        
        # Get user by ID
        user = User.query.get(current_user['user_id'])
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        # Deactivate user (same as admin deactivation)
        user.is_active = False
        user.subscription_status = 'inactive'
        db.session.commit()
        
        return jsonify({
            'message': 'Account deactivated successfully',
            'user_id': current_user['user_id'],
            'status': 'inactive'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"Error deactivating own account: {e}")
        import traceback
        traceback.print_exc()
        
        # Check for specific constraint violations
        if 'check constraint' in str(e).lower():
            return jsonify({'error': 'Invalid data format. Please contact support.'}), 400
        elif 'foreign key' in str(e).lower():
            return jsonify({'error': 'Cannot deactivate account due to existing data. Please contact support.'}), 400
        else:
            return jsonify({'error': 'Failed to deactivate account'}), 500

# Admin: Get all properties (including inactive)
@app.route('/api/admin/properties', methods=['GET'])
@require_auth
def get_all_properties_admin():
    """Get all properties for admin management"""
    try:
        # Verify admin access
        current_user = request.user
        if current_user['user_type'] != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        # Get query parameters for pagination
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        
        # Get all properties (not filtered by status)
        query = Property.query
        
        # Pagination
        pagination = query.order_by(Property.created_at.desc()).paginate(
            page=page, per_page=per_page, error_out=False
        )
        
        properties = pagination.items
        property_list = []
        
        for prop in properties:
            # Get agent information
            agent = User.query.get(prop.agent_id)
            agent_name = agent.full_name if agent else 'Unknown Agent'
            agent_email = agent.email if agent else 'N/A'
            
            # Get agent profile for additional info
            agent_profile = None
            if agent:
                agent_profile = AgentProfile.query.filter_by(user_id=agent.id).first()
            
            property_data = {
                'id': prop.id,
                'title': prop.title,
                'description': prop.description,
                'property_type': prop.property_type,
                'address': prop.address,
                'street_address': prop.street_address,
                'city': prop.city,
                'state': prop.state,
                'zip_code': prop.zip_code,
                'size_sqft': float(prop.size_sqft),
                'floors': prop.floors,
                'year_built': prop.year_built,
                'zoning': prop.zoning,
                'parking_spaces': prop.parking_spaces,
                'asking_price': float(prop.asking_price),
                'price_type': prop.price_type,
                'status': prop.status,
                'latitude': float(prop.latitude) if prop.latitude else None,
                'longitude': float(prop.longitude) if prop.longitude else None,
                'created_at': prop.created_at.isoformat() if prop.created_at else None,
                'updated_at': prop.updated_at.isoformat() if prop.updated_at else None,
                'agent_id': prop.agent_id,
                'agent_name': agent_name,
                'agent_email': agent_email,
                'agent_license': agent_profile.license_number if agent_profile else None,
                'agent_company': agent_profile.company_name if agent_profile else None
            }
            property_list.append(property_data)
        
        # Calculate total statistics (from all properties, not just current page)
        total_active = Property.query.filter_by(status='active').count()
        total_pending = Property.query.filter_by(status='pending').count()
        total_sold = Property.query.filter_by(status='sold').count()
        total_under_contract = Property.query.filter_by(status='under-contract').count()
        
        return jsonify({
            'properties': property_list,
            'total': pagination.total,
            'page': page,
            'per_page': per_page,
            'pages': pagination.pages,
            'statistics': {
                'total_active': total_active,
                'total_pending': total_pending,
                'total_sold': total_sold,
                'total_under_contract': total_under_contract
            }
        }), 200
        
    except Exception as e:
        print(f"Error fetching properties for admin: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Failed to fetch properties'}), 500

# Admin: Delete property
@app.route('/api/admin/properties/<int:property_id>', methods=['DELETE'])
@require_auth
def delete_property_admin(property_id):
    """Delete a property (admin only)"""
    try:
        # Verify admin access
        current_user = request.user
        if current_user['user_type'] != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        # Get property
        property = Property.query.get(property_id)
        if not property:
            return jsonify({'error': 'Property not found'}), 404
        
        # Delete related records first to avoid foreign key constraint errors
        # Delete property views
        PropertyView.query.filter_by(property_id=property_id).delete()
        
        # Delete property amenities
        PropertyAmenity.query.filter_by(property_id=property_id).delete()
        
        # Delete property images
        PropertyImage.query.filter_by(property_id=property_id).delete()
        
        # Delete property bookmarks (if any bookmark references this property)
        # Note: Bookmarks store property_id in reference_id when bookmark_type is 'property'
        Bookmark.query.filter_by(bookmark_type='property', reference_id=property_id).delete()
        
        # Now delete the property itself
        db.session.delete(property)
        db.session.commit()
        
        return jsonify({
            'message': 'Property deleted successfully',
            'property_id': property_id
        }), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"Error deleting property: {e}")
        import traceback
        traceback.print_exc()
        
        # Provide more specific error messages
        error_msg = str(e)
        if 'foreign key' in error_msg.lower() or 'constraint' in error_msg.lower():
            return jsonify({
                'error': 'Cannot delete property: It is referenced by other records (e.g., bookings, inquiries, or reviews). Please remove related data first.'
            }), 400
        elif 'not found' in error_msg.lower():
            return jsonify({'error': 'Property not found'}), 404
        else:
            # Return more informative error message
            return jsonify({
                'error': f'Failed to delete property: {error_msg}'
            }), 500

# User completely deletes their own account (hard delete)
@app.route('/api/user/delete-account', methods=['POST'])
@require_auth
def delete_own_account():
    try:
        # Get current user
        current_user = request.user
        
        # Get user by ID
        user = User.query.get(current_user['user_id'])
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        user_id = user.id
        
        # Delete related data first (to avoid foreign key constraints)
        print(f"🗑️ Deleting all data for user {user_id}...")
        
        # Delete user's bookmarks
        bookmarks_deleted = Bookmark.query.filter_by(user_id=user_id).delete()
        print(f"📚 Deleted {bookmarks_deleted} bookmarks")
        
        # Delete user's price predictions
        predictions_deleted = PricePrediction.query.filter_by(user_id=user_id).delete()
        print(f"📊 Deleted {predictions_deleted} predictions")
        
        # Delete user's agent profile if exists
        agent_profile_deleted = AgentProfile.query.filter_by(user_id=user_id).delete()
        print(f"👤 Deleted {agent_profile_deleted} agent profiles")
        
        # Delete user's agent regions if exists
        agent_regions_deleted = AgentRegion.query.filter_by(agent_id=user_id).delete()
        print(f"🗺️ Deleted {agent_regions_deleted} agent regions")
        
        # Delete user's properties if they're an agent
        properties_deleted = Property.query.filter_by(agent_id=user_id).delete()
        print(f"🏢 Deleted {properties_deleted} properties")
        
        # Delete user's feedback/reviews
        feedback_deleted = BusinessInquiry.query.filter_by(user_id=user_id).delete()
        print(f"💬 Deleted {feedback_deleted} feedback entries")
        
        # Finally delete the user account
        db.session.delete(user)
        db.session.commit()
        
        print(f"✅ User {user_id} completely deleted from database")
        
        return jsonify({
            'message': 'Account completely deleted. You can sign up again with the same email.',
            'user_id': user_id,
            'status': 'deleted',
            'data_removed': {
                'bookmarks': bookmarks_deleted,
                'predictions': predictions_deleted,
                'agent_profile': agent_profile_deleted,
                'agent_regions': agent_regions_deleted,
                'properties': properties_deleted,
                'feedback': feedback_deleted
            }
        }), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"Error deleting account: {e}")
        import traceback
        traceback.print_exc()
        
        # Check for specific constraint violations
        if 'foreign key' in str(e).lower():
            return jsonify({'error': 'Cannot delete account due to existing data. Please contact support.'}), 400
        else:
            return jsonify({'error': 'Failed to delete account'}), 500

# Suspend user account
# Admin: Update user type
@app.route('/api/admin/users/<int:user_id>', methods=['PUT'])
@require_auth
def update_user(user_id):
    """Update user information (e.g., user_type)"""
    try:
        # Get current user to verify admin status
        current_user = request.user
        if current_user['user_type'] != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        # Get user by ID
        user = User.query.get(user_id)
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        # Check if trying to modify own account (admin should be able to, but be careful)
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Request data is required'}), 400
        
        # Update user_type if provided
        if 'user_type' in data:
            new_user_type = data['user_type'].lower()
            
            # Validate user_type
            if not validate_user_type(new_user_type):
                return jsonify({'error': f'Invalid user_type. Must be one of: free, premium, agent, admin'}), 400
            
            # Prevent changing own user_type to non-admin (unless they want to)
            if user.id == current_user['user_id'] and new_user_type != 'admin':
                return jsonify({'error': 'Cannot change your own user type from admin'}), 400
            
            # Update user type
            user.user_type = new_user_type
            db.session.commit()
        
        # Return updated user data
        return jsonify({
            'message': 'User updated successfully',
            'user': {
                'id': user.id,
                'email': user.email,
                'full_name': user.full_name,
                'user_type': user.user_type,
                'subscription_status': user.subscription_status,
                'is_active': user.is_active
            }
        }), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"Error updating user: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to update user: {str(e)}'}), 500

@app.route('/api/admin/users/<int:user_id>/suspend', methods=['POST'])
@require_auth
def suspend_user(user_id):
    try:
        # Get current user to verify admin status
        current_user = request.user
        if current_user['user_type'] != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        # Get user by ID
        user = User.query.get(user_id)
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        # Check if trying to suspend self
        if user.id == current_user['user_id']:
            return jsonify({'error': 'Cannot suspend your own account'}), 400
        
        # Suspend user (set subscription status to cancelled)
        user.subscription_status = 'cancelled'  # Fixed: use valid constraint value
        db.session.commit()
        
        return jsonify({
            'message': 'User suspended successfully',
            'user_id': user_id,
            'status': 'suspended'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"Error suspending user: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Failed to suspend user'}), 500

# Reactivate user account
@app.route('/api/admin/users/<int:user_id>/reactivate', methods=['POST'])
@require_auth
def reactivate_user(user_id):
    try:
        # Get current user to verify admin status
        current_user = request.user
        if current_user['user_type'] != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        # Get user by ID
        user = User.query.get(user_id)
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        # Reactivate user
        user.is_active = True
        user.subscription_status = 'active'  # Fixed: lowercase to match database constraint
        db.session.commit()
        
        return jsonify({
            'message': 'User reactivated successfully',
            'user_id': user_id,
            'status': 'active'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"Error reactivating user: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Failed to reactivate user'}), 500

# Admin completely deletes a user account (hard delete)
@app.route('/api/admin/users/<int:user_id>/delete', methods=['DELETE'])
@require_auth
def admin_delete_user(user_id):
    try:
        # Get current user to verify admin status
        current_user = request.user
        if current_user['user_type'] != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        # Get user by ID
        user = User.query.get(user_id)
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        # Check if trying to delete self
        if user.id == current_user['user_id']:
            return jsonify({'error': 'Cannot delete your own account'}), 400
        
        print(f"🗑️ Admin deleting user {user_id} ({user.email})...")
        
        # Delete related data first (to avoid foreign key constraints)
        bookmarks_deleted = Bookmark.query.filter_by(user_id=user_id).delete()
        predictions_deleted = PricePrediction.query.filter_by(user_id=user_id).delete()
        agent_profile_deleted = AgentProfile.query.filter_by(user_id=user_id).delete()
        agent_regions_deleted = AgentRegion.query.filter_by(agent_id=user_id).delete()
        properties_deleted = Property.query.filter_by(agent_id=user_id).delete()
        feedback_deleted = BusinessInquiry.query.filter_by(user_id=user_id).delete()
        
        # Finally delete the user account
        db.session.delete(user)
        db.session.commit()
        
        print(f"✅ Admin deleted user {user_id} completely from database")
        
        return jsonify({
            'message': 'User account completely deleted by admin',
            'user_id': user_id,
            'deleted_by': current_user['user_id'],
            'data_removed': {
                'bookmarks': bookmarks_deleted,
                'predictions': predictions_deleted,
                'agent_profile': agent_profile_deleted,
                'agent_regions': agent_regions_deleted,
                'properties': properties_deleted,
                'feedback': feedback_deleted
            }
        }), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"Error admin deleting user: {e}")
        import traceback
        traceback.print_exc()
        
        if 'foreign key' in str(e).lower():
            return jsonify({'error': 'Cannot delete user due to existing data. Please contact support.'}), 400
        else:
            return jsonify({'error': 'Failed to delete user'}), 500

# Get all reviews (user_reviews) for admin management
@app.route('/api/admin/reviews', methods=['GET'])
@require_auth
def get_all_reviews():
    """Get all user reviews for admin management with optional ML filtering"""
    try:
        # Get current user to verify admin status
        current_user = request.user
        if current_user['user_type'] != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        # Get all reviews with pagination
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        
        # Get ML filter options
        filter_legit = request.args.get('filter_legit', 'false').lower() == 'true'
        include_ml_predictions = request.args.get('include_ml', 'true').lower() == 'true'
        min_confidence = float(request.args.get('min_confidence', 0.7))
        
        # Query all reviews with user information
        query = text("""
            SELECT r.id, r.review_text, r.rating, r.review_date, r.is_verified,
                   r.admin_response, r.admin_response_date,
                   u.full_name, u.email, u.user_type
            FROM user_reviews r
            JOIN users u ON r.user_id = u.id
            ORDER BY r.review_date DESC
        """)
        
        result = db.session.execute(query)
        all_reviews = []
        
        for row in result:
            all_reviews.append({
                'id': row.id,
                'review_text': row.review_text,
                'rating': row.rating,
                'review_date': row.review_date.isoformat() if row.review_date else None,
                'is_verified': row.is_verified,
                'admin_response': row.admin_response,
                'admin_response_date': row.admin_response_date.isoformat() if row.admin_response_date else None,
                'user_name': row.full_name,
                'user_email': row.email,
                'user_type': row.user_type
            })
        
        # Apply ML filtering if requested and available
        reviews_to_filter = all_reviews
        if ML_FILTER_AVAILABLE and (filter_legit or include_ml_predictions):
            try:
                ml_filter = get_ml_filter()
                if ml_filter and ml_filter.model:
                    legit_reviews, all_with_predictions = ml_filter.filter_legit_reviews(
                        all_reviews, 
                        min_confidence=min_confidence
                    )
                    
                    # Add ML predictions to all reviews if requested
                    if include_ml_predictions:
                        all_reviews = all_with_predictions
                    
                    # Filter to only legit reviews if requested
                    if filter_legit:
                        reviews_to_filter = legit_reviews
                    else:
                        reviews_to_filter = all_with_predictions
            except Exception as ml_error:
                print(f"ML filtering error (continuing without filter): {ml_error}")
                # Continue without ML filtering if there's an error
        
        # Simple pagination
        start_idx = (page - 1) * per_page
        end_idx = start_idx + per_page
        paginated_reviews = reviews_to_filter[start_idx:end_idx]
        
        # Calculate total statistics from all reviews (not just current page)
        # Count by rating
        rating_stats = {
            5: 0,
            4: 0,
            3: 0,
            2: 0,
            1: 0
        }
        for review in reviews_to_filter:
            if review.get('rating') and 1 <= review['rating'] <= 5:
                rating_stats[review['rating']] += 1
        
        return jsonify({
            'reviews': paginated_reviews,
            'total': len(reviews_to_filter),
            'current_page': page,
            'per_page': per_page,
            'ml_enabled': ML_FILTER_AVAILABLE and include_ml_predictions,
            'filtered': filter_legit,
            'statistics': {
                'rating_distribution': rating_stats
            }
        }), 200
        
    except Exception as e:
        print(f"Error getting reviews: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Failed to get reviews'}), 500

# Get all feedback (business inquiries) for admin review
@app.route('/api/admin/feedback', methods=['GET'])
@require_auth
def get_all_feedback():
    try:
        # Get current user to verify admin status
        current_user = request.user
        if current_user['user_type'] != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        # Get all feedback with pagination
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        
        # Get ML filter options
        filter_legit = request.args.get('filter_legit', 'false').lower() == 'true'
        include_ml_predictions = request.args.get('include_ml', 'true').lower() == 'true'
        min_confidence = float(request.args.get('min_confidence', 0.7))
        
        # Query all business inquiries (feedback) with user information
        # First check if admin_response columns exist
        try:
            # Try to check if columns exist
            check_query = text("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name = 'business_inquiries' AND column_name IN ('admin_response', 'admin_response_date')
            """)
            existing_columns = [row[0] for row in db.session.execute(check_query).fetchall()]
            has_admin_response = 'admin_response' in existing_columns
            has_admin_response_date = 'admin_response_date' in existing_columns
            
            # Build query based on column existence
            if has_admin_response and has_admin_response_date:
                query = text("""
                    SELECT bi.id, bi.message, bi.inquiry_type, bi.status, bi.created_at,
                           bi.name, bi.email, bi.phone,
                           bi.admin_response, bi.admin_response_date,
                           u.full_name, u.email as user_email, u.user_type
                    FROM business_inquiries bi
                    LEFT JOIN users u ON bi.user_id = u.id
                    ORDER BY bi.created_at DESC
                """)
            else:
                # Columns don't exist yet, query without them
                query = text("""
                    SELECT bi.id, bi.message, bi.inquiry_type, bi.status, bi.created_at,
                           bi.name, bi.email, bi.phone,
                           u.full_name, u.email as user_email, u.user_type
                    FROM business_inquiries bi
                    LEFT JOIN users u ON bi.user_id = u.id
                    ORDER BY bi.created_at DESC
                """)
        except:
            # If check fails, query without admin_response columns
            query = text("""
                SELECT bi.id, bi.message, bi.inquiry_type, bi.status, bi.created_at,
                       bi.name, bi.email, bi.phone,
                       u.full_name, u.email as user_email, u.user_type
                FROM business_inquiries bi
                LEFT JOIN users u ON bi.user_id = u.id
                ORDER BY bi.created_at DESC
            """)
            has_admin_response = False
            has_admin_response_date = False
        
        result = db.session.execute(query)
        all_feedbacks = []
        
        for row in result:
            # Use user's name/email if available, otherwise use inquiry name/email
            user_name = row.full_name if row.full_name else row.name
            user_email = row.user_email if row.user_email else row.email
            
            # Try to get admin_response if columns exist and are in result
            admin_response = None
            admin_response_date = None
            try:
                if has_admin_response and hasattr(row, 'admin_response'):
                    admin_response = row.admin_response  # Include even if None or empty string
                if has_admin_response_date and hasattr(row, 'admin_response_date') and row.admin_response_date:
                    admin_response_date = row.admin_response_date.isoformat()
            except:
                pass
            
            all_feedbacks.append({
                'id': row.id,
                'message': row.message,
                'inquiry_type': row.inquiry_type,
                'status': row.status,
                'created_at': row.created_at.isoformat() if row.created_at else None,
                'user_name': user_name,
                'user_email': user_email,
                'phone': row.phone,
                'user_type': row.user_type if row.user_type else 'guest',
                'admin_response': admin_response,
                'admin_response_date': admin_response_date
            })
        
        # Apply ML filtering if requested and available
        feedbacks_to_filter = all_feedbacks
        if ML_FILTER_AVAILABLE and (filter_legit or include_ml_predictions):
            try:
                ml_filter = get_ml_filter()
                if ml_filter and ml_filter.model:
                    legit_feedbacks, all_with_predictions = ml_filter.filter_legit_reviews(
                        all_feedbacks, 
                        min_confidence=min_confidence
                    )
                    
                    # Add ML predictions to all feedbacks if requested
                    if include_ml_predictions:
                        all_feedbacks = all_with_predictions
                    
                    # Filter to only legit feedbacks if requested
                    if filter_legit:
                        feedbacks_to_filter = legit_feedbacks
                    else:
                        feedbacks_to_filter = all_with_predictions
            except Exception as ml_error:
                print(f"ML filtering error (continuing without filter): {ml_error}")
                # Continue without ML filtering if there's an error
        
        # Simple pagination
        start_idx = (page - 1) * per_page
        end_idx = start_idx + per_page
        paginated_feedbacks = feedbacks_to_filter[start_idx:end_idx]
        
        # Calculate total statistics from all feedback (not just current page)
        total_responded = 0
        total_not_responded = 0
        for feedback in feedbacks_to_filter:
            if feedback.get('admin_response'):
                total_responded += 1
            else:
                total_not_responded += 1
        
        return jsonify({
            'feedback': paginated_feedbacks,
            'total': len(feedbacks_to_filter),
            'current_page': page,
            'per_page': per_page,
            'ml_enabled': ML_FILTER_AVAILABLE and include_ml_predictions,
            'filtered': filter_legit,
            'statistics': {
                'total_responded': total_responded,
                'total_not_responded': total_not_responded
            }
        }), 200
        
    except Exception as e:
        print(f"Error getting feedback: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Failed to get feedback'}), 500

# Get all feedback form types (for user feedback form)
@app.route('/api/feedback/form-types', methods=['GET'])
def get_feedback_form_types():
    """Get all active feedback form types for the user feedback form"""
    try:
        from sqlalchemy import text
        # Check if table exists first
        try:
            query = text("""
                SELECT id, name, value, status, display_order
                FROM feedback_form_types
                WHERE status = 'active'
                ORDER BY display_order ASC, name ASC
            """)
            result = db.session.execute(query)
            types = []
            for row in result:
                types.append({
                    'id': row.id,
                    'name': row.name,
                    'value': row.value,
                    'status': row.status,
                    'display_order': row.display_order
                })
            
            # If no types found in database, return default types
            if not types:
                return jsonify({
                    'types': [
                        {'id': 1, 'name': 'General Feedback', 'value': 'general', 'status': 'active', 'display_order': 0},
                        {'id': 2, 'name': 'Support Request', 'value': 'support', 'status': 'active', 'display_order': 1},
                        {'id': 3, 'name': 'Property Viewing Inquiry', 'value': 'property_viewing', 'status': 'active', 'display_order': 2},
                        {'id': 4, 'name': 'Price Quote Request', 'value': 'price_quote', 'status': 'active', 'display_order': 3}
                    ]
                }), 200
            
            return jsonify({'types': types}), 200
        except Exception as table_error:
            # Table doesn't exist yet, return default types
            print(f"Feedback form types table not found: {table_error}")
            return jsonify({
                'types': [
                    {'id': 1, 'name': 'General Feedback', 'value': 'general', 'status': 'active', 'display_order': 0},
                    {'id': 2, 'name': 'Support Request', 'value': 'support', 'status': 'active', 'display_order': 1},
                    {'id': 3, 'name': 'Property Viewing Inquiry', 'value': 'property_viewing', 'status': 'active', 'display_order': 2},
                    {'id': 4, 'name': 'Price Quote Request', 'value': 'price_quote', 'status': 'active', 'display_order': 3}
                ]
            }), 200
    except Exception as e:
        print(f"Error getting feedback form types: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Failed to get feedback form types'}), 500

# Admin: Get all feedback form types (including inactive)
@app.route('/api/admin/feedback/form-types', methods=['GET'])
@require_auth
def get_all_feedback_form_types():
    """Get all feedback form types for admin management"""
    try:
        current_user = request.user
        if current_user['user_type'] != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        from sqlalchemy import text
        try:
            query = text("""
                SELECT id, name, value, status, display_order, created_at, updated_at
                FROM feedback_form_types
                ORDER BY display_order ASC, name ASC
            """)
            result = db.session.execute(query)
            types = []
            for row in result:
                types.append({
                    'id': row.id,
                    'name': row.name,
                    'value': row.value,
                    'status': row.status,
                    'display_order': row.display_order,
                    'created_at': row.created_at.isoformat() if row.created_at else None,
                    'updated_at': row.updated_at.isoformat() if row.updated_at else None
                })
            
            # If no types found, return empty array
            return jsonify({'types': types}), 200
        except Exception as table_error:
            # Table doesn't exist yet, return empty array
            print(f"Feedback form types table not found: {table_error}")
            return jsonify({'types': []}), 200
    except Exception as e:
        print(f"Error getting all feedback form types: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Failed to get feedback form types'}), 500

# Admin: Create feedback form type
@app.route('/api/admin/feedback/form-types', methods=['POST'])
@require_auth
def create_feedback_form_type():
    """Create a new feedback form type"""
    try:
        current_user = request.user
        if current_user['user_type'] != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        data = request.get_json()
        if not data or not data.get('name') or not data.get('value'):
            return jsonify({'error': 'Name and value are required'}), 400
        
        from sqlalchemy import text
        
        # Check if table exists, if not create it
        try:
            check_query = text("""
                SELECT EXISTS (
                    SELECT FROM information_schema.tables 
                    WHERE table_name = 'feedback_form_types'
                )
            """)
            table_exists = db.session.execute(check_query).scalar()
            
            if not table_exists:
                # Create the table
                create_table_query = text("""
                    CREATE TABLE feedback_form_types (
                        id SERIAL PRIMARY KEY,
                        name VARCHAR(100) NOT NULL,
                        value VARCHAR(50) UNIQUE NOT NULL,
                        status VARCHAR(20) DEFAULT 'active',
                        display_order INTEGER DEFAULT 0,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                """)
                db.session.execute(create_table_query)
                db.session.commit()
                print("Created feedback_form_types table")
        except Exception as check_error:
            print(f"Error checking/creating table: {check_error}")
            db.session.rollback()
        
        # Get max display_order
        max_order_query = text("SELECT COALESCE(MAX(display_order), 0) FROM feedback_form_types")
        max_order = db.session.execute(max_order_query).scalar() or 0
        
        # Insert new type
        insert_query = text("""
            INSERT INTO feedback_form_types (name, value, status, display_order)
            VALUES (:name, :value, :status, :display_order)
            RETURNING id, name, value, status, display_order, created_at, updated_at
        """)
        result = db.session.execute(insert_query, {
            'name': data['name'],
            'value': data['value'],
            'status': data.get('status', 'active'),
            'display_order': data.get('display_order', max_order + 1)
        })
        new_type = result.fetchone()
        db.session.commit()
        
        return jsonify({
            'type': {
                'id': new_type.id,
                'name': new_type.name,
                'value': new_type.value,
                'status': new_type.status,
                'display_order': new_type.display_order,
                'created_at': new_type.created_at.isoformat() if new_type.created_at else None,
                'updated_at': new_type.updated_at.isoformat() if new_type.updated_at else None
            }
        }), 201
    except Exception as e:
        db.session.rollback()
        print(f"Error creating feedback form type: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to create feedback form type: {str(e)}'}), 500

# Admin: Update feedback form type
@app.route('/api/admin/feedback/form-types/<int:type_id>', methods=['PUT'])
@require_auth
def update_feedback_form_type(type_id):
    """Update a feedback form type"""
    try:
        current_user = request.user
        if current_user['user_type'] != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Request data is required'}), 400
        
        from sqlalchemy import text
        
        # Build update query dynamically
        update_fields = []
        params = {'type_id': type_id}
        
        if 'name' in data:
            update_fields.append('name = :name')
            params['name'] = data['name']
        if 'value' in data:
            update_fields.append('value = :value')
            params['value'] = data['value']
        if 'status' in data:
            update_fields.append('status = :status')
            params['status'] = data['status']
        if 'display_order' in data:
            update_fields.append('display_order = :display_order')
            params['display_order'] = data['display_order']
        
        if not update_fields:
            return jsonify({'error': 'No fields to update'}), 400
        
        update_fields.append('updated_at = CURRENT_TIMESTAMP')
        
        update_query = text(f"""
            UPDATE feedback_form_types
            SET {', '.join(update_fields)}
            WHERE id = :type_id
            RETURNING id, name, value, status, display_order, created_at, updated_at
        """)
        
        result = db.session.execute(update_query, params)
        updated_type = result.fetchone()
        
        if not updated_type:
            return jsonify({'error': 'Feedback form type not found'}), 404
        
        db.session.commit()
        
        return jsonify({
            'type': {
                'id': updated_type.id,
                'name': updated_type.name,
                'value': updated_type.value,
                'status': updated_type.status,
                'display_order': updated_type.display_order,
                'created_at': updated_type.created_at.isoformat() if updated_type.created_at else None,
                'updated_at': updated_type.updated_at.isoformat() if updated_type.updated_at else None
            }
        }), 200
    except Exception as e:
        db.session.rollback()
        print(f"Error updating feedback form type: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to update feedback form type: {str(e)}'}), 500

# Admin: Delete feedback form type
@app.route('/api/admin/feedback/form-types/<int:type_id>', methods=['DELETE'])
@require_auth
def delete_feedback_form_type(type_id):
    """Delete a feedback form type"""
    try:
        current_user = request.user
        if current_user['user_type'] != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        from sqlalchemy import text
        
        delete_query = text("""
            DELETE FROM feedback_form_types
            WHERE id = :type_id
            RETURNING id
        """)
        
        result = db.session.execute(delete_query, {'type_id': type_id})
        deleted = result.fetchone()
        
        if not deleted:
            return jsonify({'error': 'Feedback form type not found'}), 404
        
        db.session.commit()
        return jsonify({'message': 'Feedback form type deleted successfully'}), 200
    except Exception as e:
        db.session.rollback()
        print(f"Error deleting feedback form type: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to delete feedback form type: {str(e)}'}), 500

# Verify/publish review (make it public)
@app.route('/api/admin/feedback/verify', methods=['POST'])
@require_auth
def verify_feedback():
    """Verify/publish a review (works for user_reviews table)"""
    try:
        # Get current user to verify admin status
        current_user = request.user
        if current_user['user_type'] != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        data = request.get_json()
        if not data or not data.get('feedback_id'):
            return jsonify({'error': 'Review ID is required'}), 400
        
        review_id = data['feedback_id']
        
        # Update the review verification status (make it public)
        query = text("""
            UPDATE user_reviews 
            SET is_verified = TRUE
            WHERE id = :id
        """)
        db.session.execute(query, {
            'id': review_id
        })
        db.session.commit()
        
        return jsonify({'message': 'Review verified and published successfully'}), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"Error verifying feedback: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Failed to verify feedback'}), 500

# Respond to feedback/review (send message to user)
@app.route('/api/admin/feedback/respond', methods=['POST'])
@require_auth
def respond_to_feedback():
    """Respond to either a review (user_reviews) or feedback (business_inquiries)"""
    try:
        # Get current user to verify admin status
        current_user = request.user
        if current_user['user_type'] != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        data = request.get_json()
        if not data or not data.get('feedback_id'):
            return jsonify({'error': 'Feedback ID is required'}), 400
        
        feedback_id = data['feedback_id']
        admin_response = data.get('admin_response', '')
        feedback_type = data.get('type', 'review')  # Default to 'review' for backward compatibility
        
        if not admin_response.strip():
            return jsonify({'error': 'Response message is required'}), 400
        
        if feedback_type == 'review':
            # Add admin response to the review (user_reviews table)
            query = text("""
                UPDATE user_reviews 
                SET admin_response = :admin_response,
                    admin_response_date = NOW()
                WHERE id = :id
            """)
            db.session.execute(query, {
                'id': feedback_id,
                'admin_response': admin_response
            })
        else:
            # Add admin response to the feedback (business_inquiries table)
            # First check if admin_response column exists
            try:
                # Check if columns exist
                check_query = text("""
                    SELECT column_name 
                    FROM information_schema.columns 
                    WHERE table_name = 'business_inquiries' AND column_name = 'admin_response'
                """)
                result = db.session.execute(check_query).fetchone()
                
                # If columns don't exist, add them
                if not result:
                    try:
                        db.session.execute(text("ALTER TABLE business_inquiries ADD COLUMN admin_response TEXT"))
                        db.session.execute(text("ALTER TABLE business_inquiries ADD COLUMN admin_response_date TIMESTAMP"))
                        db.session.commit()
                        print("Added admin_response columns to business_inquiries table")
                    except Exception as alter_error:
                        print(f"Could not add admin_response columns: {alter_error}")
                        db.session.rollback()
                
                # Update with admin_response - set status to 'resolved' to indicate it's been responded to
                query = text("""
                    UPDATE business_inquiries 
                    SET admin_response = :admin_response,
                        admin_response_date = NOW(),
                        status = 'resolved'
                    WHERE id = :id
                """)
                db.session.execute(query, {
                    'id': feedback_id,
                    'admin_response': admin_response
                })
            except Exception as e:
                print(f"Error updating business_inquiries: {e}")
                raise
        
        db.session.commit()
        
        return jsonify({'message': 'Response sent successfully'}), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"Error responding to feedback: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Failed to send response'}), 500

# Get feedback/review by ID (supports both user_reviews and business_inquiries)
@app.route('/api/admin/feedback/<int:feedback_id>', methods=['GET'])
@require_auth
def get_feedback_by_id(feedback_id):
    try:
        current_user = request.user
        if current_user['user_type'] != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        # Check query parameter to determine type
        feedback_type = request.args.get('type', 'review')  # Default to 'review' for backward compatibility
        
        if feedback_type == 'review':
            # Get review details from user_reviews table
            query = text("""
                SELECT 
                    ur.id,
                    ur.review_type,
                    ur.rating,
                    ur.review_text,
                    ur.review_date,
                    ur.is_verified,
                    ur.admin_response,
                    ur.admin_response_date,
                    u.full_name as user_name,
                    u.email as user_email
                FROM user_reviews ur
                JOIN users u ON ur.user_id = u.id
                WHERE ur.id = :feedback_id
            """)
            
            result = db.session.execute(query, {'feedback_id': feedback_id})
            feedback = result.fetchone()
            
            if not feedback:
                return jsonify({'error': 'Review not found'}), 404
            
            # Convert to dictionary
            feedback_dict = {
                'id': feedback.id,
                'review_type': feedback.review_type,
                'rating': feedback.rating,
                'review_text': feedback.review_text,
                'review_date': feedback.review_date.isoformat() if feedback.review_date else None,
                'created_at': feedback.review_date.isoformat() if feedback.review_date else None,
                'is_verified': feedback.is_verified,
                'admin_response': feedback.admin_response,
                'admin_response_date': feedback.admin_response_date.isoformat() if feedback.admin_response_date else None,
                'user_name': feedback.user_name,
                'user_email': feedback.user_email,
                'type': 'review'
            }
        else:
            # Get feedback details from business_inquiries table
            # Check if admin_response columns exist first
            has_admin_response_col = False
            has_admin_response_date_col = False
            
            try:
                check_query = text("""
                    SELECT column_name 
                    FROM information_schema.columns 
                    WHERE table_name = 'business_inquiries' AND column_name IN ('admin_response', 'admin_response_date')
                """)
                existing_columns = [row[0] for row in db.session.execute(check_query).fetchall()]
                has_admin_response_col = 'admin_response' in existing_columns
                has_admin_response_date_col = 'admin_response_date' in existing_columns
                
                if has_admin_response_col and has_admin_response_date_col:
                    query = text("""
                        SELECT 
                            bi.id,
                            bi.inquiry_type,
                            bi.message,
                            bi.status,
                            bi.created_at,
                            bi.name,
                            bi.email,
                            bi.phone,
                            u.full_name as user_name,
                            u.email as user_email,
                            bi.admin_response,
                            bi.admin_response_date
                        FROM business_inquiries bi
                        LEFT JOIN users u ON bi.user_id = u.id
                        WHERE bi.id = :feedback_id
                    """)
                else:
                    query = text("""
                        SELECT 
                            bi.id,
                            bi.inquiry_type,
                            bi.message,
                            bi.status,
                            bi.created_at,
                            bi.name,
                            bi.email,
                            bi.phone,
                            u.full_name as user_name,
                            u.email as user_email
                        FROM business_inquiries bi
                        LEFT JOIN users u ON bi.user_id = u.id
                        WHERE bi.id = :feedback_id
                    """)
            except:
                # If check fails, query without admin_response columns
                query = text("""
                    SELECT 
                        bi.id,
                        bi.inquiry_type,
                        bi.message,
                        bi.status,
                        bi.created_at,
                        bi.name,
                        bi.email,
                        bi.phone,
                        u.full_name as user_name,
                        u.email as user_email
                    FROM business_inquiries bi
                    LEFT JOIN users u ON bi.user_id = u.id
                    WHERE bi.id = :feedback_id
                """)
            
            result = db.session.execute(query, {'feedback_id': feedback_id})
            feedback = result.fetchone()
            
            if not feedback:
                return jsonify({'error': 'Feedback not found'}), 404
            
            # Use user's name/email if available, otherwise use inquiry name/email
            user_name = feedback.user_name if feedback.user_name else feedback.name
            user_email = feedback.user_email if feedback.user_email else feedback.email
            
            # Try to get admin_response and admin_response_date from the result
            admin_response = None
            admin_response_date = None
            try:
                # Try accessing as attribute (if column exists in SELECT)
                if has_admin_response_col and hasattr(feedback, 'admin_response') and feedback.admin_response:
                    admin_response = feedback.admin_response
                if has_admin_response_date_col and hasattr(feedback, 'admin_response_date') and feedback.admin_response_date:
                    admin_response_date = feedback.admin_response_date.isoformat()
            except:
                # Column might not exist in database
                pass
            
            # Convert to dictionary
            feedback_dict = {
                'id': feedback.id,
                'inquiry_type': feedback.inquiry_type,
                'message': feedback.message,
                'review_text': feedback.message,  # For compatibility with ViewFeedback component
                'status': feedback.status,
                'created_at': feedback.created_at.isoformat() if feedback.created_at else None,
                'review_date': feedback.created_at.isoformat() if feedback.created_at else None,
                'user_name': user_name,
                'user_email': user_email,
                'phone': feedback.phone,
                'rating': None,  # Business inquiries don't have ratings
                'is_verified': False,  # Business inquiries don't have verification
                'admin_response': admin_response,
                'admin_response_date': admin_response_date,
                'type': 'feedback'
        }
        
        return jsonify(feedback_dict), 200
        
    except Exception as e:
        print(f"Error getting feedback by ID: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Failed to get feedback'}), 500

# Unverify feedback (remove from homepage)
@app.route('/api/admin/feedback/unverify', methods=['POST'])
@require_auth
def unverify_feedback():
    try:
        current_user = request.user
        if current_user['user_type'] != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        data = request.get_json()
        if not data or not data.get('feedback_id'):
            return jsonify({'error': 'Feedback ID is required'}), 400
        
        feedback_id = data['feedback_id']
        
        # Remove verification status (make it private again)
        query = text("""
            UPDATE user_reviews 
            SET is_verified = FALSE
            WHERE id = :id
        """)
        db.session.execute(query, {'id': feedback_id})
        db.session.commit()
        
        return jsonify({'message': 'Feedback removed from homepage successfully'}), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"Error unverifying feedback: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Failed to remove feedback from homepage'}), 500

# Add missing columns to user_reviews table
@app.route('/api/db/add-missing-columns', methods=['POST'])
def add_missing_columns():
    try:
        # Check if columns exist
        check_query = text("""
            SELECT column_name 
            FROM information_schema.columns 
            WHERE table_name = 'user_reviews' 
            AND column_name IN ('verified_date', 'verified_by')
        """)
        
        result = db.session.execute(check_query)
        existing_columns = [row[0] for row in result]
        
        # Add verified_date column if it doesn't exist
        if 'verified_date' not in existing_columns:
            db.session.execute(text("ALTER TABLE user_reviews ADD COLUMN verified_date TIMESTAMP"))
            print("Added verified_date column")
        
        # Add verified_by column if it doesn't exist
        if 'verified_by' not in existing_columns:
            db.session.execute(text("ALTER TABLE user_reviews ADD COLUMN verified_by INTEGER REFERENCES users(id)"))
            print("Added verified_by column")
        
        db.session.commit()
        
        return jsonify({
            'message': 'Missing columns added successfully',
            'added_columns': [col for col in ['verified_date', 'verified_by'] if col not in existing_columns]
        }), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"Error adding missing columns: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Failed to add missing columns'}), 500

# Test database connection
@app.route('/api/test-db')
def test_db():
    try:
        # Test basic database operations
        user_count = User.query.count()
        return jsonify({
            'message': 'Database connection successful',
            'user_count': user_count,
            'tables': ['users', 'user_reviews', 'properties']
        }), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Database connection failed: {str(e)}'}), 500

# Create database tables
@app.route('/api/init-db')
def init_db():
    try:
        db.create_all()
        return jsonify({'message': 'Database tables created successfully'}), 200
    except Exception as e:
        return jsonify({'error': f'Failed to create tables: {str(e)}'}), 500

# Bookmark endpoints
@app.route('/api/bookmarks', methods=['GET'])
@require_auth
def get_user_bookmarks():
    """Get all bookmarks for the current user"""
    try:
        current_user = request.user
        
        # Get bookmarks directly from bookmarks table (simpler and more reliable)
        query = text("""
            SELECT 
                b.id,
                b.bookmark_type,
                b.reference_id,
                b.address,
                b.floor_area,
                b.level,
                b.unit_number,
                b.property_type,
                b.address_2,
                b.floor_area_2,
                b.level_2,
                b.unit_number_2,
                b.property_type_2,
                b.created_at
            FROM bookmarks b
            WHERE b.user_id = :user_id
            ORDER BY b.created_at DESC
        """)
        
        result = db.session.execute(query, {'user_id': current_user['user_id']})
        bookmarks = result.fetchall()
        
        # Group bookmarks by type
        bookmarked_addresses = []
        bookmarked_predictions = []
        bookmarked_comparisons = []
        
        for bookmark in bookmarks:
            if bookmark.bookmark_type == 'comparison':
                # For comparison bookmarks, include both properties
                bookmark_data = {
                    'id': bookmark.id,
                    'address': bookmark.address or 'Address not available',
                    'floor_area': bookmark.floor_area,
                    'level': bookmark.level,
                    'unit_number': bookmark.unit_number,
                    'property_type': bookmark.property_type or 'Property',
                    'address_2': bookmark.address_2 or 'Second address not available',
                    'floor_area_2': bookmark.floor_area_2,
                    'level_2': bookmark.level_2,
                    'unit_number_2': bookmark.unit_number_2,
                    'property_type_2': bookmark.property_type_2 or 'Property',
                    'bookmarked_date': bookmark.created_at.isoformat() if bookmark.created_at else None
                }
            else:
                # For property and prediction bookmarks
                bookmark_data = {
                    'id': bookmark.id,
                    'address': bookmark.address or 'Address not available',
                    'floor_area': bookmark.floor_area,
                    'level': bookmark.level,
                    'unit_number': bookmark.unit_number,
                    'property_type': bookmark.property_type or 'Property',
                    'bookmarked_date': bookmark.created_at.isoformat() if bookmark.created_at else None
                }
            
            if bookmark.bookmark_type == 'property':
                bookmarked_addresses.append(bookmark_data)
            elif bookmark.bookmark_type == 'prediction':
                bookmarked_predictions.append(bookmark_data)
            elif bookmark.bookmark_type == 'comparison':
                bookmarked_comparisons.append(bookmark_data)
        
        return jsonify({
            'bookmarked_addresses': bookmarked_addresses,
            'bookmarked_predictions': bookmarked_predictions,
            'bookmarked_comparisons': bookmarked_comparisons
        }), 200
        
    except Exception as e:
        print(f"Error getting bookmarks: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Failed to get bookmarks'}), 500

@app.route('/api/bookmarks', methods=['POST'])
@require_auth
def create_bookmark():
    """Create a new bookmark"""
    try:
        current_user = request.user
        data = request.get_json()
        
        # Validate required fields
        if not data.get('bookmark_type') or not data.get('address'):
            return jsonify({'error': 'Bookmark type and address are required'}), 400
        
        # Check if bookmark already exists for this user and address
        existing_bookmark = Bookmark.query.filter_by(
            user_id=current_user['user_id'],
            bookmark_type=data['bookmark_type'],
            address=data['address']
        ).first()
        
        if existing_bookmark:
            return jsonify({'error': 'Bookmark already exists'}), 400
        
        # Create new bookmark
        bookmark = Bookmark(
            user_id=current_user['user_id'],
            bookmark_type=data['bookmark_type'],
            reference_id=int(datetime.utcnow().timestamp()),  # Use timestamp as reference
            address=data.get('address'),
            floor_area=data.get('floor_area'),
            level=data.get('level'),
            unit_number=data.get('unit_number'),
            property_type=data.get('property_type'),
            address_2=data.get('address_2'),
            floor_area_2=data.get('floor_area_2'),
            level_2=data.get('level_2'),
            unit_number_2=data.get('unit_number_2'),
            property_type_2=data.get('property_type_2')
        )
        
        db.session.add(bookmark)
        db.session.commit()
        
        return jsonify({
            'message': 'Bookmark created successfully',
            'bookmark_id': bookmark.id
        }), 201
        
    except Exception as e:
        db.session.rollback()
        print(f"Error creating bookmark: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Failed to create bookmark'}), 500

@app.route('/api/bookmarks/<int:bookmark_id>', methods=['DELETE'])
@require_auth
def delete_bookmark(bookmark_id):
    """Delete a bookmark"""
    try:
        current_user = request.user
        
        # Get bookmark and verify ownership
        bookmark = Bookmark.query.filter_by(id=bookmark_id, user_id=current_user['user_id']).first()
        if not bookmark:
            return jsonify({'error': 'Bookmark not found'}), 404
        
        db.session.delete(bookmark)
        db.session.commit()
        
        return jsonify({'message': 'Bookmark deleted successfully'}), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"Error deleting bookmark: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Failed to delete bookmark'}), 500

# First-time user onboarding endpoints
@app.route('/api/onboarding/complete', methods=['POST'])
@require_auth
def complete_user_onboarding():
    """Mark user as having completed first-time onboarding"""
    try:
        current_user = request.user
        user_id = current_user['user_id']
        
        # Update user's first_time_user status
        user = User.query.get(user_id)
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        user.first_time_user = False
        db.session.commit()
        
        print(f"Successfully completed user onboarding for user {user_id}")
        return jsonify({'message': 'Onboarding completed successfully'}), 200
        
    except Exception as e:
        db.session.rollback()
        import traceback
        print(f"Error completing user onboarding: {e}")
        traceback.print_exc()
        return jsonify({'error': f'Failed to complete onboarding: {str(e)}'}), 500

@app.route('/api/onboarding/complete-agent', methods=['POST'])
@require_auth
def complete_agent_onboarding():
    """Mark agent as having completed first-time region selection"""
    try:
        current_user = request.user
        user_id = current_user['user_id']
        
        # Verify user is an agent
        user = User.query.get(user_id)
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        if user.user_type != 'agent':
            return jsonify({'error': 'User is not an agent'}), 403
        
        # Get or create agent profile
        agent_profile = AgentProfile.query.filter_by(user_id=user_id).first()
        if not agent_profile:
            # Create agent profile if it doesn't exist
            print(f"Agent profile not found for user {user_id}, creating one...")
            agent_profile = AgentProfile(
                user_id=user_id,
                first_time_agent=False
            )
            db.session.add(agent_profile)
        
        # Update agent profile's first_time_agent status
        agent_profile.first_time_agent = False
        db.session.commit()
        
        print(f"Successfully completed agent onboarding for user {user_id}")
        return jsonify({'message': 'Agent onboarding completed successfully'}), 200
        
    except Exception as e:
        db.session.rollback()
        import traceback
        print(f"Error completing agent onboarding: {e}")
        traceback.print_exc()
        return jsonify({'error': f'Failed to complete agent onboarding: {str(e)}'}), 500

@app.route('/api/onboarding/agent-status', methods=['GET'])
@require_auth
def get_agent_status():
    """Get agent's first-time status"""
    try:
        current_user = request.user
        
        # Verify user is an agent
        user = User.query.get(current_user['user_id'])
        if not user or user.user_type != 'agent':
            return jsonify({'error': 'User is not an agent'}), 403
        
        # Get agent profile's first_time_agent status
        agent_profile = AgentProfile.query.filter_by(user_id=current_user['user_id']).first()
        if agent_profile:
            return jsonify({
                'first_time_agent': agent_profile.first_time_agent,
                'user_id': current_user['user_id']
            }), 200
        else:
            # Create agent profile if it doesn't exist (for existing agents)
            agent_profile = AgentProfile(
                user_id=current_user['user_id'],
                first_time_agent=False  # Assume existing agents have completed onboarding
            )
            db.session.add(agent_profile)
            db.session.commit()
            
            return jsonify({
                'first_time_agent': agent_profile.first_time_agent,
                'user_id': current_user['user_id']
            }), 200
        
    except Exception as e:
        print(f"Error getting agent status: {e}")
        return jsonify({'error': 'Failed to get agent status'}), 500

@app.route('/api/onboarding/user-status', methods=['GET'])
@require_auth
def get_user_status():
    """Get user's first-time status"""
    try:
        current_user = request.user
        
        # Get user
        user = User.query.get(current_user['user_id'])
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        return jsonify({
            'first_time_user': user.first_time_user,
            'user_type': user.user_type,
            'user_id': current_user['user_id']
        }), 200
        
    except Exception as e:
        print(f"Error getting user status: {e}")
        return jsonify({'error': 'Failed to get user status'}), 500

# Agent Dashboard API endpoints
@app.route('/api/agent/dashboard-stats', methods=['GET'])
@require_auth
def get_agent_dashboard_stats():
    """Get agent dashboard statistics"""
    try:
        current_user = request.user
        print(f"Getting dashboard stats for user: {current_user['user_id']}")
        
        # Verify user is an agent
        user = User.query.get(current_user['user_id'])
        if not user or user.user_type != 'agent':
            return jsonify({'error': 'User is not an agent'}), 403
        
        # Get agent's active listings count
        try:
            active_listings = Property.query.filter_by(
                agent_id=current_user['user_id'], 
                status='active'
            ).count()
        except Exception as e:
            print(f"Error counting active listings: {e}")
            active_listings = 0
        
        # Get agent's assigned regions count
        try:
            assigned_regions = db.session.query(db.func.count(AgentRegion.id)).filter_by(
                agent_id=current_user['user_id']
            ).scalar() or 0
        except Exception as e:
            print(f"Error counting assigned regions: {e}")
            assigned_regions = 0
        
        # Get total inquiries for agent's properties (simplified)
        try:
            total_inquiries = BusinessInquiry.query.filter_by(
                assigned_to=current_user['user_id']
            ).count()
        except Exception as e:
            print(f"Error counting inquiries: {e}")
            total_inquiries = 0
        
        # Get total properties count for this agent
        try:
            total_properties = Property.query.filter_by(agent_id=current_user['user_id']).count()
        except Exception as e:
            print(f"Error counting total properties: {e}")
            total_properties = 0
        
        # Get total listing views for agent's properties (last 30 days)
        try:
            thirty_days_ago = datetime.utcnow() - timedelta(days=30)
            total_views = db.session.query(db.func.count(PropertyView.id)).join(
                Property, PropertyView.property_id == Property.id
            ).filter(
                Property.agent_id == current_user['user_id'],
                PropertyView.viewed_at >= thirty_days_ago
            ).scalar() or 0
        except Exception as e:
            print(f"Error counting property views (table might not exist yet): {e}")
            total_views = 0  # Fallback to 0 if table doesn't exist
        
        return jsonify({
            'active_listings': active_listings,
            'total_inquiries': total_inquiries,
            'total_properties': total_properties,
            'listing_views': total_views,
            'assigned_regions': assigned_regions,
            'max_regions': 3
        }), 200
        
    except Exception as e:
        print(f"Error getting agent dashboard stats: {e}")
        return jsonify({'error': 'Failed to get dashboard stats'}), 500

@app.route('/api/agent/regions', methods=['GET'])
@require_auth
def get_agent_regions():
    """Get agent's assigned regions"""
    try:
        current_user = request.user
        
        # Verify user is an agent
        user = User.query.get(current_user['user_id'])
        if not user or user.user_type != 'agent':
            return jsonify({'error': 'User is not an agent'}), 403
        
        # Get agent's regions
        regions = AgentRegion.query.filter_by(agent_id=current_user['user_id']).all()
        
        regions_data = []
        for region in regions:
            regions_data.append({
                'id': region.id,
                'region_name': region.region_name,
                'region_type': region.region_type,
                'region_value': region.region_value
            })
        
        return jsonify({'regions': regions_data}), 200
        
    except Exception as e:
        print(f"Error getting agent regions: {e}")
        return jsonify({'error': 'Failed to get agent regions'}), 500

@app.route('/api/agent/regions', methods=['POST'])
@require_auth
def update_agent_regions():
    """Update agent's assigned regions"""
    try:
        current_user = request.user
        
        # Verify user is an agent
        user = User.query.get(current_user['user_id'])
        if not user or user.user_type != 'agent':
            return jsonify({'error': 'User is not an agent'}), 403
        
        data = request.get_json()
        selected_regions = data.get('regions', [])
        
        if not selected_regions:
            return jsonify({'error': 'At least one region must be selected'}), 400
        
        if len(selected_regions) > 3:
            return jsonify({'error': 'Maximum of 3 regions allowed'}), 400
        
        # Delete existing regions for this agent
        AgentRegion.query.filter_by(agent_id=current_user['user_id']).delete()
        
        # Convert to integers to ensure type consistency
        selected_regions = [int(r) for r in selected_regions]
        
        # Fetch selected regions from database
        regions = Region.query.filter(Region.id.in_(selected_regions)).all()
        
        # Debug logging
        print(f"DEBUG: Selected region IDs: {selected_regions}")
        print(f"DEBUG: Found regions: {[(r.id, r.district, r.location, r.is_active) for r in regions]}")
        
        # Check if all regions exist
        if len(regions) != len(selected_regions):
            # Get the IDs that were not found
            found_ids = [r.id for r in regions]
            missing_ids = [r for r in selected_regions if r not in found_ids]
            print(f"DEBUG: Missing region IDs: {missing_ids}")
            return jsonify({
                'error': f'Invalid region IDs: {missing_ids}. Expected {len(selected_regions)} regions but found {len(regions)}'
            }), 400
        
        if not regions:
            return jsonify({'error': 'No valid regions found'}), 400
        
        # Check for inactive regions
        inactive_regions = [r for r in regions if not r.is_active]
        if inactive_regions:
            return jsonify({
                'error': f'One or more selected regions are inactive: {[r.location for r in inactive_regions]}'
            }), 400
        
        # All regions are valid and active - add them
        for region in regions:
            new_region = AgentRegion(
                agent_id=current_user['user_id'],
                region_name=f'District {region.district}',
                region_type='district',
                region_value=region.location
            )
            db.session.add(new_region)
        
        db.session.commit()
        
        return jsonify({'message': 'Regions updated successfully'}), 200
        
    except Exception as e:
        print(f"Error updating agent regions: {e}")
        import traceback
        traceback.print_exc()
        db.session.rollback()
        return jsonify({'error': f'Failed to update regions: {str(e)}'}), 500

@app.route('/api/agent/recent-activity', methods=['GET'])
@require_auth
def get_agent_recent_activity():
    """Get agent's recent activity"""
    try:
        current_user = request.user
        
        # Verify user is an agent
        user = User.query.get(current_user['user_id'])
        if not user or user.user_type != 'agent':
            return jsonify({'error': 'User is not an agent'}), 403
        
        # Get recent property status changes
        recent_properties = Property.query.filter_by(
            agent_id=current_user['user_id']
        ).order_by(Property.updated_at.desc()).limit(5).all()
        
        activity_data = []
        for property in recent_properties:
            activity_data.append({
                'id': property.id,
                'type': 'status',
                'title': f'Listing Status Updated: {property.address} - {property.status.title()}',
                'details': f'{property.status.title()} on: {property.updated_at.strftime("%Y-%m-%d")}',
                'time': '1 day ago'  # You can implement proper time calculation
            })
        
        return jsonify({'activities': activity_data}), 200
        
    except Exception as e:
        print(f"Error getting agent recent activity: {e}")
        return jsonify({'error': 'Failed to get recent activity'}), 500

@app.route('/api/properties/agent', methods=['GET'])
@require_auth
def get_agent_properties():
    """Get properties for the current agent"""
    try:
        current_user = request.user
        
        # Verify user is an agent
        user = User.query.get(current_user['user_id'])
        if not user or user.user_type != 'agent':
            return jsonify({'error': 'User is not an agent'}), 403
        
        # Get agent's properties (all properties, not limited to 5)
        properties = Property.query.filter_by(agent_id=current_user['user_id']).order_by(Property.created_at.desc()).all()
        
        properties_data = []
        for property in properties:
            # Get primary image
            primary_image = PropertyImage.query.filter_by(
                property_id=property.id, 
                is_primary=True
            ).first()
            
            image_url = primary_image.image_url if primary_image else 'https://images.unsplash.com/photo-1560472354-b33ff0c44a43?w=300&h=200&fit=crop'
            
            properties_data.append({
                'id': property.id,
                'title': property.title,
                'description': property.description,
                'property_type': property.property_type,
                'address': property.address,
                'street_address': property.street_address,
                'city': property.city,
                'state': property.state,
                'zip_code': property.zip_code,
                'size_sqft': property.size_sqft,
                'floors': property.floors,
                'year_built': property.year_built,
                'zoning': property.zoning,
                'parking_spaces': property.parking_spaces,
                'asking_price': property.asking_price,
                'price_type': property.price_type,
                'status': property.status,
                'latitude': property.latitude,
                'longitude': property.longitude,
                'image': image_url
            })
        
        return jsonify(properties_data), 200
        
    except Exception as e:
        print(f"Error getting agent properties: {e}")
        return jsonify({'error': 'Failed to get agent properties'}), 500

@app.route('/api/properties/agent/<int:property_id>', methods=['GET'])
@require_auth
def get_agent_property(property_id):
    """Get a single property for the current agent"""
    try:
        print(f"Getting agent property {property_id}")
        current_user = request.user
        print(f"Current user: {current_user}")
        
        if not current_user:
            print("No current user found in request")
            return jsonify({'error': 'Authentication required'}), 401
        
        user = User.query.get(current_user['user_id'])
        print(f"User from DB: {user.full_name if user else 'Not found'}")
        print(f"User type: {user.user_type if user else 'None'}")
        
        if not user or user.user_type != 'agent':
            print(f"User validation failed: user={user}, user_type={user.user_type if user else 'None'}")
            return jsonify({'error': 'User is not an agent'}), 403
        
        print(f"Looking for property {property_id} for agent {current_user['user_id']}")
        property = Property.query.filter_by(
            id=property_id, 
            agent_id=current_user['user_id']
        ).first()
        
        print(f"Property found: {property.title if property else 'Not found'}")
        
        if not property:
            return jsonify({'error': 'Property not found'}), 404
        
        # Get agent details
        agent = User.query.get(property.agent_id)
        agent_profile = AgentProfile.query.filter_by(user_id=property.agent_id).first()
        
        print(f"Agent found: {agent.full_name if agent else 'None'}")
        print(f"Agent profile found: {agent_profile.company_name if agent_profile else 'None'}")
        
        # Get all images
        images = PropertyImage.query.filter_by(property_id=property.id).all()
        print(f"Images found: {len(images)}")
        
        # Get amenities
        amenities = PropertyAmenity.query.filter_by(property_id=property.id).all()
        print(f"Amenities found: {len(amenities)}")
        
        property_data = {
            'id': property.id,
            'title': property.title,
            'description': property.description,
            'property_type': property.property_type,
            'address': property.address,
            'street_address': property.street_address,
            'city': property.city,
            'state': property.state,
            'zip_code': property.zip_code,
            'size_sqft': property.size_sqft,
            'floors': property.floors,
            'year_built': property.year_built,
            'zoning': property.zoning,
            'parking_spaces': property.parking_spaces,
            'asking_price': property.asking_price,
            'price_type': property.price_type,
            'status': property.status,
            'latitude': property.latitude,
            'longitude': property.longitude,
            'created_at': property.created_at.isoformat() if property.created_at else None,
            'updated_at': property.updated_at.isoformat() if property.updated_at else None,
            'agent': {
                'id': agent.id if agent else None,
                'name': agent.full_name if agent else None,
                'email': agent.email if agent else None,
                'phone': agent.phone_number if agent else None,
                'company': agent_profile.company_name if agent_profile else None,
                'license': agent_profile.license_number if agent_profile else None,
                'experience': agent_profile.years_experience if agent_profile else None,
                'bio': agent_profile.bio if agent_profile else None,
                'specializations': agent_profile.specializations if agent_profile else None,
                'contact_preferences': agent_profile.contact_preferences if agent_profile else None
            } if agent else None,
            'images': [{'id': img.id, 'url': img.image_url, 'name': img.image_name, 'is_primary': img.is_primary} for img in images],
            'amenities': [{'name': amenity.amenity_name, 'has_amenity': amenity.has_amenity} for amenity in amenities]
        }
        
        return jsonify(property_data), 200
    except Exception as e:
        print(f"Error getting agent property: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to get agent property: {str(e)}'}), 500

@app.route('/api/properties/agent', methods=['POST'])
@require_auth
def create_agent_property():
    """Create a new property for the current agent"""
    try:
        current_user = request.user
        user = User.query.get(current_user['user_id'])
        if not user or user.user_type != 'agent':
            return jsonify({'error': 'User is not an agent'}), 403
        
        data = request.get_json()
        print(f"Creating property with data: {data}")
        print(f"Latitude: {data.get('latitude')}, Longitude: {data.get('longitude')}")
        
        # Create new property
        new_property = Property(
            agent_id=current_user['user_id'],
            title=data.get('title', ''),
            description=data.get('description', ''),
            property_type=data.get('property_type'),
            address=data.get('address'),
            street_address=data.get('street_address'),
            city=data.get('city'),
            state=data.get('state'),
            zip_code=data.get('zip_code'),
            size_sqft=data.get('size_sqft'),
            floors=data.get('floors'),
            year_built=data.get('year_built'),
            zoning=data.get('zoning'),
            parking_spaces=data.get('parking_spaces'),
            asking_price=data.get('asking_price'),
            price_type=data.get('price_type'),
            status=data.get('status', 'active'),
            latitude=data.get('latitude'),
            longitude=data.get('longitude')
        )
        
        db.session.add(new_property)
        db.session.commit()
        
        # Add amenities if provided
        if data.get('amenities'):
            for amenity_name, has_amenity in data['amenities'].items():
                amenity = PropertyAmenity(
                    property_id=new_property.id,
                    amenity_name=amenity_name,
                    has_amenity=has_amenity
                )
                db.session.add(amenity)
        
        # Add images if provided
        if data.get('images'):
            for i, image_data in enumerate(data['images']):
                image = PropertyImage(
                    property_id=new_property.id,
                    image_url=image_data['url'],
                    image_name=image_data.get('name', f'image_{i+1}'),
                    is_primary=i == 0  # First image is primary
                )
                db.session.add(image)
        
        db.session.commit()
        
        return jsonify({'message': 'Property created successfully', 'id': new_property.id}), 201
    except Exception as e:
        print(f"Error creating agent property: {e}")
        db.session.rollback()
        return jsonify({'error': 'Failed to create property'}), 500

@app.route('/api/properties/agent/<int:property_id>', methods=['PUT'])
@require_auth
def update_agent_property(property_id):
    """Update a property for the current agent"""
    try:
        current_user = request.user
        user = User.query.get(current_user['user_id'])
        if not user or user.user_type != 'agent':
            return jsonify({'error': 'User is not an agent'}), 403
        
        property = Property.query.filter_by(
            id=property_id, 
            agent_id=current_user['user_id']
        ).first()
        
        if not property:
            return jsonify({'error': 'Property not found'}), 404
        
        data = request.get_json()
        
        # Update property fields
        for field in ['title', 'description', 'property_type', 'address', 'street_address', 
                     'city', 'state', 'zip_code', 'size_sqft', 'floors', 'year_built', 
                     'zoning', 'parking_spaces', 'asking_price', 'price_type', 'status', 
                     'latitude', 'longitude']:
            if field in data:
                setattr(property, field, data[field])
        
        # Update amenities if provided
        if data.get('amenities'):
            # Delete existing amenities
            PropertyAmenity.query.filter_by(property_id=property.id).delete()
            
            # Add new amenities
            for amenity_name, has_amenity in data['amenities'].items():
                amenity = PropertyAmenity(
                    property_id=property.id,
                    amenity_name=amenity_name,
                    has_amenity=has_amenity
                )
                db.session.add(amenity)
        
        # Update images if provided
        if data.get('images'):
            # Delete existing images
            PropertyImage.query.filter_by(property_id=property.id).delete()
            
            # Add new images
            for i, image_data in enumerate(data['images']):
                image = PropertyImage(
                    property_id=property.id,
                    image_url=image_data['url'],
                    image_name=image_data.get('name', f'image_{i+1}'),
                    is_primary=i == 0  # First image is primary
                )
                db.session.add(image)
        
        db.session.commit()
        
        return jsonify({'message': 'Property updated successfully'}), 200
    except Exception as e:
        print(f"Error updating agent property: {e}")
        db.session.rollback()
        return jsonify({'error': 'Failed to update property'}), 500

@app.route('/api/properties/agent/<int:property_id>', methods=['DELETE'])
@require_auth
def delete_agent_property(property_id):
    """Delete a property for the current agent"""
    try:
        print(f"🗑️ DELETE request for property {property_id}")
        current_user = request.user
        print(f"Current user: {current_user}")
        
        user = User.query.get(current_user['user_id'])
        print(f"User found: {user.full_name if user else 'None'}")
        print(f"User type: {user.user_type if user else 'None'}")
        
        if not user or user.user_type != 'agent':
            print("❌ User is not an agent")
            return jsonify({'error': 'User is not an agent'}), 403
        
        property = Property.query.filter_by(
            id=property_id, 
            agent_id=current_user['user_id']
        ).first()
        
        print(f"Property found: {property.title if property else 'None'}")
        
        if not property:
            print("❌ Property not found")
            return jsonify({'error': 'Property not found'}), 404
        
        print(f"Deleting property: {property.title} (ID: {property.id})")
        
        # Delete related data first
        try:
            amenities_deleted = PropertyAmenity.query.filter_by(property_id=property.id).delete()
            print(f"Deleted {amenities_deleted} amenities")
        except Exception as e:
            print(f"Error deleting amenities: {e}")
            raise
        
        try:
            images_deleted = PropertyImage.query.filter_by(property_id=property.id).delete()
            print(f"Deleted {images_deleted} images")
        except Exception as e:
            print(f"Error deleting images: {e}")
            raise
        
        try:
            # Delete property views
            views_deleted = PropertyView.query.filter_by(property_id=property.id).delete()
            print(f"Deleted {views_deleted} property views")
        except Exception as e:
            print(f"Error deleting property views: {e}")
            # Don't raise here as PropertyView table might not exist yet
        
        try:
            # Delete the property
            db.session.delete(property)
            db.session.commit()
            print("Property deleted successfully from database")
        except Exception as e:
            print(f"Error deleting property: {e}")
            raise
        
        print(f"✅ Property {property_id} deleted successfully")
        return jsonify({'message': 'Property deleted successfully'}), 200
    except Exception as e:
        print(f"Error deleting agent property: {e}")
        import traceback
        traceback.print_exc()
        db.session.rollback()
        return jsonify({'error': f'Failed to delete property: {str(e)}'}), 500

# Add referral code column and generate codes for premium users
@app.route('/api/db/add-referral-codes', methods=['POST'])
def add_referral_codes():
    try:
        # Check if referral_code column exists
        check_query = text("""
            SELECT column_name 
            FROM information_schema.columns 
            WHERE table_name = 'users' 
            AND column_name = 'referral_code'
        """)
        
        result = db.session.execute(check_query)
        existing_columns = [row[0] for row in result]
        
        if 'referral_code' not in existing_columns:
            # Add referral_code column
            db.session.execute(text("ALTER TABLE users ADD COLUMN referral_code VARCHAR(20) UNIQUE"))
            print("Added referral_code column")
        
        # Generate referral codes for premium users who don't have one
        update_query = text("""
            UPDATE users 
            SET referral_code = CONCAT('REF', LPAD(id::text, 6, '0'))
            WHERE user_type = 'premium' AND (referral_code IS NULL OR referral_code = '')
        """)
        
        db.session.execute(update_query)
        print("Generated referral codes for premium users")
        
        db.session.commit()
        
        return jsonify({
            'message': 'Referral codes added successfully',
            'added_column': 'referral_code' if 'referral_code' not in existing_columns else None,
            'codes_generated': True
        }), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"Error adding referral codes: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to add referral codes: {str(e)}'}), 500

# Check database contents
@app.route('/api/db/check-contents', methods=['GET'])
def check_database_contents():
    try:
        # Check counts of different tables
        user_count = User.query.count()
        property_count = Property.query.count()
        prediction_count = PricePrediction.query.count()
        bookmark_count = Bookmark.query.count()
        
        return jsonify({
            'users': user_count,
            'properties': property_count,
            'price_predictions': prediction_count,
            'bookmarks': bookmark_count
        }), 200
        
    except Exception as e:
        print(f"Error checking database contents: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to check database: {str(e)}'}), 500

# Migration endpoint to add additional columns to bookmarks table
@app.route('/api/db/add-bookmark-details', methods=['POST'])
def add_bookmark_details():
    try:
        # Add multiple columns to bookmarks table if they don't exist
        columns_to_add = [
            ('address', 'VARCHAR(500)'),
            ('floor_area', 'DECIMAL(10,2)'),
            ('level', 'VARCHAR(50)'),
            ('unit_number', 'VARCHAR(50)'),
            ('property_type', 'VARCHAR(100)'),
            ('address_2', 'VARCHAR(500)'),
            ('floor_area_2', 'DECIMAL(10,2)'),
            ('level_2', 'VARCHAR(50)'),
            ('unit_number_2', 'VARCHAR(50)'),
            ('property_type_2', 'VARCHAR(100)')
        ]
        
        for column_name, column_type in columns_to_add:
            query = text(f"""
                ALTER TABLE bookmarks 
                ADD COLUMN IF NOT EXISTS {column_name} {column_type}
            """)
            db.session.execute(query)
            print(f"Added column: {column_name}")
        
        db.session.commit()
        return jsonify({'message': 'Additional bookmark columns added successfully'}), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"Error adding bookmark columns: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to add bookmark columns: {str(e)}'}), 500

# Test endpoint to add sample bookmarks (for development only)
@app.route('/api/db/add-sample-bookmarks', methods=['POST'])
def add_sample_bookmarks():
    try:
        # Check if we have any users
        users = User.query.limit(1).all()
        if not users:
            return jsonify({'error': 'No users found. Please create a user first.'}), 400
        
        user_id = users[0].id
        
        # Add sample bookmarks with detailed property information for testing
        sample_bookmarks = [
            Bookmark(
                user_id=user_id, 
                bookmark_type='property', 
                reference_id=1,
                address='123 Jurong East Ave 1, Singapore 600123',
                floor_area=5000,
                level='Ground Floor',
                unit_number='A-01',
                property_type='Industrial Property'
            ),
            Bookmark(
                user_id=user_id, 
                bookmark_type='prediction', 
                reference_id=1,
                address='456 Tuas South Ave 2, Singapore 637456',
                floor_area=8000,
                level='Level 1',
                unit_number='B-15',
                property_type='Factory Building'
            ),
            Bookmark(
                user_id=user_id, 
                bookmark_type='comparison', 
                reference_id=1,
                address='789 Woodlands Industrial Park, Singapore 738789',
                floor_area=12000,
                level='Ground Floor',
                unit_number='C-03',
                property_type='Warehouse',
                address_2='321 Changi Business Park, Singapore 486321',
                floor_area_2=3000,
                level_2='Level 3',
                unit_number_2='D-22',
                property_type_2='Office Space'
            ),
        ]
        
        for bookmark in sample_bookmarks:
            # Check if bookmark already exists
            existing = Bookmark.query.filter_by(
                user_id=bookmark.user_id,
                bookmark_type=bookmark.bookmark_type,
                reference_id=bookmark.reference_id
            ).first()
            
            if not existing:
                db.session.add(bookmark)
        
        db.session.commit()
        
        return jsonify({
            'message': 'Sample bookmarks added successfully',
            'bookmarks_added': len(sample_bookmarks)
        }), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"Error adding sample bookmarks: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to add sample bookmarks: {str(e)}'}), 500

# Endpoint to populate existing bookmarks with sample data
@app.route('/api/db/populate-bookmarks', methods=['POST'])
def populate_bookmarks():
    try:
        # Update existing bookmarks with sample data
        sample_data = [
            {
                'id': 1,
                'address': '123 Jurong East Ave 1, Singapore 600123',
                'floor_area': 5000,
                'level': 'Ground Floor',
                'unit_number': 'A-01',
                'property_type': 'Industrial Property'
            },
            {
                'id': 2,
                'address': '456 Tuas South Ave 2, Singapore 637456',
                'floor_area': 8000,
                'level': 'Level 1',
                'unit_number': 'B-15',
                'property_type': 'Factory Building'
            },
            {
                'id': 3,
                'address': '789 Woodlands Industrial Park, Singapore 738789',
                'floor_area': 12000,
                'level': 'Ground Floor',
                'unit_number': 'C-03',
                'property_type': 'Warehouse',
                'address_2': '321 Changi Business Park, Singapore 486321',
                'floor_area_2': 3000,
                'level_2': 'Level 3',
                'unit_number_2': 'D-22',
                'property_type_2': 'Office Space'
            },
            {
                'id': 4,
                'address': '123 Jurong East Ave 1, Singapore 600123',
                'floor_area': 5000,
                'level': 'Ground Floor',
                'unit_number': 'A-01',
                'property_type': 'Industrial Property'
            },
            {
                'id': 5,
                'address': '456 Tuas South Ave 2, Singapore 637456',
                'floor_area': 8000,
                'level': 'Level 1',
                'unit_number': 'B-15',
                'property_type': 'Factory Building'
            },
            {
                'id': 6,
                'address': '789 Woodlands Industrial Park, Singapore 738789',
                'floor_area': 12000,
                'level': 'Ground Floor',
                'unit_number': 'C-03',
                'property_type': 'Warehouse',
                'address_2': '321 Changi Business Park, Singapore 486321',
                'floor_area_2': 3000,
                'level_2': 'Level 3',
                'unit_number_2': 'D-22',
                'property_type_2': 'Office Space'
            }
        ]
        
        for data in sample_data:
            query = text("""
                UPDATE bookmarks 
                SET address = :address, 
                    floor_area = :floor_area, 
                    level = :level, 
                    unit_number = :unit_number, 
                    property_type = :property_type,
                    address_2 = :address_2,
                    floor_area_2 = :floor_area_2,
                    level_2 = :level_2,
                    unit_number_2 = :unit_number_2,
                    property_type_2 = :property_type_2
                WHERE id = :id
            """)
            db.session.execute(query, data)
        
        db.session.commit()
        return jsonify({'message': 'Bookmarks populated with sample data successfully'}), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"Error populating bookmarks: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to populate bookmarks: {str(e)}'}), 500

# Bookmarks table already exists in database_fyp.sql
# No need to create it here

# Add admin response columns to user_reviews table
@app.route('/api/db/add-admin-response-columns', methods=['POST'])
def add_admin_response_columns():
    try:
        # Check if columns already exist
        check_query = text("""
            SELECT column_name 
            FROM information_schema.columns 
            WHERE table_name = 'user_reviews' 
            AND column_name IN ('admin_response', 'admin_response_date', 'admin_id')
        """)
        
        result = db.session.execute(check_query)
        existing_columns = [row[0] for row in result]
        
        if 'admin_response' not in existing_columns:
            # Add admin_response column
            db.session.execute(text("ALTER TABLE user_reviews ADD COLUMN admin_response TEXT"))
            print("Added admin_response column")
        
        if 'admin_response_date' not in existing_columns:
            # Add admin_response_date column
            db.session.execute(text("ALTER TABLE user_reviews ADD COLUMN admin_response_date TIMESTAMP"))
            print("Added admin_response_date column")
        
        if 'admin_id' not in existing_columns:
            # Add admin_id column
            db.session.execute(text("ALTER TABLE user_reviews ADD COLUMN admin_id INTEGER REFERENCES users(id)"))
            print("Added admin_id column")
        
        if 'verified_date' not in existing_columns:
            # Add verified_date column
            db.session.execute(text("ALTER TABLE user_reviews ADD COLUMN verified_date TIMESTAMP"))
            print("Added verified_date column")
        
        if 'verified_by' not in existing_columns:
            # Add verified_by column
            db.session.execute(text("ALTER TABLE user_reviews ADD COLUMN verified_by INTEGER REFERENCES users(id)"))
            print("Added verified_by column")
        
        db.session.commit()
        return jsonify({'message': 'Admin response columns added successfully'}), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"Error adding admin response columns: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to add columns: {str(e)}'}), 500

# Upload property images
@app.route('/api/properties/upload-images', methods=['POST'])
@require_auth
def upload_property_images():
    print("=== Property Images Upload Debug ===")
    print(f"Request method: {request.method}")
    print(f"Request files: {list(request.files.keys())}")
    
    # Get user from authenticated token
    user_id = request.user['user_id']
    print(f"User ID: {user_id}")
    
    user = User.query.get(user_id)
    if not user:
        print("User not found")
        return jsonify({'error': 'User not found'}), 404
    
    # Check if user is an agent
    if user.user_type != 'agent':
        print("User is not an agent")
        return jsonify({'error': 'Access denied. Property image upload only available for agent users.'}), 403
    
    # Check if files were uploaded
    if 'property_images' not in request.files:
        print("No property_images in request.files")
        return jsonify({'error': 'No property images provided'}), 400
    
    files = request.files.getlist('property_images')
    print(f"Number of files: {len(files)}")
    
    if not files or files[0].filename == '':
        print("No files selected")
        return jsonify({'error': 'No files selected'}), 400
    
    # Check file types
    allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
    uploaded_images = []
    
    try:
        for file in files:
            if file.filename == '':
                continue
                
            print(f"Processing file: {file.filename}")
            print(f"File content type: {file.content_type}")
            
            # Check file type
            if not file.filename.lower().endswith(tuple('.' + ext for ext in allowed_extensions)):
                print(f"Invalid file type: {file.filename}")
                continue
            
            # Generate unique filename
            import os
            from werkzeug.utils import secure_filename
            filename = secure_filename(file.filename)
            timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
            unique_filename = f"property_{user_id}_{timestamp}_{filename}"
            print(f"Generated filename: {unique_filename}")
            
            # Create upload directory if it doesn't exist
            upload_dir = os.path.join(os.getcwd(), 'uploads', 'properties')
            print(f"Upload directory: {upload_dir}")
            os.makedirs(upload_dir, exist_ok=True)
            print(f"Upload directory created/exists: {os.path.exists(upload_dir)}")
            
            # Save file
            file_path = os.path.join(upload_dir, unique_filename)
            print(f"Full file path: {file_path}")
            file.save(file_path)
            print(f"File saved successfully: {os.path.exists(file_path)}")
            
            # Create image URL
            image_url = f"/uploads/properties/{unique_filename}"
            uploaded_images.append({
                'filename': unique_filename,
                'original_name': filename,
                'url': image_url
            })
        
        print(f"Successfully uploaded {len(uploaded_images)} images")
        
        return jsonify({
            'message': f'Successfully uploaded {len(uploaded_images)} images',
            'images': uploaded_images
        }), 200
        
    except Exception as e:
        print(f"Error uploading property images: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Failed to upload property images'}), 500

# Serve uploaded property images
@app.route('/uploads/properties/<filename>')
def serve_property_image(filename):
    """Serve uploaded property images"""
    print(f"=== Serving Property Image ===")
    print(f"Requested filename: {filename}")
    
    upload_dir = os.path.join(os.getcwd(), 'uploads', 'properties')
    print(f"Upload directory: {upload_dir}")
    
    file_path = os.path.join(upload_dir, filename)
    print(f"Full file path: {file_path}")
    print(f"File exists: {os.path.exists(file_path)}")
    
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        return jsonify({'error': 'File not found'}), 404
    
    try:
        response = send_from_directory(upload_dir, filename)
        # Add CORS headers
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Methods'] = 'GET'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        return response
    except Exception as e:
        print(f"Error serving file: {e}")
        return jsonify({'error': 'Error serving file'}), 500


# Trial system endpoints
@app.route('/api/trial/check', methods=['GET'])
@require_auth
def check_trial_status():
    """Check if user has trial available or used"""
    try:
        current_user = request.user
        user = User.query.get(current_user['user_id'])
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        # Get or create user profile
        user_profile = UserProfile.query.filter_by(user_id=user.id).first()
        if not user_profile:
            user_profile = UserProfile(user_id=user.id)
            db.session.add(user_profile)
            db.session.commit()
        
        return jsonify({
            'trial_used': user_profile.trial_used,
            'trial_predictions_used': user_profile.trial_predictions_used,
            'max_trial_predictions': user_profile.max_trial_predictions,
            'trial_available': not user_profile.trial_used and user_profile.trial_predictions_used < user_profile.max_trial_predictions,
            'trial_start_date': user_profile.trial_start_date.isoformat() if user_profile.trial_start_date else None,
            'trial_end_date': user_profile.trial_end_date.isoformat() if user_profile.trial_end_date else None
        }), 200
        
    except Exception as e:
        print(f"Error checking trial status: {e}")
        return jsonify({'error': 'Failed to check trial status'}), 500

@app.route('/api/trial/start', methods=['POST'])
@require_auth
def start_trial():
    """Start trial for user"""
    try:
        current_user = request.user
        user = User.query.get(current_user['user_id'])
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        # Get or create user profile
        user_profile = UserProfile.query.filter_by(user_id=user.id).first()
        if not user_profile:
            user_profile = UserProfile(user_id=user.id)
            db.session.add(user_profile)
        
        # Check if trial already used
        if user_profile.trial_used:
            return jsonify({'error': 'Trial already used'}), 400
        
        # Start trial
        user_profile.trial_start_date = datetime.utcnow()
        user_profile.trial_end_date = datetime.utcnow() + timedelta(days=7)  # 7-day trial
        user_profile.trial_used = True
        
        db.session.commit()
        
        return jsonify({
            'message': 'Trial started successfully',
            'trial_start_date': user_profile.trial_start_date.isoformat(),
            'trial_end_date': user_profile.trial_end_date.isoformat(),
            'max_predictions': user_profile.max_trial_predictions
        }), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"Error starting trial: {e}")
        return jsonify({'error': 'Failed to start trial'}), 500

@app.route('/api/trial/use-prediction', methods=['POST'])
@require_auth
def use_trial_prediction():
    """Use a trial prediction"""
    try:
        current_user = request.user
        user = User.query.get(current_user['user_id'])
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        # Get user profile
        user_profile = UserProfile.query.filter_by(user_id=user.id).first()
        if not user_profile:
            return jsonify({'error': 'User profile not found'}), 404
        
        # Check if trial is available
        if user_profile.trial_predictions_used >= user_profile.max_trial_predictions:
            return jsonify({'error': 'Trial predictions exhausted'}), 400
        
        # Increment trial predictions used
        user_profile.trial_predictions_used += 1
        db.session.commit()
        
        return jsonify({
            'message': 'Trial prediction used successfully',
            'predictions_remaining': user_profile.max_trial_predictions - user_profile.trial_predictions_used
        }), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"Error using trial prediction: {e}")
        return jsonify({'error': 'Failed to use trial prediction'}), 500

# User preferences and recommendations endpoints
@app.route('/api/onboarding/save-preferences', methods=['POST'])
@require_auth
def save_user_preferences():
    """Save user property type and location preferences"""
    try:
        current_user = request.user
        data = request.get_json()
        
        # Get user
        user = User.query.get(current_user['user_id'])
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        # Update user preferences
        user.property_type_preferences = data.get('propertyTypes', [])
        user.location_preferences = data.get('locations', [])
        # Don't reset first_time_user here - only set during initial onboarding
        
        db.session.commit()
        
        return jsonify({
            'message': 'Preferences saved successfully',
            'propertyTypes': user.property_type_preferences,
            'locations': user.location_preferences
        }), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"Error saving user preferences: {e}")
        return jsonify({'error': 'Failed to save preferences'}), 500

@app.route('/api/recommendations', methods=['GET'])
@require_auth
def get_user_recommendations():
    """Get personalized property recommendations based on user preferences"""
    try:
        current_user = request.user
        
        # Get user with preferences
        user = User.query.get(current_user['user_id'])
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        # Get user preferences - ensure they're lists even if stored as None or invalid JSON
        property_types = user.property_type_preferences if user.property_type_preferences is not None else []
        if not isinstance(property_types, list):
            property_types = []
        
        locations = user.location_preferences if user.location_preferences is not None else []
        if not isinstance(locations, list):
            locations = []
        
        # Get all properties
        properties = Property.query.filter_by(status='active').all()
        
        if not properties:
            return jsonify({'recommendations': []}), 200
        
        # Filter properties based on preferences
        filtered_properties = []
        
        for property in properties:
            score = 0
            
            # Score based on property type preferences
            if property_types and property.property_type:
                # Safely handle property_type - might not be a string
                property_type_str = str(property.property_type) if property.property_type else ''
                property_type_lower = property_type_str.lower()
                for pref_type in property_types:
                    pref_type_str = str(pref_type) if pref_type else ''
                    pref_type_lower = pref_type_str.lower()
                    
                    # Direct match
                    if property_type_lower == pref_type_lower:
                        score += 10
                    # Related matches
                    elif pref_type_lower == 'warehouse' and property_type_lower in ['storage', 'logistics']:
                        score += 8
                    elif pref_type_lower in ['single-user factory', 'multiple-user factory'] and property_type_lower in ['factory', 'industrial']:
                        score += 8
                    elif pref_type_lower == 'office' and property_type_lower in ['business parks', 'commercial']:
                        score += 6
                    elif pref_type_lower == 'retail' and property_type_lower in ['shop house', 'commercial']:
                        score += 6
            
            # Score based on location preferences
            if locations and property.address:
                # Safely handle address and city - they might be None or not strings
                address_lower = str(property.address).lower() if property.address else ''
                city_lower = str(property.city).lower() if property.city else ''
                
                for location in locations:
                    if location == 'CBD' and (
                        'raffles place' in address_lower or 
                        'marina bay' in address_lower or 
                        'cecil' in address_lower or
                        'raffles' in city_lower or
                        'marina' in city_lower
                    ):
                        score += 8
                    elif location == 'City Fringe' and (
                        'orchard' in address_lower or 
                        'tanglin' in address_lower or 
                        'bukit timah' in address_lower or
                        'balestier' in address_lower or
                        'toa payoh' in address_lower or
                        'orchard' in city_lower or
                        'tanglin' in city_lower
                    ):
                        score += 8
                    elif location == 'Industrial Areas' and (
                        'jurong' in address_lower or 
                        'tuas' in address_lower or 
                        'changi' in address_lower or
                        'woodlands' in address_lower or
                        'industrial' in address_lower or
                        'factory' in address_lower or
                        'warehouse' in address_lower or
                        'jurong' in city_lower or
                        'tuas' in city_lower or
                        'changi' in city_lower or
                        'woodlands' in city_lower
                    ):
                        score += 8
                    elif location == 'General':
                        score += 2  # Small boost for general preference
            
            # Add property to filtered list with score
            if score > 0:
                filtered_properties.append({
                    'property': property,
                    'score': score
                })
        
        # Sort by score (highest first) and get top 6 recommendations
        filtered_properties.sort(key=lambda x: x['score'], reverse=True)
        top_recommendations = filtered_properties[:6]
        
        # Format recommendations
        recommendations = []
        for item in top_recommendations:
            property = item['property']
            
            # Get primary image
            primary_image = PropertyImage.query.filter_by(
                property_id=property.id, 
                is_primary=True
            ).first()
            
            # Get agent info - safely handle if agent_id is None
            agent = None
            agent_profile = None
            if property.agent_id:
                agent = User.query.get(property.agent_id)
                if agent:
                    agent_profile = AgentProfile.query.filter_by(user_id=property.agent_id).first()
            
            recommendations.append({
                'id': property.id,
                'title': property.title,
                'description': property.description,
                'property_type': property.property_type,
                'address': property.address,
                'city': property.city,
                'size_sqft': float(property.size_sqft) if property.size_sqft is not None else None,
                'floors': property.floors,
                'year_built': property.year_built,
                'parking_spaces': property.parking_spaces,
                'price': float(property.asking_price) if property.asking_price is not None else None,
                'price_type': property.price_type,
                'status': property.status,
                'latitude': float(property.latitude) if property.latitude else None,
                'longitude': float(property.longitude) if property.longitude else None,
                'image': primary_image.image_url if primary_image else None,
                'agent': {
                    'id': agent.id if agent else None,
                    'name': agent.full_name if agent else 'Unknown Agent',
                    'phone': agent.phone_number if agent else None,
                    'email': agent.email if agent else None,
                    'company': agent_profile.company_name if agent_profile else None,
                    'license': agent_profile.license_number if agent_profile else None
                } if agent else None,
                'score': item['score']
            })
        
        return jsonify({
            'recommendations': recommendations,
            'userPreferences': {
                'propertyTypes': property_types,
                'locations': locations
            }
        }), 200
        
    except Exception as e:
        import traceback
        error_traceback = traceback.format_exc()
        print(f"Error getting recommendations: {e}")
        print(f"Full traceback: {error_traceback}")
        return jsonify({'error': f'Failed to get recommendations: {str(e)}'}), 500

@app.route('/api/user/preferences', methods=['GET'])
@require_auth
def get_user_preferences():
    """Get current user's preferences"""
    try:
        current_user = request.user
        
        user = User.query.get(current_user['user_id'])
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        return jsonify({
            'propertyTypes': user.property_type_preferences or [],
            'locations': user.location_preferences or []
        }), 200
        
    except Exception as e:
        print(f"Error getting user preferences: {e}")
        return jsonify({'error': 'Failed to get preferences'}), 500

# Support Content API Endpoints

@app.route('/api/support/faq', methods=['GET'])
def get_faq_entries():
    """Get all FAQ entries"""
    try:
        faqs = FAQEntry.query.filter_by(is_active=True).order_by(FAQEntry.display_order).all()
        
        faq_list = []
        for faq in faqs:
            faq_list.append({
                'id': faq.id,
                'question': faq.question,
                'answer': faq.answer,
                'category': faq.category,
                'display_order': faq.display_order
            })
        
        return jsonify({
            'success': True,
            'faqs': faq_list
        }), 200
        
    except Exception as e:
        print(f"Error getting FAQ entries: {e}")
        return jsonify({'error': 'Failed to get FAQ entries'}), 500

# FAQ Management API Endpoints

@app.route('/api/faq/section', methods=['GET'])
def get_faq_section():
    """Get FAQ section details"""
    try:
        section = FAQSection.query.first()
        if not section:
            # Create default section if none exists
            section = FAQSection(section_title='Frequently Asked Questions')
            db.session.add(section)
            db.session.commit()
        
        return jsonify({
            'success': True,
            'section': {
                'id': section.id,
                'section_title': section.section_title,
                'created_at': section.created_at.isoformat() if section.created_at else None,
                'updated_at': section.updated_at.isoformat() if section.updated_at else None
            }
        }), 200
        
    except Exception as e:
        print(f"Error getting FAQ section: {e}")
        return jsonify({'error': 'Failed to get FAQ section'}), 500

@app.route('/api/faq/section', methods=['PUT'])
@require_auth
def update_faq_section():
    """Update FAQ section details"""
    try:
        data = request.get_json()
        section_title = data.get('section_title', '').strip()
        
        if not section_title:
            return jsonify({'error': 'Section title is required'}), 400
        
        section = FAQSection.query.first()
        if not section:
            section = FAQSection(section_title=section_title)
            db.session.add(section)
        else:
            section.section_title = section_title
            section.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'FAQ section updated successfully'
        }), 200
        
    except Exception as e:
        print(f"Error updating FAQ section: {e}")
        db.session.rollback()
        return jsonify({'error': 'Failed to update FAQ section'}), 500

@app.route('/api/faq/entries', methods=['GET'])
def get_faq_entries_admin():
    """Get all FAQ entries for admin management"""
    try:
        faqs = FAQEntry.query.order_by(FAQEntry.display_order, FAQEntry.created_at).all()
        
        faq_list = []
        for faq in faqs:
            faq_list.append({
                'id': faq.id,
                'question': faq.question,
                'answer': faq.answer,
                'category': faq.category,
                'display_order': faq.display_order,
                'is_active': faq.is_active,
                'created_at': faq.created_at.isoformat(),
                'updated_at': faq.updated_at.isoformat()
            })
        
        return jsonify({
            'success': True,
            'faqs': faq_list
        }), 200
        
    except Exception as e:
        print(f"Error getting FAQ entries: {e}")
        return jsonify({'error': 'Failed to get FAQ entries'}), 500

@app.route('/api/faq/entries/<int:faq_id>', methods=['GET'])
def get_faq_entry(faq_id):
    """Get a specific FAQ entry"""
    try:
        faq = FAQEntry.query.get(faq_id)
        if not faq:
            return jsonify({'error': 'FAQ entry not found'}), 404
        
        return jsonify({
            'success': True,
            'faq': {
                'id': faq.id,
                'question': faq.question,
                'answer': faq.answer,
                'category': faq.category,
                'display_order': faq.display_order,
                'is_active': faq.is_active,
                'created_at': faq.created_at.isoformat(),
                'updated_at': faq.updated_at.isoformat()
            }
        }), 200
        
    except Exception as e:
        print(f"Error getting FAQ entry: {e}")
        return jsonify({'error': 'Failed to get FAQ entry'}), 500

@app.route('/api/faq/entries', methods=['POST'])
@require_auth
def create_faq_entry():
    """Create a new FAQ entry"""
    try:
        data = request.get_json()
        question = data.get('question', '').strip()
        answer = data.get('answer', '').strip()
        category = data.get('category', '').strip()
        display_order = data.get('display_order', 0)
        
        if not question or not answer:
            return jsonify({'error': 'Question and answer are required'}), 400
        
        # Get the next display order if not provided
        if display_order == 0:
            last_faq = FAQEntry.query.order_by(FAQEntry.display_order.desc()).first()
            display_order = (last_faq.display_order + 1) if last_faq else 1
        
        faq = FAQEntry(
            question=question,
            answer=answer,
            category=category,
            display_order=display_order
        )
        
        db.session.add(faq)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'FAQ entry created successfully',
            'faq': {
                'id': faq.id,
                'question': faq.question,
                'answer': faq.answer,
                'category': faq.category,
                'display_order': faq.display_order,
                'is_active': faq.is_active
            }
        }), 201
        
    except Exception as e:
        print(f"Error creating FAQ entry: {e}")
        db.session.rollback()
        return jsonify({'error': 'Failed to create FAQ entry'}), 500

@app.route('/api/faq/entries/<int:faq_id>', methods=['PUT'])
@require_auth
def update_faq_entry(faq_id):
    """Update an existing FAQ entry"""
    try:
        data = request.get_json()
        faq = FAQEntry.query.get(faq_id)
        
        if not faq:
            return jsonify({'error': 'FAQ entry not found'}), 404
        
        question = data.get('question', '').strip()
        answer = data.get('answer', '').strip()
        category = data.get('category', '').strip()
        display_order = data.get('display_order', faq.display_order)
        is_active = data.get('is_active', faq.is_active)
        
        if not question or not answer:
            return jsonify({'error': 'Question and answer are required'}), 400
        
        faq.question = question
        faq.answer = answer
        faq.category = category
        faq.display_order = display_order
        faq.is_active = is_active
        faq.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'FAQ entry updated successfully',
            'faq': {
                'id': faq.id,
                'question': faq.question,
                'answer': faq.answer,
                'category': faq.category,
                'display_order': faq.display_order,
                'is_active': faq.is_active
            }
        }), 200
        
    except Exception as e:
        print(f"Error updating FAQ entry: {e}")
        db.session.rollback()
        return jsonify({'error': 'Failed to update FAQ entry'}), 500

@app.route('/api/faq/entries/<int:faq_id>', methods=['DELETE'])
@require_auth
def delete_faq_entry(faq_id):
    """Delete an FAQ entry"""
    try:
        faq = FAQEntry.query.get(faq_id)
        
        if not faq:
            return jsonify({'error': 'FAQ entry not found'}), 404
        
        db.session.delete(faq)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'FAQ entry deleted successfully'
        }), 200
        
    except Exception as e:
        print(f"Error deleting FAQ entry: {e}")
        db.session.rollback()
        return jsonify({'error': 'Failed to delete FAQ entry'}), 500

@app.route('/api/support/contact', methods=['GET'])
def get_contact_info():
    """Get support contact information"""
    try:
        # Get contact info from database
        contact_items = db.session.execute(text("""
            SELECT contact_type, contact_value, display_order 
            FROM support_contact_info 
            WHERE is_active = true 
            ORDER BY display_order
        """)).fetchall()
        
        contact_info = {}
        for item in contact_items:
            contact_info[item.contact_type] = item.contact_value
        
        return jsonify({
            'success': True,
            'contact': contact_info
        }), 200
        
    except Exception as e:
        print(f"Error getting contact info: {e}")
        return jsonify({'error': 'Failed to get contact info'}), 500

@app.route('/api/support/legal/<content_type>', methods=['GET'])
def get_legal_content(content_type):
    """Get legal content (disclaimer, privacy_policy, terms_of_use)"""
    try:
        valid_types = ['disclaimer', 'privacy_policy', 'terms_of_use']
        if content_type not in valid_types:
            return jsonify({'error': 'Invalid content type'}), 400
        
        result = db.session.execute(text("""
            SELECT title, content, version 
            FROM legal_content 
            WHERE content_type = :content_type AND is_active = true
            ORDER BY created_at DESC
            LIMIT 1
        """), {'content_type': content_type}).fetchone()
        
        if not result:
            return jsonify({'error': 'Content not found'}), 404
        
        return jsonify({
            'success': True,
            'content': {
                'title': result.title,
                'content': result.content,
                'version': result.version
            }
        }), 200
        
    except Exception as e:
        print(f"Error getting legal content: {e}")
        return jsonify({'error': 'Failed to get legal content'}), 500

@app.route('/api/support/legal', methods=['GET'])
def get_all_legal_content():
    """Get all legal content"""
    try:
        result = db.session.execute(text("""
            SELECT content_type, title, content, version 
            FROM legal_content 
            WHERE is_active = true
            ORDER BY content_type
        """)).fetchall()
        
        legal_content = {}
        for row in result:
            legal_content[row.content_type] = {
                'title': row.title,
                'content': row.content,
                'version': row.version
            }
        
        return jsonify({
            'success': True,
            'legal_content': legal_content
        }), 200
        
    except Exception as e:
        print(f"Error getting all legal content: {e}")
        return jsonify({'error': 'Failed to get legal content'}), 500

# Hero Content API Endpoints

@app.route('/api/hero/content', methods=['GET'])
def get_hero_content():
    """Get hero section content"""
    try:
        result = db.session.execute(text("""
            SELECT section_name, headline, subheading, hero_background_url, 
                   marketing_video_url, button_text, button_url
            FROM hero_content 
            WHERE is_active = true
            ORDER BY updated_at DESC
            LIMIT 1
        """)).fetchone()
        
        if not result:
            return jsonify({'error': 'Hero content not found'}), 404
        
        return jsonify({
            'success': True,
            'hero_content': {
                'section_name': result.section_name,
                'headline': result.headline,
                'subheading': result.subheading,
                'hero_background_url': result.hero_background_url,
                'marketing_video_url': result.marketing_video_url,
                'button_text': result.button_text,
                'button_url': result.button_url
            }
        }), 200
        
    except Exception as e:
        print(f"Error getting hero content: {e}")
        return jsonify({'error': 'Failed to get hero content'}), 500

@app.route('/api/hero/content/<section_name>', methods=['GET'])
def get_hero_content_by_section(section_name):
    """Get specific hero section content by section name"""
    try:
        result = db.session.execute(text("""
            SELECT section_name, headline, subheading, hero_background_url, 
                   marketing_video_url, button_text, button_url
            FROM hero_content 
            WHERE section_name = :section_name AND is_active = true
            ORDER BY updated_at DESC
            LIMIT 1
        """), {'section_name': section_name}).fetchone()
        
        if not result:
            return jsonify({'error': 'Hero content not found'}), 404
        
        return jsonify({
            'success': True,
            'hero_content': {
                'section_name': result.section_name,
                'headline': result.headline,
                'subheading': result.subheading,
                'hero_background_url': result.hero_background_url,
                'marketing_video_url': result.marketing_video_url,
                'button_text': result.button_text,
                'button_url': result.button_url
            }
        }), 200
        
    except Exception as e:
        print(f"Error getting hero content: {e}")
        return jsonify({'error': 'Failed to get hero content'}), 500

# Hero Content File Upload Endpoints

@app.route('/api/hero/upload-background', methods=['POST'])
@require_auth
def upload_hero_background():
    """Upload hero background image"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Check file type
        allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
        if not ('.' in file.filename and file.filename.rsplit('.', 1)[1].lower() in allowed_extensions):
            return jsonify({'error': 'Invalid file type. Allowed: PNG, JPG, JPEG, GIF, WEBP'}), 400
        
        # Generate unique filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"hero_bg_{timestamp}_{file.filename}"
        
        # Ensure directory exists
        upload_dir = os.path.join('admin', 'hero', 'backgrounds')
        os.makedirs(upload_dir, exist_ok=True)
        
        # Save file
        file_path = os.path.join(upload_dir, filename)
        file.save(file_path)
        
        # Return the URL path
        file_url = f"/admin/hero/backgrounds/{filename}"
        
        return jsonify({
            'success': True,
            'file_url': file_url,
            'filename': filename
        }), 200
        
    except Exception as e:
        print(f"Error uploading hero background: {e}")
        return jsonify({'error': 'Failed to upload background image'}), 500

@app.route('/api/hero/upload-video', methods=['POST'])
@require_auth
def upload_hero_video():
    """Upload hero marketing video"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Check file type
        allowed_extensions = {'mp4', 'webm', 'ogg', 'mov'}
        if not ('.' in file.filename and file.filename.rsplit('.', 1)[1].lower() in allowed_extensions):
            return jsonify({'error': 'Invalid file type. Allowed: MP4, WEBM, OGG, MOV'}), 400
        
        # Generate unique filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"hero_video_{timestamp}_{file.filename}"
        
        # Ensure directory exists
        upload_dir = os.path.join('admin', 'hero', 'videos')
        os.makedirs(upload_dir, exist_ok=True)
        
        # Save file
        file_path = os.path.join(upload_dir, filename)
        file.save(file_path)
        
        # Return the URL path
        file_url = f"/admin/hero/videos/{filename}"
        
        return jsonify({
            'success': True,
            'file_url': file_url,
            'filename': filename
        }), 200
        
    except Exception as e:
        print(f"Error uploading hero video: {e}")
        return jsonify({'error': 'Failed to upload video'}), 500

@app.route('/api/hero/update-content', methods=['POST'])
@require_auth
def update_hero_content():
    """Update hero content in database"""
    try:
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['headline', 'subheading', 'button_text', 'button_url']
        for field in required_fields:
            if field not in data:
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        # Update or insert hero content
        result = db.session.execute(text("""
            INSERT INTO hero_content (section_name, headline, subheading, hero_background_url, 
                                    marketing_video_url, button_text, button_url)
            VALUES ('main_hero', :headline, :subheading, :hero_background_url, 
                    :marketing_video_url, :button_text, :button_url)
            ON CONFLICT (section_name) 
            DO UPDATE SET 
                headline = EXCLUDED.headline,
                subheading = EXCLUDED.subheading,
                hero_background_url = EXCLUDED.hero_background_url,
                marketing_video_url = EXCLUDED.marketing_video_url,
                button_text = EXCLUDED.button_text,
                button_url = EXCLUDED.button_url,
                updated_at = CURRENT_TIMESTAMP
        """), {
            'headline': data['headline'],
            'subheading': data['subheading'],
            'hero_background_url': data.get('hero_background_url'),
            'marketing_video_url': data.get('marketing_video_url'),
            'button_text': data['button_text'],
            'button_url': data['button_url']
        })
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Hero content updated successfully'
        }), 200
        
    except Exception as e:
        print(f"Error updating hero content: {e}")
        db.session.rollback()
        return jsonify({'error': 'Failed to update hero content'}), 500

# Serve uploaded files
@app.route('/admin/hero/backgrounds/<filename>')
def serve_hero_background(filename):
    """Serve hero background images"""
    return send_from_directory(os.path.join('admin', 'hero', 'backgrounds'), filename)

@app.route('/admin/hero/videos/<filename>')
def serve_hero_video(filename):
    """Serve hero videos"""
    return send_from_directory(os.path.join('admin', 'hero', 'videos'), filename)

# HowItWorks Content API Endpoints

@app.route('/api/howitworks/properties', methods=['GET'])
def get_howitworks_properties():
    """Get all HowItWorks properties"""
    try:
        result = db.session.execute(text("""
            SELECT id, property_order, title, address, level, unit_area, property_type, image_url
            FROM howitworks_properties 
            WHERE is_active = true
            ORDER BY property_order
        """)).fetchall()
        
        properties = []
        for row in result:
            properties.append({
                'id': row.id,
                'property_order': row.property_order,
                'title': row.title,
                'address': row.address,
                'level': row.level,
                'unit_area': row.unit_area,
                'property_type': row.property_type,
                'image_url': row.image_url
            })
        
        return jsonify({
            'success': True,
            'properties': properties
        }), 200
        
    except Exception as e:
        print(f"Error getting HowItWorks properties: {e}")
        return jsonify({'error': 'Failed to get HowItWorks properties'}), 500

@app.route('/api/howitworks/properties/<int:property_id>', methods=['GET'])
def get_howitworks_property(property_id):
    """Get specific HowItWorks property by ID"""
    try:
        result = db.session.execute(text("""
            SELECT id, property_order, title, address, level, unit_area, property_type, image_url
            FROM howitworks_properties 
            WHERE id = :property_id AND is_active = true
        """), {'property_id': property_id}).fetchone()
        
        if not result:
            return jsonify({'error': 'Property not found'}), 404
        
        return jsonify({
            'success': True,
            'property': {
                'id': result.id,
                'property_order': result.property_order,
                'title': result.title,
                'address': result.address,
                'level': result.level,
                'unit_area': result.unit_area,
                'property_type': result.property_type,
                'image_url': result.image_url
            }
        }), 200
        
    except Exception as e:
        print(f"Error getting HowItWorks property: {e}")
        return jsonify({'error': 'Failed to get HowItWorks property'}), 500

@app.route('/api/howitworks/properties/<int:property_id>', methods=['PUT'])
@require_auth
def update_howitworks_property(property_id):
    """Update HowItWorks property"""
    try:
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['title', 'address', 'property_type']
        for field in required_fields:
            if field not in data:
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        # Update property
        result = db.session.execute(text("""
            UPDATE howitworks_properties 
            SET title = :title, address = :address, level = :level, 
                unit_area = :unit_area, property_type = :property_type, 
                image_url = :image_url, updated_at = CURRENT_TIMESTAMP
            WHERE id = :property_id AND is_active = true
        """), {
            'property_id': property_id,
            'title': data['title'],
            'address': data['address'],
            'level': data.get('level'),
            'unit_area': data.get('unit_area'),
            'property_type': data['property_type'],
            'image_url': data.get('image_url')
        })
        
        if result.rowcount == 0:
            return jsonify({'error': 'Property not found'}), 404
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Property updated successfully'
        }), 200

    except Exception as e:
        print(f"Error updating HowItWorks property: {e}")
        db.session.rollback()
        return jsonify({'error': 'Failed to update HowItWorks property'}), 500

@app.route('/api/howitworks/properties', methods=['POST'])
@require_auth
def create_howitworks_property():
    """Create a new HowItWorks property"""
    try:
        data = request.get_json()

        # Validate required fields
        required_fields = ['title', 'address', 'property_type']
        for field in required_fields:
            if field not in data or data.get(field) is None:
                return jsonify({'error': f'Missing required field: {field}'}), 400

        # Determine property order; default append to end if not provided
        provided_order = data.get('property_order')
        if provided_order is None:
            result = db.session.execute(text("""
                SELECT COALESCE(MAX(property_order), 0) + 1 AS next_order
                FROM howitworks_properties
                WHERE is_active = true
            """)).fetchone()
            property_order = result.next_order if result else 1
        else:
            property_order = int(provided_order)

        insert_result = db.session.execute(text("""
            INSERT INTO howitworks_properties (property_order, title, address, level, unit_area, property_type, image_url, is_active, created_at, updated_at)
            VALUES (:property_order, :title, :address, :level, :unit_area, :property_type, :image_url, true, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
            RETURNING id
        """), {
            'property_order': property_order,
            'title': data['title'],
            'address': data['address'],
            'level': data.get('level'),
            'unit_area': data.get('unit_area'),
            'property_type': data['property_type'],
            'image_url': data.get('image_url')
        })

        new_id = insert_result.fetchone()[0]
        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'Property created successfully',
            'id': new_id,
            'property_order': property_order
        }), 201
    except Exception as e:
        print(f"Error creating HowItWorks property: {e}")
        db.session.rollback()
        return jsonify({'error': 'Failed to create HowItWorks property'}), 500

# Features Content API Endpoints

@app.route('/api/features/steps', methods=['GET'])
def get_features_steps():
    """Get all features steps"""
    try:
        # Query only columns that exist in the database to avoid schema mismatch errors
        result = db.session.execute(text("""
            SELECT id, step_number, step_title, step_description, step_image
            FROM features_steps 
            WHERE is_active = true
            ORDER BY step_number
        """)).fetchall()
        
        steps = []
        for row in result:
            steps.append({
                'id': row.id,
                'step_number': row.step_number,
                'step_title': row.step_title,
                'step_description': row.step_description,
                'step_image': row.step_image,
                'step_video': None  # Column doesn't exist in database
            })
        
        return jsonify({
            'success': True,
            'steps': steps
        }), 200
        
    except Exception as e:
        import traceback
        error_traceback = traceback.format_exc()
        print(f"Error getting features steps: {e}")
        print(f"Full traceback: {error_traceback}")
        return jsonify({'error': f'Failed to get features steps: {str(e)}'}), 500

@app.route('/api/features/steps/<int:step_id>', methods=['GET'])
def get_features_step(step_id):
    """Get specific features step by ID"""
    try:
        # Query only columns that exist in the database to avoid schema mismatch errors
        result = db.session.execute(text("""
            SELECT id, step_number, step_title, step_description, step_image
            FROM features_steps 
            WHERE id = :step_id AND is_active = true
        """), {'step_id': step_id}).fetchone()
        
        if not result:
            return jsonify({'error': 'Step not found'}), 404
        
        return jsonify({
            'success': True,
            'step': {
                'id': result.id,
                'step_number': result.step_number,
                'step_title': result.step_title,
                'step_description': result.step_description,
                'step_image': result.step_image,
                'step_video': None  # Column doesn't exist in database
            }
        }), 200
        
    except Exception as e:
        import traceback
        error_traceback = traceback.format_exc()
        print(f"Error getting features step: {e}")
        print(f"Full traceback: {error_traceback}")
        return jsonify({'error': f'Failed to get features step: {str(e)}'}), 500

@app.route('/api/features/steps/<int:step_id>', methods=['PUT'])
@require_auth
def update_features_step(step_id):
    """Update features step"""
    try:
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['step_title', 'step_description']
        for field in required_fields:
            if field not in data:
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        # Update step (step_video column doesn't exist in database)
        result = db.session.execute(text("""
            UPDATE features_steps 
            SET step_title = :step_title, step_description = :step_description, 
                step_image = :step_image, updated_at = CURRENT_TIMESTAMP
            WHERE id = :step_id AND is_active = true
        """), {
            'step_id': step_id,
            'step_title': data['step_title'],
            'step_description': data['step_description'],
            'step_image': data.get('step_image')
        })
        
        if result.rowcount == 0:
            return jsonify({'error': 'Step not found'}), 404
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Step updated successfully'
        }), 200
        
    except Exception as e:
        print(f"Error updating features step: {e}")
        db.session.rollback()
        return jsonify({'error': 'Failed to update step'}), 500

@app.route('/api/features/steps', methods=['POST'])
@require_auth
def create_features_step():
    """Create new features step"""
    try:
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['step_title', 'step_description']
        for field in required_fields:
            if field not in data:
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        # Get the next step number
        result = db.session.execute(text("""
            SELECT COALESCE(MAX(step_number), 0) + 1 as next_step_number
            FROM features_steps
        """)).fetchone()
        
        next_step_number = result.next_step_number
        
        # Create new step (step_video column doesn't exist in database)
        result = db.session.execute(text("""
            INSERT INTO features_steps (step_number, step_title, step_description, step_image)
            VALUES (:step_number, :step_title, :step_description, :step_image)
            RETURNING id
        """), {
            'step_number': next_step_number,
            'step_title': data['step_title'],
            'step_description': data['step_description'],
            'step_image': data.get('step_image', '')
        })
        
        new_step_id = result.fetchone().id
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Step created successfully',
            'step_id': new_step_id
        }), 201
        
    except Exception as e:
        print(f"Error creating features step: {e}")
        db.session.rollback()
        return jsonify({'error': 'Failed to create step'}), 500

@app.route('/api/features/steps/<int:step_id>', methods=['DELETE'])
@require_auth
def delete_features_step(step_id):
    """Delete features step (hard delete - permanently remove from database)"""
    try:
        # First check if step exists
        check_result = db.session.execute(text("""
            SELECT id FROM features_steps 
            WHERE id = :step_id AND is_active = true
        """), {'step_id': step_id}).fetchone()
        
        if not check_result:
            return jsonify({'error': 'Step not found'}), 404
        
        # Hard delete step - permanently remove from database
        result = db.session.execute(text("""
            DELETE FROM features_steps 
            WHERE id = :step_id
        """), {'step_id': step_id})
        
        if result.rowcount == 0:
            return jsonify({'error': 'Step not found'}), 404
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Step deleted successfully'
        }), 200
        
    except Exception as e:
        print(f"Error deleting features step: {e}")
        db.session.rollback()
        return jsonify({'error': 'Failed to delete step'}), 500

@app.route('/api/features/section-title', methods=['GET'])
def get_features_section_title():
    """Get features section title and tutorial video"""
    try:
        # Try to get the section, or create a default one if it doesn't exist
        section = FeaturesSection.query.filter_by(id=1).first()
        
        if not section:
            # Create a default section if it doesn't exist
            section = FeaturesSection(
                id=1,
                section_title='How it Works',
                tutorial_video_url=None
            )
            db.session.add(section)
            db.session.commit()
        
        return jsonify({
            'success': True,
            'section_title': section.section_title,
            'tutorial_video_url': section.tutorial_video_url
        }), 200
        
    except Exception as e:
        print(f"Error getting features section title: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Failed to get section title'}), 500

@app.route('/api/features/section-title', methods=['PUT'])
@require_auth
def update_features_section_title():
    """Update features section title and tutorial video"""
    try:
        data = request.get_json()
        
        if 'section_title' not in data:
            return jsonify({'error': 'Missing section_title field'}), 400
        
        # Get or create the section
        section = FeaturesSection.query.filter_by(id=1).first()
        
        if section:
            # Update existing section
            section.section_title = data['section_title']
            if 'tutorial_video_url' in data:
                section.tutorial_video_url = data['tutorial_video_url'] or None
        else:
            # Create new section
            section = FeaturesSection(
                id=1,
                section_title=data['section_title'],
                tutorial_video_url=data.get('tutorial_video_url')
            )
            db.session.add(section)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Section title updated successfully'
        }), 200
        
    except Exception as e:
        print(f"Error updating features section title: {e}")
        import traceback
        traceback.print_exc()
        db.session.rollback()
        return jsonify({'error': 'Failed to update section title'}), 500

@app.route('/api/features/upload-video', methods=['POST'])
@require_auth
def upload_features_video():
    """Upload features step tutorial video"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Check file type
        allowed_extensions = {'mp4', 'webm', 'ogg', 'mov'}
        if not ('.' in file.filename and file.filename.rsplit('.', 1)[1].lower() in allowed_extensions):
            return jsonify({'error': 'Invalid file type. Allowed: MP4, WEBM, OGG, MOV'}), 400
        
        # Generate unique filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"features_video_{timestamp}_{file.filename}"
        
        # Ensure directory exists
        upload_dir = os.path.join('admin', 'features', 'videos')
        os.makedirs(upload_dir, exist_ok=True)
        
        # Save file
        file_path = os.path.join(upload_dir, filename)
        file.save(file_path)
        
        # Return the URL path
        file_url = f"/admin/features/videos/{filename}"
        
        return jsonify({
            'success': True,
            'file_url': file_url,
            'filename': filename
        }), 200
        
    except Exception as e:
        print(f"Error uploading features video: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to upload video: {str(e)}'}), 500

@app.route('/api/features/tutorial-video', methods=['PUT'])
@require_auth
def update_tutorial_video():
    """Update tutorial video URL for the section"""
    try:
        data = request.get_json()
        
        if 'tutorial_video_url' not in data:
            return jsonify({'error': 'Missing tutorial_video_url field'}), 400
        
        # Update or insert tutorial video URL
        result = db.session.execute(text("""
            INSERT INTO features_section (id, tutorial_video_url, updated_at)
            VALUES (1, :tutorial_video_url, CURRENT_TIMESTAMP)
            ON CONFLICT (id) 
            DO UPDATE SET 
                tutorial_video_url = :tutorial_video_url,
                updated_at = CURRENT_TIMESTAMP
        """), {'tutorial_video_url': data['tutorial_video_url']})
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Tutorial video URL updated successfully'
        }), 200
        
    except Exception as e:
        print(f"Error updating tutorial video: {e}")
        db.session.rollback()
        return jsonify({'error': 'Failed to update tutorial video URL'}), 500

@app.route('/admin/features/videos/<filename>')
def serve_features_video(filename):
    """Serve features tutorial videos"""
    try:
        upload_dir = os.path.join(os.getcwd(), 'admin', 'features', 'videos')
        file_path = os.path.join(upload_dir, filename)
        
        if not os.path.exists(file_path):
            print(f"File not found: {file_path}")
            return jsonify({'error': 'File not found'}), 404
        
        response = send_from_directory(upload_dir, filename)
        # Add CORS headers
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Methods'] = 'GET'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        return response
    except Exception as e:
        print(f"Error serving features video: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Error serving file'}), 500

# Team API Endpoints

@app.route('/api/team/section', methods=['GET'])
def get_team_section():
    """Get team section details"""
    try:
        team_section = TeamSection.query.filter_by(id=1).first()
        
        if team_section:
            section_title = team_section.section_title
            section_subtitle = team_section.section_subtitle
        else:
            section_title = 'Our Team'
            section_subtitle = 'Worem ipsum dolor sit amet, consectetur adipiscing elit. Nunc vulputate libero et velit interdum, ac aliquet odio mattis.'
        
        return jsonify({
            'success': True,
            'section_title': section_title,
            'section_subtitle': section_subtitle
        }), 200
        
    except Exception as e:
        print(f"Error getting team section: {e}")
        return jsonify({'error': 'Failed to get team section'}), 500

@app.route('/api/team/section', methods=['PUT'])
@require_auth
def update_team_section():
    """Update team section details"""
    try:
        data = request.get_json()
        
        if 'section_title' not in data:
            return jsonify({'error': 'Missing section_title field'}), 400
        
        # Get or create team section
        team_section = TeamSection.query.filter_by(id=1).first()
        
        if team_section:
            # Update existing
            team_section.section_title = data['section_title']
            team_section.section_subtitle = data.get('section_subtitle', '')
            team_section.updated_at = datetime.utcnow()
        else:
            # Create new
            team_section = TeamSection(
                id=1,
                section_title=data['section_title'],
                section_subtitle=data.get('section_subtitle', '')
            )
            db.session.add(team_section)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Team section updated successfully'
        }), 200
        
    except Exception as e:
        print(f"Error updating team section: {e}")
        db.session.rollback()
        return jsonify({'error': 'Failed to update team section'}), 500

@app.route('/api/team/members', methods=['GET'])
def get_team_members():
    """Get all team members"""
    try:
        members = TeamMember.query.filter_by(is_active=True).order_by(TeamMember.display_order, TeamMember.created_at).all()
        
        members_list = []
        for member in members:
            # Convert relative image URL to full URL
            image_url = member.image_url
            if image_url and not image_url.startswith('http'):
                image_url = f"http://localhost:5001{image_url}"
            
            members_list.append({
                'id': member.id,
                'name': member.name,
                'role': member.role,
                'description': member.description,
                'image_url': image_url,
                'social_links': member.social_links or {},
                'display_order': member.display_order
            })
        
        return jsonify({
            'success': True,
            'members': members_list
        }), 200
        
    except Exception as e:
        print(f"Error getting team members: {e}")
        return jsonify({'error': 'Failed to get team members'}), 500

@app.route('/api/team/members/<int:member_id>', methods=['GET'])
def get_team_member(member_id):
    """Get specific team member by ID"""
    try:
        member = TeamMember.query.filter_by(id=member_id, is_active=True).first()
        
        if not member:
            return jsonify({'error': 'Team member not found'}), 404
        
        # Convert relative image URL to full URL
        image_url = member.image_url
        if image_url and not image_url.startswith('http'):
            image_url = f"http://localhost:5001{image_url}"
        
        return jsonify({
            'success': True,
            'member': {
                'id': member.id,
                'name': member.name,
                'role': member.role,
                'description': member.description,
                'image_url': image_url,
                'social_links': member.social_links or {},
                'display_order': member.display_order
            }
        }), 200
        
    except Exception as e:
        print(f"Error getting team member: {e}")
        return jsonify({'error': 'Failed to get team member'}), 500

@app.route('/api/team/members', methods=['POST'])
@require_auth
def create_team_member():
    """Create new team member"""
    try:
        data = request.get_json()
        print(f"Received team member data: {data}")
        
        # Validate required fields
        required_fields = ['name', 'role']
        for field in required_fields:
            if field not in data:
                print(f"Missing required field: {field}")
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        # Get the next display order
        print("Getting next display order...")
        last_member = TeamMember.query.order_by(TeamMember.display_order.desc()).first()
        next_order = (last_member.display_order + 1) if last_member else 1
        print(f"Next display order: {next_order}")
        
        # Create new team member using ORM
        print("Creating team member...")
        new_member = TeamMember(
            name=data['name'],
            role=data['role'],
            description=data.get('description', ''),
            image_url=data.get('image_url', ''),  # This will be a file path, not base64
            social_links=data.get('social_links', {}),
            display_order=next_order
        )
        
        db.session.add(new_member)
        db.session.commit()
        
        print(f"Created team member with ID: {new_member.id}")
        print("Team member committed to database")
        
        return jsonify({
            'success': True,
            'message': 'Team member created successfully',
            'member_id': new_member.id
        }), 201
        
    except Exception as e:
        print(f"Error creating team member: {e}")
        db.session.rollback()
        return jsonify({'error': 'Failed to create team member'}), 500

@app.route('/api/team/members/<int:member_id>', methods=['PUT'])
@require_auth
def update_team_member(member_id):
    """Update team member"""
    try:
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['name', 'role']
        for field in required_fields:
            if field not in data:
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        # Get team member using ORM
        member = TeamMember.query.filter_by(id=member_id, is_active=True).first()
        
        if not member:
            return jsonify({'error': 'Team member not found'}), 404
        
        # Update team member using ORM
        member.name = data['name']
        member.role = data['role']
        member.description = data.get('description', '')
        member.image_url = data.get('image_url', '')
        member.social_links = data.get('social_links', {})
        member.display_order = data.get('display_order', 0)
        member.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Team member updated successfully'
        }), 200
        
    except Exception as e:
        print(f"Error updating team member: {e}")
        db.session.rollback()
        return jsonify({'error': 'Failed to update team member'}), 500

@app.route('/api/team/members/<int:member_id>', methods=['DELETE'])
@require_auth
def delete_team_member(member_id):
    """Delete team member (hard delete)"""
    try:
        # Get team member using ORM
        member = TeamMember.query.filter_by(id=member_id, is_active=True).first()
        
        if not member:
            return jsonify({'error': 'Team member not found'}), 404
        
        # Hard delete team member using ORM
        db.session.delete(member)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Team member deleted successfully'
        }), 200
        
    except Exception as e:
        print(f"Error deleting team member: {e}")
        db.session.rollback()
        return jsonify({'error': 'Failed to delete team member'}), 500

# Upload team member profile picture
@app.route('/api/team/members/upload-profile-picture', methods=['POST'])
@require_auth
def upload_team_member_profile_picture():
    """Upload team member profile picture"""
    print("=== Team Member Profile Picture Upload Debug ===")
    print(f"Request method: {request.method}")
    print(f"Request files: {list(request.files.keys())}")
    print(f"Request form: {list(request.form.keys())}")
    
    # Get user from authenticated token
    user_id = request.user['user_id']
    print(f"User ID: {user_id}")
    
    user = User.query.get(user_id)
    if not user:
        print("User not found")
        return jsonify({'error': 'User not found'}), 404
    
    # Check if user is admin
    if user.user_type != 'admin':
        print("User is not an admin")
        return jsonify({'error': 'Access denied. Profile picture upload only available for admin users.'}), 403
    
    # Check if file was uploaded
    if 'profile_picture' not in request.files:
        print("No profile_picture in request.files")
        return jsonify({'error': 'No profile picture file provided'}), 400
    
    file = request.files['profile_picture']
    print(f"File name: {file.filename}")
    print(f"File content type: {file.content_type}")
    
    # Check if file is empty
    if file.filename == '':
        print("File filename is empty")
        return jsonify({'error': 'No file selected'}), 400
    
    # Check file type
    allowed_extensions = {'png', 'jpg', 'jpeg', 'gif'}
    if not file.filename.lower().endswith(tuple('.' + ext for ext in allowed_extensions)):
        return jsonify({'error': 'Invalid file type. Only PNG, JPG, JPEG, and GIF files are allowed.'}), 400
    
    try:
        # Generate unique filename
        import os
        from werkzeug.utils import secure_filename
        filename = secure_filename(file.filename)
        timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
        unique_filename = f"team_member_{timestamp}_{filename}"
        print(f"Generated filename: {unique_filename}")
        
        # Create upload directory if it doesn't exist
        upload_dir = os.path.join(os.getcwd(), 'admin', 'team')
        print(f"Upload directory: {upload_dir}")
        os.makedirs(upload_dir, exist_ok=True)
        print(f"Upload directory created/exists: {os.path.exists(upload_dir)}")
        
        # Save file
        file_path = os.path.join(upload_dir, unique_filename)
        print(f"Full file path: {file_path}")
        file.save(file_path)
        print(f"File saved successfully: {os.path.exists(file_path)}")
        
        # Create profile picture URL
        profile_url = f"/admin/team/{unique_filename}"
        print(f"Profile URL: {profile_url}")
        
        return jsonify({
            'message': 'Profile picture uploaded successfully',
            'profile_picture_url': profile_url
        }), 200
        
    except Exception as e:
        print(f"Error uploading profile picture: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Failed to upload profile picture'}), 500

# Serve uploaded team member profile pictures
@app.route('/admin/team/<filename>')
def serve_team_member_profile_picture(filename):
    """Serve uploaded team member profile pictures"""
    print(f"=== Serving Team Member Profile Picture ===")
    print(f"Requested filename: {filename}")
    
    upload_dir = os.path.join(os.getcwd(), 'admin', 'team')
    print(f"Upload directory: {upload_dir}")
    
    file_path = os.path.join(upload_dir, filename)
    print(f"Full file path: {file_path}")
    print(f"File exists: {os.path.exists(file_path)}")
    
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        return jsonify({'error': 'File not found'}), 404
    
    try:
        response = send_from_directory(upload_dir, filename)
        # Add CORS headers
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Methods'] = 'GET'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        return response
    except Exception as e:
        print(f"Error serving file: {e}")
        return jsonify({'error': 'Error serving file'}), 500

# Database migration endpoint
@app.route('/api/migrate/update-image-url-column', methods=['POST'])
def migrate_image_url_column():
    """Update team_members.image_url column to TEXT to support base64 data URLs"""
    try:
        # Execute the ALTER TABLE command
        db.session.execute(text("ALTER TABLE team_members ALTER COLUMN image_url TYPE TEXT;"))
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Successfully updated team_members.image_url to TEXT!'
        }), 200
        
    except Exception as e:
        print(f"Error updating image_url column: {e}")
        db.session.rollback()
        return jsonify({'error': f'Failed to update image_url column: {str(e)}'}), 500

# Contact Information API Endpoints

@app.route('/api/contact/info', methods=['GET'])
def get_contact_information():
    """Get contact information"""
    try:
        result = db.session.execute(text("""
            SELECT contact_type, contact_value, display_order
            FROM support_contact_info 
            WHERE is_active = true
            ORDER BY display_order, created_at
        """)).fetchall()
        
        contact_info = []
        for row in result:
            contact_info.append({
                'contact_type': row.contact_type,
                'contact_value': row.contact_value,
                'display_order': row.display_order
            })
        
        return jsonify({
            'success': True,
            'contact_info': contact_info
        }), 200
        
    except Exception as e:
        print(f"Error getting contact info: {e}")
        return jsonify({'error': 'Failed to get contact information'}), 500

@app.route('/api/contact/info', methods=['PUT'])
@require_auth
def update_contact_information():
    """Update contact information"""
    try:
        data = request.get_json()
        
        if 'contact_info' not in data:
            return jsonify({'error': 'Missing contact_info field'}), 400
        
        # Clear existing contact info
        db.session.execute(text("UPDATE support_contact_info SET is_active = false"))
        
        # Insert new contact info
        for item in data['contact_info']:
            if 'contact_type' in item and 'contact_value' in item:
                db.session.execute(text("""
                    INSERT INTO support_contact_info (contact_type, contact_value, display_order, is_active)
                    VALUES (:contact_type, :contact_value, :display_order, true)
                """), {
                    'contact_type': item['contact_type'],
                    'contact_value': item['contact_value'],
                    'display_order': item.get('display_order', 0)
                })
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Contact information updated successfully'
        }), 200
        
    except Exception as e:
        print(f"Error updating contact info: {e}")
        db.session.rollback()
        return jsonify({'error': 'Failed to update contact information'}), 500

# Legal Content API Endpoints
@app.route('/api/legal/<content_type>', methods=['GET'])
def get_legal_content_by_type(content_type):
    """Get legal content by type (disclaimer, privacy_policy, terms_of_use)"""
    try:
        # Validate content type
        valid_types = ['disclaimer', 'privacy_policy', 'terms_of_use']
        if content_type not in valid_types:
            return jsonify({'error': 'Invalid content type'}), 400
        
        legal_content = LegalContent.query.filter_by(
            content_type=content_type, 
            is_active=True
        ).order_by(LegalContent.created_at.desc()).first()
        
        if legal_content:
            return jsonify({
                'success': True,
                'data': {
                    'id': legal_content.id,
                    'title': legal_content.title,
                    'content': legal_content.content,
                    'version': legal_content.version,
                    'is_active': legal_content.is_active,
                    'created_at': legal_content.created_at.isoformat() if legal_content.created_at else None,
                    'updated_at': legal_content.updated_at.isoformat() if legal_content.updated_at else None
                }
            }), 200
        else:
            return jsonify({
                'success': True,
                'data': None,
                'message': f'No {content_type} content found'
            }), 200
            
    except Exception as e:
        print(f"Error getting legal content: {e}")
        return jsonify({'error': 'Failed to get legal content'}), 500

@app.route('/api/legal/<content_type>', methods=['PUT'])
@require_auth
def update_legal_content_by_type(content_type):
    """Update legal content by type"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        # Validate content type
        valid_types = ['disclaimer', 'privacy_policy', 'terms_of_use']
        if content_type not in valid_types:
            return jsonify({'error': 'Invalid content type'}), 400
        
        # Validate required fields
        if 'content' not in data:
            return jsonify({'error': 'Content is required'}), 400
        
        # Check if content exists
        existing = LegalContent.query.filter_by(
            content_type=content_type, 
            is_active=True
        ).order_by(LegalContent.created_at.desc()).first()
        
        if existing:
            # Update existing content
            existing.content = data['content']
            existing.title = data.get('title', existing.title)
            existing.updated_at = datetime.utcnow()
        else:
            # Create new content
            new_content = LegalContent(
                content_type=content_type,
                title=data.get('title', ''),
                content=data['content'],
                version=data.get('version', '1.0'),
                is_active=True
            )
            db.session.add(new_content)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'{content_type.replace("_", " ").title()} updated successfully'
        }), 200
        
    except Exception as e:
        print(f"Error updating legal content: {e}")
        db.session.rollback()
        return jsonify({'error': 'Failed to update legal content'}), 500

@app.route('/api/legal/all', methods=['GET'])
def get_all_legal_content_by_type():
    """Get all legal content"""
    try:
        results = LegalContent.query.filter_by(is_active=True).order_by(
            LegalContent.content_type, 
            LegalContent.created_at.desc()
        ).all()
        
        legal_content = {}
        for result in results:
            content_type = result.content_type
            if content_type not in legal_content:
                legal_content[content_type] = {
                    'id': result.id,
                    'title': result.title,
                    'content': result.content,
                    'version': result.version,
                    'is_active': result.is_active,
                    'created_at': result.created_at.isoformat() if result.created_at else None,
                    'updated_at': result.updated_at.isoformat() if result.updated_at else None
                }
        
        return jsonify({
            'success': True,
            'data': legal_content
        }), 200
        
    except Exception as e:
        print(f"Error getting all legal content: {e}")
        return jsonify({'error': 'Failed to get legal content'}), 500

# Reviews API Endpoints
@app.route('/api/reviews/verified', methods=['GET'])
def get_verified_reviews():
    """Get all verified user reviews"""
    try:
        reviews = db.session.execute(text("""
            SELECT ur.id, ur.user_id, ur.review_type, ur.rating, ur.review_text, 
                   ur.review_date, ur.is_verified, ur.admin_response, ur.admin_response_date,
                   u.full_name, u.profile_image_url
            FROM user_reviews ur
            LEFT JOIN users u ON ur.user_id = u.id
            WHERE ur.is_verified = true
            ORDER BY ur.review_date DESC
        """)).fetchall()
        
        reviews_data = []
        for review in reviews:
            reviews_data.append({
                'id': review[0],
                'user_id': review[1],
                'review_type': review[2],
                'rating': review[3],
                'review_text': review[4],
                'review_date': review[5].isoformat() if review[5] else None,
                'is_verified': review[6],
                'admin_response': review[7],
                'admin_response_date': review[8].isoformat() if review[8] else None,
                'author_name': review[9] or 'Anonymous User',
                'author_avatar': review[10] or '/assets/default-avatar.png',
                'author_role': 'User'  # Default role for user reviews
            })
        
        return jsonify({
            'success': True,
            'data': reviews_data
        }), 200
        
    except Exception as e:
        print(f"Error getting verified reviews: {e}")
        return jsonify({'error': 'Failed to get verified reviews'}), 500

# ==================== DATABASE SEEDING ====================

def seed_regions():
    """Seed the database with Singapore postal districts"""
    
    try:
        # Check if regions already exist
        if Region.query.count() > 0:
            print("Regions already exist, skipping seeding")
            return
        
        print("Seeding regions...")
        
        # Regions data
        regions_data = [
            ('01', '01, 02, 03, 04, 05, 06', 'Raffles Place, Cecil, Marina, People\'s Park'),
            ('02', '07, 08', 'Anson, Tanjong Pagar'),
            ('03', '14, 15, 16', 'Queenstown, Tiong Bahru'),
            ('04', '09, 10', 'Telok Blangah, Harbourfront'),
            ('05', '11, 12, 13', 'Pasir Panjang, Hong Leong Garden, Clementi New Town'),
            ('06', '17', 'High Street, Beach Road (part)'),
            ('07', '18, 19', 'Middle Road, Golden Mile'),
            ('08', '20, 21', 'Little India'),
            ('09', '22, 23', 'Orchard, Cairnhill, River Valley'),
            ('10', '24, 25, 26, 27', 'Ardmore, Bukit Timah, Holland Road, Tanglin'),
            ('11', '28, 29, 30', 'Watten Estate, Novena, Thomson'),
            ('12', '31, 32, 33', 'Balestier, Toa Payoh, Serangoon'),
            ('13', '34, 35, 36, 37', 'Macpherson, Braddell'),
            ('14', '38, 39, 40, 41', 'Geylang, Eunos'),
            ('15', '42, 43, 44, 45', 'Katong, Joo Chiat, Amber Road'),
            ('16', '46, 47, 48', 'Bedok, Upper East Coast, Eastwood, Kew Drive'),
            ('17', '49, 50, 81', 'Loyang, Changi'),
            ('18', '51, 52', 'Tampines, Pasir Ris'),
            ('19', '53, 54, 55, 82', 'Serangoon Garden, Hougang, Punggol'),
            ('20', '56, 57', 'Bishan, Ang Mo Kio'),
            ('21', '58, 59', 'Upper Bukit Timah, Clementi Park, Ulu Pandan'),
            ('22', '60, 61, 62, 63, 64', 'Jurong'),
            ('23', '65, 66, 67, 68', 'Hillview, Dairy Farm, Bukit Panjang, Choa Chu Kang'),
            ('24', '69, 70, 71', 'Lim Chu Kang, Tengah'),
            ('25', '72, 73', 'Kranji, Woodgrove'),
            ('26', '77, 78', 'Upper Thomson, Springleaf'),
            ('27', '75, 76', 'Yishun, Sembawang'),
            ('28', '79, 80', 'Seletar'),
        ]
        
        # Insert regions
        for district, sector, location in regions_data:
            region = Region(
                district=district,
                sector=sector,
                location=location,
                is_active=True
            )
            db.session.add(region)
        
        db.session.commit()
        print(f"✅ Successfully seeded {Region.query.count()} regions!")
        
    except Exception as e:
        print(f"❌ Error seeding regions: {e}")

def seed_subscription_plans():
    """Seed the database with initial subscription plans and important features"""
    
    try:
        # Check if data already exists
        if SubscriptionPlan.query.count() > 0:
            print("Subscription plans already exist, skipping seeding")
            return
        
        print("Seeding subscription plans and important features...")
        
        # Create Free Plan
        free_plan = SubscriptionPlan(
            plan_name='Free Plan',
            plan_type='free',
            monthly_price=0.00,
            yearly_price=0.00,
            description='Perfect for getting started with basic property insights',
            display_order=1
        )
        db.session.add(free_plan)
        db.session.flush()
        
        # Add Free Plan features
        free_features = [
            'Price Prediction',
            'Bookmark Prediction',
            'Bookmark Address'
        ]
        
        for i, feature in enumerate(free_features):
            feature_obj = SubscriptionPlanFeature(
                plan_id=free_plan.id,
                feature_name=feature,
                feature_description=f'Access to {feature.lower()} functionality',
                is_included=True,
                display_order=i + 1
            )
            db.session.add(feature_obj)
        
        # Create Premium Plan
        premium_plan = SubscriptionPlan(
            plan_name='Premium Plan',
            plan_type='premium',
            monthly_price=29.99,
            yearly_price=299.99,
            description='Advanced features for serious property investors',
            is_popular=True,
            display_order=2
        )
        db.session.add(premium_plan)
        db.session.flush()
        
        # Add Premium Plan features
        premium_features = [
            'Price Prediction',
            'Compare Prediction',
            'Bookmark Prediction',
            'Bookmark Address',
            'Download Report'
        ]
        
        for i, feature in enumerate(premium_features):
            feature_obj = SubscriptionPlanFeature(
                plan_id=premium_plan.id,
                feature_name=feature,
                feature_description=f'Advanced {feature.lower()} functionality',
                is_included=True,
                display_order=i + 1
            )
            db.session.add(feature_obj)
        
        # Create Agent Plan
        agent_plan = SubscriptionPlan(
            plan_name='Agent Plan',
            plan_type='agent',
            monthly_price=49.99,
            yearly_price=499.99,
            description='Professional tools for real estate agents',
            display_order=3
        )
        db.session.add(agent_plan)
        db.session.flush()
        
        # Add Agent Plan features
        agent_features = [
            'Agent Listing',
            'Agent Regions'
        ]
        
        for i, feature in enumerate(agent_features):
            feature_obj = SubscriptionPlanFeature(
                plan_id=agent_plan.id,
                feature_name=feature,
                feature_description=f'Professional {feature.lower()} management',
                is_included=True,
                display_order=i + 1
            )
            db.session.add(feature_obj)
        
        # Create Important Features for comparison
        important_features = [
            '24/7 Support',
            'Mobile App Access',
            'Email Notifications',
            'Data Export',
            'Advanced Analytics',
            'Priority Support',
            'API Access'
        ]
        
        for i, feature in enumerate(important_features):
            feature_obj = ImportantFeature(
                feature_name=feature,
                feature_description=f'Feature: {feature}',
                display_order=i + 1
            )
            db.session.add(feature_obj)
        
        # Commit all changes
        db.session.commit()
        
        print("✅ Successfully seeded subscription plans and important features!")
        print(f"Created {SubscriptionPlan.query.count()} subscription plans")
        print(f"Created {SubscriptionPlanFeature.query.count()} plan features")
        print(f"Created {ImportantFeature.query.count()} important features")
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error seeding database: {e}")

def seed_features_section():
    """Seed the features_section table with initial data"""
    
    try:
        # Check if data already exists
        if FeaturesSection.query.count() > 0:
            print("Features section already exists, skipping seeding")
            return
        
        print("Seeding features section...")
        
        # Create default features section
        features_section = FeaturesSection(
            id=1,
            section_title='How it Works',
            tutorial_video_url=None
        )
        db.session.add(features_section)
        db.session.commit()
        
        print("✅ Successfully seeded features section!")
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error seeding features section: {e}")

# Admin endpoint to manually seed data
@app.route('/api/admin/seed-subscription-plans', methods=['POST'])
@require_auth
def admin_seed_subscription_plans():
    """Admin endpoint to manually seed subscription plans"""
    try:
        # Check if user is admin
        user_id = request.user['user_id']
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        # Clear existing data first
        SubscriptionPlanFeature.query.delete()
        SubscriptionPlan.query.delete()
        ImportantFeature.query.delete()
        db.session.commit()
        
        # Seed new data
        seed_subscription_plans()
        
        return jsonify({
            'success': True,
            'message': 'Subscription plans seeded successfully'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'Failed to seed subscription plans'}), 500
    
# ============================
# ML Prediction Endpoint
# ============================
from flask import Flask, request, jsonify
import joblib
import numpy as np
import os

# Load the model only once when the app starts
MODEL_PATH = os.path.join(os.path.dirname(__file__), 'real_estate_model_enhanced_20251003_0044.pkl')
model = joblib.load(MODEL_PATH)

@app.route('/api/predict', methods=['POST'])
def predict():
    """
    Expects JSON like:
    {
        "features": [1200, 3, 2, 1, 5]   # numerical input for model
    }
    """
    try:
        data = request.get_json()
        if not data or "features" not in data:
            return jsonify({"error": "Missing 'features' in request"}), 400

        features = np.array(data["features"]).reshape(1, -1)
        prediction = model.predict(features)[0]

        return jsonify({
            "status": "success",
            "prediction": float(prediction)
        })
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500


# -------------------------
# ML PREDICTION ENDPOINTS
# -------------------------

# Global cache for ML data to avoid reloading on every request
_ml_cache = {
    'df_industrial': None,
    'df_commercial_clean': None,
    'df_office_rental': None,
    'df_retail_rental': None,
    'all_addresses': None,
    'ml_functions': None,
    'postal_districts': None,
    'last_cleanup': None
}

# Cache cleanup function to prevent memory leaks
def cleanup_ml_cache():
    """Clean up ML cache if it's been too long since last cleanup"""
    import time
    current_time = time.time()
    
    # Clean up cache every 30 minutes
    if _ml_cache['last_cleanup'] is None or (current_time - _ml_cache['last_cleanup']) > 1800:
        _ml_cache['df_industrial'] = None
        _ml_cache['df_commercial_clean'] = None
        _ml_cache['df_office_rental'] = None
        _ml_cache['df_retail_rental'] = None
        _ml_cache['all_addresses'] = None
        _ml_cache['postal_districts'] = None
        _ml_cache['last_cleanup'] = current_time

def run_ml_prediction(property_data):
    """Run ML prediction using cached data approach"""
    try:
        import time
        start_time = time.time()
        
        # Print separator for new prediction
        print("\n" + "="*80)
        print("🔍 NEW PREDICTION REQUEST")
        print("="*80)
        print(f"📋 Property Data: {property_data}")
        print("-"*80)
        
        # Clean up cache if needed
        cleanup_ml_cache()
        
        # Add the machinelearning directory to the path
        current_dir = os.path.dirname(os.path.abspath(__file__))
        ml_dir = os.path.join(current_dir, '..', 'machinelearning')
        sys.path.insert(0, ml_dir)
        
        # Import the multi-model ML functions directly (cache after first import)
        if _ml_cache['ml_functions'] is None:
            try:
                from multi_model_predictor import get_multi_model_predictor
                from ml_predictor_enhanced import predict_for_propertycard, get_unique_addresses, get_enhanced_predictor
                
                # Initialize the multi-model predictor and load all models
                import time
                load_start = time.time()
                multi_predictor = get_multi_model_predictor()
                if multi_predictor.is_loaded:
                    load_time = time.time() - load_start
                    print(f"✅ Multi-model ML predictor loaded successfully in {load_time:.2f} seconds")
                else:
                    print("⚠️ Multi-model predictor failed to load, will use enhanced predictor as fallback")
                
                _ml_cache['ml_functions'] = {
                    'multi_predictor': multi_predictor,
                    'predict_for_propertycard': predict_for_propertycard,
                    'get_unique_addresses': get_unique_addresses,
                    'get_enhanced_predictor': get_enhanced_predictor
                }
                print("✅ Multi-model ML functions imported successfully")
            except ImportError as e:
                print(f"❌ Failed to import multi-model ML functions: {e}")
                import traceback
                traceback.print_exc()
                return {
                    'success': False,
                    'error': f'Failed to import multi-model ML predictor module: {e}'
                }
        
        # Load datasets (cache after first load)
        if _ml_cache['df_industrial'] is None:
            import pandas as pd
            
            # Load industrial data with optimized settings
            industrial_path = os.path.join(ml_dir, 'industrial_2022toSep2025.csv')
            try:
                import time
                start_time = time.time()
                # Load all columns and map to expected column names
                df = pd.read_csv(industrial_path)
                load_time = time.time() - start_time
                print(f"📊 Industrial data loaded in {load_time:.2f} seconds ({len(df)} records)")
                
                # Map column names to what ML predictor expects
                column_mapping = {
                    'Street Name': 'street_name',
                    'Project Name': 'project_name', 
                    'Planning Area': 'planning_area',
                    'Property Type': 'property_type',
                    'Area': 'area',
                    'Contract Date': 'contract_date',
                    'Price': 'price',
                    '$psm': 'unit_price_psm',  # Map PSM column
                    'Postal District': 'postal_district'
                }
                
                # Rename columns to match ML predictor expectations
                df = df.rename(columns=column_mapping)
                
                # Clean up string columns - replace NaN with 'N.A.'
                string_columns = ['street_name', 'project_name', 'planning_area', 'property_type']
                for col in string_columns:
                    if col in df.columns:
                        df[col] = df[col].fillna('N.A.').astype(str)
                
                # Clean up price column (remove $ and commas)
                if 'price' in df.columns:
                    df['price'] = df['price'].astype(str).str.replace('$', '', regex=False).str.replace(',', '', regex=False)
                    df['price'] = pd.to_numeric(df['price'], errors='coerce')
                
                # Clean up unit price PSM column (remove $ and commas)
                if 'unit_price_psm' in df.columns:
                    df['unit_price_psm'] = df['unit_price_psm'].astype(str).str.replace('$', '', regex=False).str.replace(',', '', regex=False)
                    df['unit_price_psm'] = pd.to_numeric(df['unit_price_psm'], errors='coerce')
                
                # Clean up area column (convert from sqm to sqft for industrial data)
                if 'area' in df.columns:
                    df['area'] = pd.to_numeric(df['area'], errors='coerce')
                    df['area'] = df['area'] * 10.764  # Convert sqm to sqft
                
                # Parse contract date - handle MM/DD/YYYY format for industrial data
                if 'contract_date' in df.columns:
                    df['contract_date'] = pd.to_datetime(df['contract_date'], format='%m/%d/%Y', errors='coerce')
                
                # Ensure postal_district is numeric for proper filtering
                if 'postal_district' in df.columns:
                    df['postal_district'] = pd.to_numeric(df['postal_district'], errors='coerce')
                    print(f"✅ Converted postal_district to numeric: {df['postal_district'].dtype}")
                
                _ml_cache['df_industrial'] = df
                print(f"✅ Successfully loaded industrial data: {len(df)} rows, columns: {list(df.columns)}")
            except Exception as e:
                print(f"❌ Failed to load industrial data: {e}")
                return {
                    'success': False,
                    'error': f'Failed to load industrial data: {e}'
                }
        
        # Load commercial data if not cached
        if _ml_cache['df_commercial_clean'] is None:
            try:
                import pandas as pd  # Ensure pandas is imported
                import time
                start_time = time.time()
                commercial_path = os.path.join(ml_dir, 'commercial(everything teeco)', 'CommercialTransaction20250917124317.csv')
                df_commercial = pd.read_csv(commercial_path)
                load_time = time.time() - start_time
                print(f"📊 Commercial data loaded in {load_time:.2f} seconds ({len(df_commercial)} records)")
                
                # Map commercial column names to expected format
                commercial_mapping = {
                    'Project Name': 'project_name',
                    'Street Name': 'street_name',
                    'Property Type': 'property_type',
                    'Transacted Price ($)': 'price',
                    'Area (SQFT)': 'area',
                    'Unit Price ($ PSF)': 'unit_price_psf',  # Map PSF column
                    'Sale Date': 'contract_date',
                    'Postal District': 'postal_district'
                }
                
                # Rename columns
                df_commercial = df_commercial.rename(columns=commercial_mapping)
                
                # Clean up string columns - replace NaN with 'N.A.'
                string_columns = ['street_name', 'project_name', 'property_type']
                for col in string_columns:
                    if col in df_commercial.columns:
                        df_commercial[col] = df_commercial[col].fillna('N.A.').astype(str)
                
                # Clean up price column
                if 'price' in df_commercial.columns:
                    df_commercial['price'] = df_commercial['price'].astype(str).str.replace('$', '', regex=False).str.replace(',', '', regex=False)
                    df_commercial['price'] = pd.to_numeric(df_commercial['price'], errors='coerce')
                
                # Clean up unit price PSF column (remove $ and commas)
                if 'unit_price_psf' in df_commercial.columns:
                    df_commercial['unit_price_psf'] = df_commercial['unit_price_psf'].astype(str).str.replace('$', '', regex=False).str.replace(',', '', regex=False)
                    df_commercial['unit_price_psf'] = pd.to_numeric(df_commercial['unit_price_psf'], errors='coerce')
                
                # Clean up area column (keep in sqft for commercial data)
                if 'area' in df_commercial.columns:
                    df_commercial['area'] = pd.to_numeric(df_commercial['area'], errors='coerce')
                    # Keep in sqft (no conversion needed)
                
                # Parse sale date - handle custom format like "Sept-25", "Aug-25"
                if 'contract_date' in df_commercial.columns:
                    def parse_custom_date(date_str):
                        try:
                            if pd.isna(date_str):
                                return pd.NaT
                            
                            # Handle formats like "Sept-25", "Aug-25"
                            if isinstance(date_str, str) and '-' in date_str:
                                month_str, year_str = date_str.split('-')
                                month_map = {
                                    'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04',
                                    'May': '05', 'Jun': '06', 'Jul': '07', 'Aug': '08',
                                    'Sep': '09', 'Sept': '09', 'Oct': '10', 'Nov': '11', 'Dec': '12'
                                }
                                month_num = month_map.get(month_str, '01')
                                year = f"20{year_str}" if len(year_str) == 2 else year_str
                                return pd.to_datetime(f"{year}-{month_num}-01")
                            else:
                                return pd.to_datetime(date_str, errors='coerce')
                        except:
                            return pd.NaT
                    
                    df_commercial['contract_date'] = df_commercial['contract_date'].apply(parse_custom_date)
                
                # Add planning area (set to 'Unknown' for commercial data)
                df_commercial['planning_area'] = 'Unknown'
                
                # Ensure postal_district is numeric for proper filtering
                if 'postal_district' in df_commercial.columns:
                    df_commercial['postal_district'] = pd.to_numeric(df_commercial['postal_district'], errors='coerce')
                    print(f"✅ Converted postal_district to numeric: {df_commercial['postal_district'].dtype}")
                
                _ml_cache['df_commercial_clean'] = df_commercial.dropna(subset=['price', 'area'])
                print(f"✅ Successfully loaded commercial data: {len(df_commercial)} rows, columns: {list(df_commercial.columns)}")
            except Exception as e:
                print(f"❌ Warning: Could not load commercial data: {e}")
                _ml_cache['df_commercial_clean'] = None
        
        # Load rental data if not cached
        if _ml_cache['df_office_rental'] is None:
            try:
                import pandas as pd  # Ensure pandas is imported
                office_rental_path = os.path.join(ml_dir, 'commercial(rental)', 'CommercialOfficeRental.csv')
                df_office_rental = pd.read_csv(office_rental_path)
                _ml_cache['df_office_rental'] = df_office_rental
                print(f"✅ Successfully loaded office rental data: {len(df_office_rental)} rows")
            except Exception as e:
                print(f"❌ Warning: Could not load office rental data: {e}")
                import traceback
                print(f"❌ Traceback: {traceback.format_exc()}")
                _ml_cache['df_office_rental'] = None
        
        if _ml_cache['df_retail_rental'] is None:
            try:
                import pandas as pd  # Ensure pandas is imported
                retail_rental_path = os.path.join(ml_dir, 'commercial(rental)', 'CommercialRetailRental.csv')
                df_retail_rental = pd.read_csv(retail_rental_path)
                _ml_cache['df_retail_rental'] = df_retail_rental
                print(f"✅ Successfully loaded retail rental data: {len(df_retail_rental)} rows")
            except Exception as e:
                print(f"❌ Warning: Could not load retail rental data: {e}")
                import traceback
                print(f"❌ Traceback: {traceback.format_exc()}")
                _ml_cache['df_retail_rental'] = None
        
        # Get all addresses (cache after first calculation) - Optimized for speed
        # Load postal districts if not cached
        if _ml_cache['postal_districts'] is None or (isinstance(_ml_cache['postal_districts'], dict) and len(_ml_cache['postal_districts']) == 0):
            try:
                import pandas as pd
                postal_districts_path = os.path.join(ml_dir, 'sg cordinates', 'sg_postal_districts.csv')
                postal_districts_df = pd.read_csv(postal_districts_path)
                
                # Create postal district mapping
                postal_districts = {}
                for _, row in postal_districts_df.iterrows():
                    district = row['Postal District']
                    sectors = str(row['Postal Sector'])  # Ensure it's a string
                    
                    # Parse sectors (e.g., "01, 02, 03" or "17")
                    if ',' in sectors:
                        sector_list = [s.strip() for s in sectors.split(',')]
                    else:
                        sector_list = [sectors.strip()]
                    
                    for sector in sector_list:
                        postal_districts[sector] = district
                
                _ml_cache['postal_districts'] = postal_districts
                print(f"✅ Loaded postal districts: {len(postal_districts)} sectors mapped")
                print(f"📋 Sample mappings: 04 -> {postal_districts.get('04')}, 27 -> {postal_districts.get('27')}")
            except Exception as e:
                print(f"⚠️ Failed to load postal districts: {e}")
                import traceback
                traceback.print_exc()
                _ml_cache['postal_districts'] = {}

        if _ml_cache['all_addresses'] is None:
            try:
                industrial_addresses = _ml_cache['ml_functions']['get_unique_addresses'](_ml_cache['df_industrial'], "Industrial")
                commercial_addresses = _ml_cache['ml_functions']['get_unique_addresses'](_ml_cache['df_commercial_clean'], "Commercial") if _ml_cache['df_commercial_clean'] is not None else []
                # Limit addresses for faster processing
                all_addresses = industrial_addresses + commercial_addresses
                _ml_cache['all_addresses'] = all_addresses[:1000]  # Limit to first 1000 addresses for speed
            except Exception as e:
                return {
                    'success': False,
                    'error': f'Failed to get addresses: {e}'
                }
        
        # Try to use multi-model predictor for direct predictions first
        multi_predictor = _ml_cache['ml_functions'].get('multi_predictor')
        direct_predictions = None
        
        if multi_predictor and multi_predictor.is_loaded:
            try:
                # Convert floor area from sqft to sqm
                # Both industrial and commercial models expect "Area (SQM)" as input
                floor_area_sqft = float(property_data.get('floorArea', 0))
                floor_area_sqm = floor_area_sqft * 0.092903  # Convert sqft to sqm (1 sqft = 0.092903 sqm)
                
                # Get predictions using multi-model predictor
                # Note: Industrial model predicts TOTAL PRICE directly
                #       Commercial model may predict PSM/PSF (handled internally)
                predictions = multi_predictor.predict_both(
                    address=property_data.get('address', ''),
                    property_type=property_data.get('propertyType', 'Office'),
                    area_sqm=floor_area_sqm,  # Model expects area in SQM
                    level=property_data.get('level', 'Ground Floor'),
                    unit=property_data.get('unit', 'N/A'),
                    tenure="Freehold"
                )
                
                # Accept predictions if at least one is available AND valid (not None)
                if predictions.get('sales_price') is not None or predictions.get('rental_price') is not None:
                    # Check if sales_price is valid (positive)
                    if predictions.get('sales_price') is not None and predictions['sales_price'] > 0:
                        direct_predictions = predictions
                        sales_str = f"${predictions['sales_price']:,.2f}" if predictions.get('sales_price') else "N/A"
                        rental_str = f"${predictions['rental_price']:,.2f}/month" if predictions.get('rental_price') else "N/A"
                        print(f"✅ Multi-model predictions: Sales={sales_str}, Rental={rental_str}")
                    else:
                        print(f"⚠️ Multi-model sales prediction was invalid (None or negative), using fallback")
                else:
                    print(f"⚠️ Multi-model predictions returned None for both sales and rental, using fallback")
            except Exception as e:
                print(f"⚠️ Multi-model prediction failed: {e}, falling back to enhanced predictor")
                import traceback
                traceback.print_exc()
        
        # Run prediction using enhanced ML predictor for full analysis
        try:
            # Ensure postal_districts is loaded
            postal_districts_for_prediction = _ml_cache.get('postal_districts') or {}
            if len(postal_districts_for_prediction) == 0:
                print(f"⚠️ WARNING: postal_districts is empty, district filtering will not work!")
            else:
                print(f"📋 Passing postal_districts with {len(postal_districts_for_prediction)} sectors to predictor")
            
            property_data_result, comparison_data_result, matched_address = _ml_cache['ml_functions']['predict_for_propertycard'](
                frontend_property_data=property_data,
                all_addresses=_ml_cache['all_addresses'],
                df_industrial=_ml_cache['df_industrial'],
                df_commercial=_ml_cache['df_commercial_clean'],
                postal_districts=postal_districts_for_prediction,
                df_retail_rental=_ml_cache['df_retail_rental'],
                df_office_rental=_ml_cache['df_office_rental']
            )
            
            # Only override with multi-model predictions if comparison_data doesn't already have adjusted values
            # (The analyze_commercial_market/analyze_industrial_market functions may have already adjusted
            # the prediction based on market data, so we should respect that adjustment)
            if direct_predictions:
                # Check if comparison_data already has a validated/adjusted price
                # If it was adjusted, it means market data validation found the ML prediction was off
                current_sales_str = comparison_data_result.get('estimatedSalesPrice', '')
                
                # Only override if comparison_data is using fallback (contains "N/A" or seems unadjusted)
                # OR if the current value seems to be from simple estimation (very round numbers)
                should_override = False
                if not current_sales_str or current_sales_str == 'N/A' or 'Loading' in current_sales_str:
                    should_override = True
                else:
                    # If comparison_data already has a properly formatted price, keep it
                    # (it may have been adjusted based on market data)
                    should_override = False
                    print(f"ℹ️ Keeping adjusted prediction from market data validation: {current_sales_str}")
                
                # Format sales price if we should override
                if should_override and direct_predictions.get('sales_price') is not None:
                    sales_price = direct_predictions['sales_price']
                    
                    # Ensure sales price is positive and reasonable
                    if sales_price < 0:
                        print(f"⚠️ Warning: Negative sales price detected: ${sales_price:,.2f}, using absolute value")
                        sales_price = abs(sales_price)
                    
                    # Minimum reasonable price check
                    floor_area_sqft = float(property_data.get('floorArea', 0))
                    if floor_area_sqft > 0:
                        min_price = floor_area_sqft * 100  # Minimum $100 PSF
                        if sales_price < min_price:
                            print(f"⚠️ Warning: Sales price ${sales_price:,.2f} below minimum ${min_price:,.2f}, using minimum")
                            sales_price = min_price
                    
                    if sales_price >= 1000000:
                        formatted_sales = f"${sales_price/1000000:.2f}M"
                    elif sales_price >= 1000:
                        formatted_sales = f"${sales_price/1000:.0f}k"
                    else:
                        formatted_sales = f"${sales_price:,.0f}"
                    comparison_data_result['estimatedSalesPrice'] = formatted_sales
                    print(f"✅ Updated sales price with multi-model prediction: {formatted_sales}")
                
                # Format rental price if available - only override if not already adjusted
                current_rental_str = comparison_data_result.get('estimatedRentalPrice', '')
                
                # Only override if comparison_data is using fallback (contains "N/A" or seems unadjusted)
                should_override_rental = False
                if not current_rental_str or current_rental_str == 'N/A' or 'Loading' in current_rental_str:
                    should_override_rental = True
                else:
                    # If comparison_data already has a properly formatted price, keep it
                    # (ML rental prediction is used directly without adjustment)
                    should_override_rental = False
                    print(f"ℹ️ Keeping ML rental prediction from enhanced predictor: {current_rental_str}")
                
                if should_override_rental and direct_predictions.get('rental_price') is not None:
                    rental_price = direct_predictions['rental_price']
                    
                    # Ensure rental price is positive
                    if rental_price < 0:
                        print(f"⚠️ Warning: Negative rental price detected: ${rental_price:,.2f}, using absolute value")
                        rental_price = abs(rental_price)
                    
                    # Minimum reasonable rental check
                    floor_area_sqft = float(property_data.get('floorArea', 0))
                    if floor_area_sqft > 0:
                        min_rental = floor_area_sqft * 1  # Minimum $1 PSF/month
                        if rental_price < min_rental:
                            print(f"⚠️ Warning: Rental price ${rental_price:,.2f} below minimum ${min_rental:,.2f}, using minimum")
                            rental_price = min_rental
                    
                    if rental_price >= 1000:
                        formatted_rental = f"${rental_price/1000:.1f}k/month"
                    else:
                        formatted_rental = f"${rental_price:,.0f}/month"
                    comparison_data_result['estimatedRentalPrice'] = formatted_rental
                    print(f"✅ Updated rental price with multi-model prediction: {formatted_rental}")
            
            # Return results
            total_time = time.time() - start_time
            print(f"⏱️ Total prediction time: {total_time:.2f} seconds")
            
            # Print separator at end of successful prediction
            print("-"*80)
            print(f"✅ PREDICTION COMPLETE (Time: {total_time:.2f}s)")
            print("="*80 + "\n")
            
            return {
                'success': True,
                'property_data': property_data_result,
                'comparison_data': comparison_data_result,
                'matched_address': matched_address
            }
            
        except Exception as e:
            print(f"❌ Prediction failed: {e}")
            import traceback
            print(f"❌ Traceback: {traceback.format_exc()}")
            print("-"*80)
            print("❌ PREDICTION FAILED")
            print("="*80 + "\n")
            return {
                'success': False,
                'error': f'Prediction failed: {e}'
            }
            
    except Exception as e:
        print(f"❌ ML prediction error: {e}")
        import traceback
        print(f"❌ Traceback: {traceback.format_exc()}")
        print("-"*80)
        print("❌ PREDICTION ERROR")
        print("="*80 + "\n")
        return {
            'success': False,
            'error': f'ML prediction error: {str(e)}'
        }

@app.route('/api/predict-price', methods=['POST'])
def predict_price():
    """Generate ML-based price prediction for property"""
    try:
        # Get current user
        token = request.headers.get('Authorization')
        print(f"🔍 Authorization header: {token}")
        
        if not token or not token.startswith('Bearer '):
            print("❌ No Bearer token found")
            return jsonify({'error': 'Authorization token required'}), 401
        
        token = token.split(' ')[1]
        print(f"🔍 JWT token: {token[:50]}...")
        print(f"🔍 Secret key: {app.config.get('SECRET_KEY', 'NOT_SET')[:20]}...")
        
        try:
            payload = jwt.decode(token, JWT_SECRET_KEY, algorithms=['HS256'])
            user_id = payload['user_id']
            print(f"✅ JWT decoded successfully, user_id: {user_id}")
        except jwt.ExpiredSignatureError as e:
            print(f"❌ Token expired: {e}")
            return jsonify({'error': 'Token expired'}), 401
        except jwt.InvalidTokenError as e:
            print(f"❌ Invalid token: {e}")
            return jsonify({'error': 'Invalid token'}), 401
        
        # Get user and check if they're a free user with limit reached
        user = User.query.get(user_id)
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        # Debug: Print user info
        print(f"🔍 User check - user_id: {user_id}, user_type: {user.user_type}")
        
        # Check if free user has exceeded prediction limit
        if user.user_type and user.user_type.lower() == 'free':
            prediction_count = PricePrediction.query.filter_by(user_id=user_id).count()
            print(f"🔍 Free user prediction count: {prediction_count}/3")
            
            if prediction_count >= 3:
                print(f"❌ Blocking prediction - limit reached: {prediction_count} >= 3")
                return jsonify({
                    'error': 'prediction_limit_reached',
                    'message': 'You have reached your free prediction limit (3 searches). Please upgrade to Premium to continue using price predictions.',
                    'current_count': prediction_count,
                    'limit': 3,
                    'upgrade_required': True
                }), 403
            else:
                print(f"✅ Allowing prediction - count: {prediction_count} < 3")
        
        # Get property data from request
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        required_fields = ['propertyType', 'address', 'floorArea']
        for field in required_fields:
            if field not in data:
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        # Prepare property data for ML prediction
        property_data = {
            'propertyType': data['propertyType'],
            'address': data['address'],
            'floorArea': str(data['floorArea']),
            'level': data.get('level', 'Ground Floor'),
            'unit': data.get('unit', 'N/A')
        }
        
        # Run ML prediction
        print(f"🔍 Running ML prediction for: {property_data}")
        ml_result = run_ml_prediction(property_data)
        print(f"🔍 ML result: {ml_result}")
        
        if not ml_result['success']:
            print(f"❌ ML prediction failed: {ml_result['error']}")
            return jsonify({'error': ml_result['error']}), 500
        
        # Extract prediction data
        property_data_result = ml_result['property_data']
        comparison_data_result = ml_result['comparison_data']
        
        # Save prediction to database
        try:
            # Convert estimated sales price to numeric
            estimated_price = None
            if comparison_data_result.get('estimatedSalesPrice') and comparison_data_result['estimatedSalesPrice'] != 'N/A':
                price_str = comparison_data_result['estimatedSalesPrice'].replace('$', '').replace(',', '')
                if 'M' in price_str:
                    estimated_price = float(price_str.replace('M', '')) * 1000000
                elif 'k' in price_str:
                    estimated_price = float(price_str.replace('k', '')) * 1000
                else:
                    estimated_price = float(price_str)
            
            # Create prediction record
            prediction = PricePrediction(
                user_id=user_id,
                property_address=property_data['address'],
                property_type=property_data['propertyType'],
                size_sqft=float(property_data['floorArea']),
                predicted_price=estimated_price or 0,
                confidence_score=85.0,  # Default confidence
                search_parameters=json.dumps(property_data)
            )
            
            db.session.add(prediction)
            db.session.commit()
            
            print(f"✅ Prediction saved to database - ID: {prediction.id}, user_id: {user_id}")
            
            # Verify the count after saving
            if user.user_type and user.user_type.lower() == 'free':
                new_count = PricePrediction.query.filter_by(user_id=user_id).count()
                print(f"🔍 Updated prediction count for user {user_id}: {new_count}")
            
            # Add prediction ID to response
            property_data_result['prediction_id'] = prediction.id
            
        except Exception as e:
            print(f"❌ Error saving prediction to database: {e}")
            import traceback
            traceback.print_exc()
            # Continue without saving to database
        
        # Return the prediction results
        return jsonify({
            'success': True,
            'property_data': property_data_result,
            'comparison_data': comparison_data_result,
            'matched_address': ml_result.get('matched_address'),
            'message': 'Price prediction generated successfully'
        })
        
    except Exception as e:
        print(f"Error in predict_price endpoint: {e}")
        return jsonify({'error': 'Internal server error'}), 500

@app.route('/api/predict-price-test', methods=['POST'])
def predict_price_test():
    """Test endpoint for ML prediction without authentication"""
    try:
        # Get property data from request
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        required_fields = ['propertyType', 'address', 'floorArea']
        for field in required_fields:
            if field not in data:
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        # Prepare property data for ML prediction
        property_data = {
            'propertyType': data['propertyType'],
            'address': data['address'],
            'floorArea': str(data['floorArea']),
            'level': data.get('level', 'Ground Floor'),
            'unit': data.get('unit', 'N/A')
        }
        
        # Run ML prediction
        print(f"🔍 [TEST] Running ML prediction for: {property_data}")
        ml_result = run_ml_prediction(property_data)
        print(f"🔍 [TEST] ML result: {ml_result}")
        
        if not ml_result['success']:
            print(f"❌ [TEST] ML prediction failed: {ml_result['error']}")
            return jsonify({'error': ml_result['error']}), 500
        
        # Return the prediction results (without saving to database)
        return jsonify({
            'success': True,
            'property_data': ml_result['property_data'],
            'comparison_data': ml_result['comparison_data'],
            'matched_address': ml_result.get('matched_address'),
            'message': 'Test prediction generated successfully (not saved to database)'
        })
        
    except Exception as e:
        print(f"Error in predict_price_test endpoint: {e}")
        return jsonify({'error': 'Internal server error'}), 500

@app.route('/api/predictions/check-limit', methods=['GET'])
@require_auth
def check_prediction_limit():
    """Check if free user has reached prediction limit"""
    try:
        current_user = request.user
        user_id = current_user['user_id']
        
        user = User.query.get(user_id)
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        # Only check limit for free users
        if user.user_type and user.user_type.lower() == 'free':
            prediction_count = PricePrediction.query.filter_by(user_id=user_id).count()
            return jsonify({
                'is_free_user': True,
                'prediction_count': prediction_count,
                'limit': 3,
                'limit_reached': prediction_count >= 3,
                'remaining': max(0, 3 - prediction_count)
            }), 200
        else:
            return jsonify({
                'is_free_user': False,
                'prediction_count': 0,
                'limit': 0,
                'limit_reached': False,
                'remaining': -1  # Unlimited for premium/agent/admin
            }), 200
            
    except Exception as e:
        print(f"Error checking prediction limit: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Failed to check prediction limit'}), 500

@app.route('/api/predictions/user/<int:user_id>', methods=['GET'])
def get_user_predictions(user_id):
    """Get all predictions for a specific user"""
    try:
        # Verify user access
        token = request.headers.get('Authorization')
        if not token or not token.startswith('Bearer '):
            return jsonify({'error': 'Authorization token required'}), 401
        
        token = token.split(' ')[1]
        try:
            payload = jwt.decode(token, JWT_SECRET_KEY, algorithms=['HS256'])
            if payload['user_id'] != user_id:
                return jsonify({'error': 'Access denied'}), 403
        except jwt.ExpiredSignatureError:
            return jsonify({'error': 'Token expired'}), 401
        except jwt.InvalidTokenError:
            return jsonify({'error': 'Invalid token'}), 401
        
        # Get user's predictions
        predictions = PricePrediction.query.filter_by(user_id=user_id).order_by(PricePrediction.prediction_date.desc()).all()
        
        predictions_data = []
        for pred in predictions:
            predictions_data.append({
                'id': pred.id,
                'property_address': pred.property_address,
                'property_type': pred.property_type,
                'size_sqft': float(pred.size_sqft),
                'predicted_price': float(pred.predicted_price),
                'confidence_score': float(pred.confidence_score) if pred.confidence_score else None,
                'prediction_date': pred.prediction_date.isoformat(),
                'search_parameters': json.loads(pred.search_parameters) if pred.search_parameters else None
            })
        
        return jsonify({
            'success': True,
            'predictions': predictions_data,
            'count': len(predictions_data)
        })
        
    except Exception as e:
        print(f"Error in get_user_predictions endpoint: {e}")
        return jsonify({'error': 'Internal server error'}), 500

# ===================== Admin Regions Management Endpoints =====================

@app.route('/api/admin/regions', methods=['GET'])
@require_auth
def get_all_regions():
    """Get all regions - accessible to both admin and agents"""
    try:
        current_user = request.user
        
        # Get all regions
        regions = Region.query.order_by(Region.district).all()
        
        regions_data = []
        for region in regions:
            regions_data.append({
                'id': region.id,
                'district': region.district,
                'sector': region.sector,
                'location': region.location,
                'is_active': region.is_active,
                'created_at': region.created_at.isoformat() if region.created_at else None,
                'updated_at': region.updated_at.isoformat() if region.updated_at else None
            })
        
        return jsonify({'regions': regions_data}), 200
        
    except Exception as e:
        print(f"Error getting regions: {e}")
        return jsonify({'error': 'Failed to get regions'}), 500

@app.route('/api/admin/regions/<int:region_id>', methods=['GET'])
@require_auth
def get_region(region_id):
    """Get a specific region"""
    try:
        current_user = request.user
        
        # Verify user is an admin
        user = User.query.get(current_user['user_id'])
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Access denied'}), 403
        
        region = Region.query.get(region_id)
        if not region:
            return jsonify({'error': 'Region not found'}), 404
        
        return jsonify({
            'id': region.id,
            'district': region.district,
            'sector': region.sector,
            'location': region.location,
            'is_active': region.is_active
        }), 200
        
    except Exception as e:
        print(f"Error getting region: {e}")
        return jsonify({'error': 'Failed to get region'}), 500

@app.route('/api/admin/regions', methods=['POST'])
@require_auth
def create_region():
    """Create a new region"""
    try:
        current_user = request.user
        
        # Verify user is an admin
        user = User.query.get(current_user['user_id'])
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Access denied'}), 403
        
        data = request.get_json()
        district = data.get('district')
        sector = data.get('sector')
        location = data.get('location')
        
        if not district or not sector or not location:
            return jsonify({'error': 'Missing required fields'}), 400
        
        # Create new region
        new_region = Region(
            district=district,
            sector=sector,
            location=location,
            is_active=True
        )
        db.session.add(new_region)
        db.session.commit()
        
        return jsonify({
            'message': 'Region created successfully',
            'region': {
                'id': new_region.id,
                'district': new_region.district,
                'sector': new_region.sector,
                'location': new_region.location,
                'is_active': new_region.is_active
            }
        }), 201
        
    except Exception as e:
        print(f"Error creating region: {e}")
        db.session.rollback()
        return jsonify({'error': 'Failed to create region'}), 500

@app.route('/api/admin/regions/<int:region_id>', methods=['PUT'])
@require_auth
def update_region(region_id):
    """Update a region"""
    try:
        current_user = request.user
        
        # Verify user is an admin
        user = User.query.get(current_user['user_id'])
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Access denied'}), 403
        
        region = Region.query.get(region_id)
        if not region:
            return jsonify({'error': 'Region not found'}), 404
        
        data = request.get_json()
        
        if 'district' in data:
            region.district = data['district']
        if 'sector' in data:
            region.sector = data['sector']
        if 'location' in data:
            region.location = data['location']
        if 'is_active' in data:
            region.is_active = data['is_active']
        
        db.session.commit()
        
        return jsonify({
            'message': 'Region updated successfully',
            'region': {
                'id': region.id,
                'district': region.district,
                'sector': region.sector,
                'location': region.location,
                'is_active': region.is_active
            }
        }), 200
        
    except Exception as e:
        print(f"Error updating region: {e}")
        db.session.rollback()
        return jsonify({'error': 'Failed to update region'}), 500

@app.route('/api/admin/regions/<int:region_id>', methods=['DELETE'])
@require_auth
def delete_region(region_id):
    """Delete a region"""
    try:
        current_user = request.user
        
        # Verify user is an admin
        user = User.query.get(current_user['user_id'])
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Access denied'}), 403
        
        region = Region.query.get(region_id)
        if not region:
            return jsonify({'error': 'Region not found'}), 404
        
        # Check if any agents are assigned to this region
        region_value = f"District {region.district}"
        agent_regions_count = AgentRegion.query.filter_by(region_value=region.location).count()
        
        if agent_regions_count > 0:
            return jsonify({'error': f'Cannot delete region. {agent_regions_count} agent(s) assigned to this region.'}), 400
        
        db.session.delete(region)
        db.session.commit()
        
        return jsonify({'message': 'Region deleted successfully'}), 200
        
    except Exception as e:
        print(f"Error deleting region: {e}")
        db.session.rollback()
        return jsonify({'error': 'Failed to delete region'}), 500

@app.route('/api/property-card/export-excel', methods=['POST'])
@require_auth
def export_property_card_excel():
    """Export property card data to Excel format"""
    try:
        # Check if user is a free user - restrict download for free users
        current_user = request.user
        user_id = current_user['user_id']
        user = User.query.get(user_id)
        
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        # Block free users from downloading reports
        if user.user_type and user.user_type.lower() == 'free':
            return jsonify({
                'error': 'Report download is not available for free users. Please upgrade to Premium to download reports.',
                'upgrade_required': True
            }), 403
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        from flask import send_file
        
        data = request.json
        
        # Check if this is a comparison (multiple properties)
        is_comparison = data.get('isComparison', False)
        property1_data = data.get('property1', {})
        property2_data = data.get('property2', {})
        
        # Single property mode (backward compatible)
        if not is_comparison:
            property_data = data.get('property', {})
            comparison_data = data.get('comparisonData', {})
            ml_prediction_data = data.get('mlPredictionData', {})
            nearby_properties = data.get('nearbyProperties', [])
            region_agents = data.get('regionAgents', [])
            
            # Use ML prediction data if available, otherwise fallback to comparison data
            effective_comparison = ml_prediction_data.get('comparison_data', {}) if ml_prediction_data else comparison_data
        else:
            # Comparison mode
            property1_comparison_data = property1_data.get('comparisonData', {})
            property1_ml_data = property1_data.get('mlPredictionData', {})
            property1_effective_comparison = property1_ml_data.get('comparison_data', {}) if property1_ml_data else property1_comparison_data
            
            property2_comparison_data = property2_data.get('comparisonData', {})
            property2_ml_data = property2_data.get('mlPredictionData', {})
            property2_effective_comparison = property2_ml_data.get('comparison_data', {}) if property2_ml_data else property2_comparison_data
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Property Report" if not is_comparison else "Property Comparison Report"
        
        # Define styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title_font = Font(bold=True, size=14, color="000000")
        title_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        row = 1
        
        # Title
        if is_comparison:
            ws.merge_cells(f'A{row}:G{row}')
            ws[f'A{row}'] = "Property Comparison Report"
        else:
            ws.merge_cells(f'A{row}:E{row}')
            ws[f'A{row}'] = "Property Analysis Report"
        ws[f'A{row}'].font = Font(bold=True, size=16)
        ws[f'A{row}'].alignment = center_alignment
        row += 2
        
        # Handle comparison mode vs single property mode
        if is_comparison:
            # COMPARISON MODE: Side-by-side layout
            prop1 = property1_data.get('property', {})
            prop2 = property2_data.get('property', {})
            prop1_nearby = property1_data.get('nearbyProperties', [])
            prop2_nearby = property2_data.get('nearbyProperties', [])
            prop1_agents = property1_data.get('regionAgents', [])
            prop2_agents = property2_data.get('regionAgents', [])
            
            # Property Details Comparison Section
            ws[f'A{row}'] = "Property Details Comparison"
            ws[f'A{row}'].font = title_font
            ws[f'A{row}'].fill = title_fill
            ws.merge_cells(f'A{row}:G{row}')
            row += 1
            
            # Comparison headers
            ws[f'A{row}'] = "Attribute"
            ws[f'B{row}'] = "Property 1"
            ws[f'C{row}'] = "Property 2"
            for col in [1, 2, 3]:
                cell = ws.cell(row=row, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = center_alignment
            row += 1
            
            # Property details comparison
            comparison_details = [
                ['Address', prop1.get('address', 'N/A'), prop2.get('address', 'N/A')],
                ['Property Type', prop1.get('propertyType', 'N/A'), prop2.get('propertyType', 'N/A')],
                ['Floor Area', f"{prop1.get('floorArea', 'N/A')} sq ft", f"{prop2.get('floorArea', 'N/A')} sq ft"],
                ['Level', prop1.get('level', 'N/A'), prop2.get('level', 'N/A')],
                ['Unit', prop1.get('unit', 'N/A'), prop2.get('unit', 'N/A')]
            ]
            
            for detail in comparison_details:
                ws.cell(row=row, column=1, value=detail[0]).font = Font(bold=True)
                ws.cell(row=row, column=1).border = border
                ws.cell(row=row, column=2, value=detail[1]).border = border
                ws.cell(row=row, column=3, value=detail[2]).border = border
                row += 1
            
            row += 1
            
            # Price Estimates Comparison
            ws[f'A{row}'] = "Price Estimates Comparison"
            ws[f'A{row}'].font = title_font
            ws[f'A{row}'].fill = title_fill
            ws.merge_cells(f'A{row}:G{row}')
            row += 1
            
            ws[f'A{row}'] = "Attribute"
            ws[f'B{row}'] = "Property 1"
            ws[f'C{row}'] = "Property 2"
            for col in [1, 2, 3]:
                cell = ws.cell(row=row, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = center_alignment
            row += 1
            
            price_comparison = [
                ['Estimated Sales Price', 
                 property1_effective_comparison.get('estimatedSalesPrice', 'N/A'),
                 property2_effective_comparison.get('estimatedSalesPrice', 'N/A')],
                ['Estimated Rental Price',
                 property1_effective_comparison.get('estimatedRentalPrice', 'N/A'),
                 property2_effective_comparison.get('estimatedRentalPrice', 'N/A')]
            ]
            
            for price_detail in price_comparison:
                ws.cell(row=row, column=1, value=price_detail[0]).font = Font(bold=True)
                ws.cell(row=row, column=1).border = border
                ws.cell(row=row, column=2, value=price_detail[1]).border = border
                ws.cell(row=row, column=3, value=price_detail[2]).border = border
                row += 1
            
            row += 1
            
            # Market Trends Comparison
            ws[f'A{row}'] = "Market Trends Comparison"
            ws[f'A{row}'].font = title_font
            ws[f'A{row}'].fill = title_fill
            ws.merge_cells(f'A{row}:G{row}')
            row += 1
            
            ws[f'A{row}'] = "Attribute"
            ws[f'B{row}'] = "Property 1"
            ws[f'C{row}'] = "Property 2"
            for col in [1, 2, 3]:
                cell = ws.cell(row=row, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = center_alignment
            row += 1
            
            trend_comparison = [
                ['Market Trend',
                 property1_effective_comparison.get('marketTrend', 'N/A'),
                 property2_effective_comparison.get('marketTrend', 'N/A')],
                ['Period',
                 property1_effective_comparison.get('marketTrendPeriod', 'N/A'),
                 property2_effective_comparison.get('marketTrendPeriod', 'N/A')]
            ]
            
            for trend_detail in trend_comparison:
                ws.cell(row=row, column=1, value=trend_detail[0]).font = Font(bold=True)
                ws.cell(row=row, column=1).border = border
                ws.cell(row=row, column=2, value=trend_detail[1]).border = border
                ws.cell(row=row, column=3, value=trend_detail[2]).border = border
                row += 1
            
            row += 1
            
            # Price Statistics Comparison
            ws[f'A{row}'] = "Price Statistics Comparison"
            ws[f'A{row}'].font = title_font
            ws[f'A{row}'].fill = title_fill
            ws.merge_cells(f'A{row}:G{row}')
            row += 1
            
            ws[f'A{row}'] = "Attribute"
            ws[f'B{row}'] = "Property 1"
            ws[f'C{row}'] = "Property 2"
            for col in [1, 2, 3]:
                cell = ws.cell(row=row, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = center_alignment
            row += 1
            
            stats_comparison = [
                ['Median Sale Price',
                 property1_effective_comparison.get('medianSalePrice', 'N/A'),
                 property2_effective_comparison.get('medianSalePrice', 'N/A')],
                ['Highest Sold Price',
                 property1_effective_comparison.get('highestSoldPriceDescription', 'N/A')[:50] + '...' if len(property1_effective_comparison.get('highestSoldPriceDescription', 'N/A')) > 50 else property1_effective_comparison.get('highestSoldPriceDescription', 'N/A'),
                 property2_effective_comparison.get('highestSoldPriceDescription', 'N/A')[:50] + '...' if len(property2_effective_comparison.get('highestSoldPriceDescription', 'N/A')) > 50 else property2_effective_comparison.get('highestSoldPriceDescription', 'N/A')]
            ]
            
            for stats_detail in stats_comparison:
                ws.cell(row=row, column=1, value=stats_detail[0]).font = Font(bold=True)
                ws.cell(row=row, column=1).border = border
                ws.cell(row=row, column=2, value=stats_detail[1]).border = border
                ws.cell(row=row, column=3, value=stats_detail[2]).border = border
                row += 1
            
            row += 1
            
            # Property 1 Transactions Section
            prop1_transactions = property1_effective_comparison.get('historicalTransactions', [])
            if prop1_transactions:
                ws[f'A{row}'] = "Property 1 - Similar Properties Transactions"
                ws[f'A{row}'].font = title_font
                ws[f'A{row}'].fill = title_fill
                ws.merge_cells(f'A{row}:G{row}')
                row += 1
                
                headers = ['Address', 'Property Type', 'Floor Area', 'Date', 'Sales Price', 'Unit Price ($ psf)', 'Postal District']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = center_alignment
                row += 1
                
                for transaction in prop1_transactions[:5]:  # Limit to 5 for comparison
                    ws.cell(row=row, column=1, value=transaction.get('address', 'N/A')).border = border
                    ws.cell(row=row, column=2, value=transaction.get('propertyType', 'N/A')).border = border
                    ws.cell(row=row, column=3, value=transaction.get('floorArea', 'N/A')).border = border
                    ws.cell(row=row, column=4, value=transaction.get('date', 'N/A')).border = border
                    ws.cell(row=row, column=5, value=transaction.get('price', 'N/A')).border = border
                    ws.cell(row=row, column=6, value=transaction.get('unitPricePsf', transaction.get('unit_price_psf', 'N/A'))).border = border
                    ws.cell(row=row, column=7, value=transaction.get('postalDistrict', 'N/A')).border = border
                    row += 1
                
                row += 1
            
            # Property 2 Transactions Section
            prop2_transactions = property2_effective_comparison.get('historicalTransactions', [])
            if prop2_transactions:
                ws[f'A{row}'] = "Property 2 - Similar Properties Transactions"
                ws[f'A{row}'].font = title_font
                ws[f'A{row}'].fill = title_fill
                ws.merge_cells(f'A{row}:G{row}')
                row += 1
                
                headers = ['Address', 'Property Type', 'Floor Area', 'Date', 'Sales Price', 'Unit Price ($ psf)', 'Postal District']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = center_alignment
                row += 1
                
                for transaction in prop2_transactions[:5]:  # Limit to 5 for comparison
                    ws.cell(row=row, column=1, value=transaction.get('address', 'N/A')).border = border
                    ws.cell(row=row, column=2, value=transaction.get('propertyType', 'N/A')).border = border
                    ws.cell(row=row, column=3, value=transaction.get('floorArea', 'N/A')).border = border
                    ws.cell(row=row, column=4, value=transaction.get('date', 'N/A')).border = border
                    ws.cell(row=row, column=5, value=transaction.get('price', 'N/A')).border = border
                    ws.cell(row=row, column=6, value=transaction.get('unitPricePsf', transaction.get('unit_price_psf', 'N/A'))).border = border
                    ws.cell(row=row, column=7, value=transaction.get('postalDistrict', 'N/A')).border = border
                    row += 1
                
                row += 1
            
            # Update filename for comparison
            prop1_address = prop1.get('address', 'property1').replace('/', '-').replace('\\', '-').replace(':', '-')
            prop2_address = prop2.get('address', 'property2').replace('/', '-').replace('\\', '-').replace(':', '-')
            address_safe = f"{prop1_address}-vs-{prop2_address}"
            
        else:
            # SINGLE PROPERTY MODE (existing code)
            # Property Details Section
            ws[f'A{row}'] = "Property Details"
            ws[f'A{row}'].font = title_font
            ws[f'A{row}'].fill = title_fill
            ws.merge_cells(f'A{row}:E{row}')
            row += 1
            
            property_details = [
                ['Address', property_data.get('address', 'N/A')],
                ['Property Type', property_data.get('propertyType', 'N/A')],
                ['Floor Area', f"{property_data.get('floorArea', 'N/A')} sq ft"],
                ['Level', property_data.get('level', 'N/A')],
                ['Unit', property_data.get('unit', 'N/A')]
            ]
            
            for detail in property_details:
                ws[f'A{row}'] = detail[0]
                ws[f'B{row}'] = detail[1]
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'A{row}'].border = border
                ws[f'B{row}'].border = border
                row += 1
            
            row += 1
        
            # Estimated Prices Section
            ws[f'A{row}'] = "Price Estimates"
            ws[f'A{row}'].font = title_font
            ws[f'A{row}'].fill = title_fill
            ws.merge_cells(f'A{row}:E{row}')
            row += 1
            
            ws[f'A{row}'] = "Estimated Sales Price"
            ws[f'B{row}'] = effective_comparison.get('estimatedSalesPrice', 'N/A')
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1
            
            ws[f'A{row}'] = "Estimated Rental Price"
            ws[f'B{row}'] = effective_comparison.get('estimatedRentalPrice', 'N/A')
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1
            
            row += 1
            
            # Market Trends Section
            ws[f'A{row}'] = "Market Trends"
            ws[f'A{row}'].font = title_font
            ws[f'A{row}'].fill = title_fill
            ws.merge_cells(f'A{row}:E{row}')
            row += 1
            
            ws[f'A{row}'] = "Market Trend"
            ws[f'B{row}'] = effective_comparison.get('marketTrend', 'N/A')
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1
            
            ws[f'A{row}'] = "Period"
            ws[f'B{row}'] = effective_comparison.get('marketTrendPeriod', 'N/A')
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1
            
            row += 1
            
            # Price Statistics Section
            ws[f'A{row}'] = "Price Statistics"
            ws[f'A{row}'].font = title_font
            ws[f'A{row}'].fill = title_fill
            ws.merge_cells(f'A{row}:E{row}')
            row += 1
            
            ws[f'A{row}'] = "Median Sale Price"
            ws[f'B{row}'] = effective_comparison.get('medianSalePrice', 'N/A')
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1
            
            ws[f'A{row}'] = "Highest Sold Price"
            ws[f'B{row}'] = effective_comparison.get('highestSoldPriceDescription', 'N/A')
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            ws[f'B{row}'].alignment = Alignment(wrap_text=True)
            row += 1
            
            row += 1
            
            # Set address_safe for single property mode
            address_safe = property_data.get('address', 'property').replace('/', '-').replace('\\', '-').replace(':', '-')
        
        # Similar Properties Transactions Section (only for single property mode)
        if not is_comparison:
            historical_transactions = effective_comparison.get('historicalTransactions', [])
            if historical_transactions:
                ws[f'A{row}'] = "Similar Properties Transactions"
                ws[f'A{row}'].font = title_font
                ws[f'A{row}'].fill = title_fill
                ws.merge_cells(f'A{row}:G{row}')
                row += 1
                
                # Headers
                headers = ['Address', 'Property Type', 'Floor Area', 'Date', 'Sales Price', 'Unit Price ($ psf)', 'Postal District']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = center_alignment
                row += 1
                
                # Transaction rows
                for transaction in historical_transactions:
                    ws.cell(row=row, column=1, value=transaction.get('address', 'N/A')).border = border
                    ws.cell(row=row, column=2, value=transaction.get('propertyType', 'N/A')).border = border
                    ws.cell(row=row, column=3, value=transaction.get('floorArea', 'N/A')).border = border
                    ws.cell(row=row, column=4, value=transaction.get('date', 'N/A')).border = border
                    ws.cell(row=row, column=5, value=transaction.get('price', 'N/A')).border = border
                    ws.cell(row=row, column=6, value=transaction.get('unitPricePsf', transaction.get('unit_price_psf', 'N/A'))).border = border
                    ws.cell(row=row, column=7, value=transaction.get('postalDistrict', 'N/A')).border = border
                    row += 1
                
                row += 1
        
        # Nearby Properties Section (only for single property mode)
        if not is_comparison and nearby_properties:
            ws[f'A{row}'] = "Nearby Properties"
            ws[f'A{row}'].font = title_font
            ws[f'A{row}'].fill = title_fill
            ws.merge_cells(f'A{row}:E{row}')
            row += 1
            
            # Headers
            headers = ['Address', 'Property Type', 'Size', 'Price', 'Action']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = center_alignment
            row += 1
            
            # Property rows
            for prop in nearby_properties:
                size_str = f"{prop.get('size_sqft', prop.get('size', 'N/A'))} sq ft" if prop.get('size_sqft') else prop.get('size', 'N/A')
                price_str = 'N/A'
                if prop.get('asking_price'):
                    price = prop['asking_price']
                    if prop.get('price_type') == 'rental':
                        price_str = f"${(price / 1000):.0f}k/month"
                    elif price >= 1000000:
                        price_str = f"${(price / 1000000):.1f}M"
                    else:
                        price_str = f"${(price / 1000):.0f}k"
                
                ws.cell(row=row, column=1, value=prop.get('address', 'N/A')).border = border
                ws.cell(row=row, column=2, value=prop.get('property_type', 'N/A')).border = border
                ws.cell(row=row, column=3, value=size_str).border = border
                ws.cell(row=row, column=4, value=price_str).border = border
                ws.cell(row=row, column=5, value='View Details').border = border
                row += 1
            
            row += 1
        
        # Contact Agents Section (only for single property mode)
        if not is_comparison and region_agents:
            ws[f'A{row}'] = "Contact Agents"
            ws[f'A{row}'].font = title_font
            ws[f'A{row}'].fill = title_fill
            ws.merge_cells(f'A{row}:E{row}')
            row += 1
            
            # Headers
            headers = ['Name', 'Company', 'License', 'Phone', 'Email']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = center_alignment
            row += 1
            
            # Agent rows
            for agent in region_agents:
                ws.cell(row=row, column=1, value=agent.get('name', 'N/A')).border = border
                ws.cell(row=row, column=2, value=agent.get('company', 'N/A')).border = border
                ws.cell(row=row, column=3, value=agent.get('license', 'N/A')).border = border
                ws.cell(row=row, column=4, value=agent.get('phone', 'N/A')).border = border
                ws.cell(row=row, column=5, value=agent.get('email', 'N/A')).border = border
                row += 1
        
        # Auto-adjust column widths
        for col in range(1, 8):
            max_length = 0
            column_letter = get_column_letter(col)
            for cell in ws[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename (address_safe is already set above based on mode)
        filename = f"property-report-{address_safe}-{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"
        # Clean filename for safe download
        filename_safe = "".join(c for c in filename if c.isalnum() or c in ('-', '_', '.'))[:200]  # Limit length
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename_safe
        )
        
    except ImportError:
        return jsonify({'error': 'openpyxl library not installed. Please install it using: pip install openpyxl'}), 500
    except Exception as e:
        print(f"Error generating Excel report: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to generate Excel report: {str(e)}'}), 500

if __name__ == '__main__':
    with app.app_context():
        db.create_all()  # Create tables if they don't exist
        seed_regions()  # Seed regions data
        seed_subscription_plans()  # Seed initial data
        seed_features_section()  # Seed features section
    app.run(debug=True, host='0.0.0.0', port=5001)
